#!/usr/bin/env python3
"""
YGORipperUI - Yu-Gi-Oh Card Price Checker & Pack Ripper
A GUI application to check card prices and rip card packs with voice recognition
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import requests
import json
import os
from datetime import datetime
import threading
import time
import re
try:
    from fuzzywuzzy import fuzz, process
    FUZZYWUZZY_AVAILABLE = True
except ImportError:
    FUZZYWUZZY_AVAILABLE = False
    print("fuzzywuzzy not available - string matching features disabled")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("openpyxl not available - Excel export features disabled")
try:
    import speech_recognition as sr
    SPEECH_RECOGNITION_AVAILABLE = True
except ImportError:
    SPEECH_RECOGNITION_AVAILABLE = False
    print("SpeechRecognition not available - voice features disabled")
import logging
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    print("PIL not available - image features disabled")
import urllib.parse
import hashlib

# Performance timing decorator for debugging lag issues
def performance_timer(func_name):
    """Decorator to time function execution for performance debugging"""
    def decorator(func):
        def wrapper(*args, **kwargs):
            start_time = time.time()
            try:
                result = func(*args, **kwargs)
                elapsed = (time.time() - start_time) * 1000  # Convert to milliseconds
                if elapsed > 10:  # Only log operations taking more than 10ms
                    print(f"[PERFORMANCE] {func_name}: {elapsed:.1f}ms")
                return result
            except Exception as e:
                elapsed = (time.time() - start_time) * 1000
                print(f"[PERFORMANCE] {func_name}: {elapsed:.1f}ms (ERROR: {e})")
                raise
        return wrapper
    return decorator

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ygo_ripper_ui.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class SessionTrackerWindow:
    """Dedicated window for tracking pack session cards with high-performance virtual scrolling"""
    def __init__(self, parent, pack_session, image_manager):
        self.parent = parent
        self.pack_session = pack_session
        self.image_manager = image_manager
        self.card_widgets = []
        
        # Thread safety and update management
        self.ui_update_lock = threading.Lock()
        self.is_updating_display = False
        self.max_concurrent_images = 2  # Limit concurrent image loads (legacy - not used with sync loading)
        
        # Virtual scrolling for performance with large collections
        self.use_virtual_scrolling = True  # Re-enabled with fixed implementation
        self.virtual_window_size = 15  # Number of cards to render at once - reduced for better performance
        self.visible_start_index = 0
        self.visible_end_index = 0
        
        # Bulk preload management
        self.bulk_preload_started = False  # Prevent multiple preload runs
        
        # Create the session tracker window with dynamic sizing
        self.window = tk.Toplevel(parent)
        self.window.title("YGO Pack Session Tracker")
        
        # Get screen dimensions for dynamic sizing
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        # Calculate dynamic window sizes (70% of screen for normal, 80% for focus)
        self.normal_width = min(max(int(screen_width * 0.6), 800), 1400)  # 60% screen width, min 800, max 1400
        self.normal_height = min(max(int(screen_height * 0.6), 600), 1000)  # 60% screen height, min 600, max 1000
        self.focus_width = min(max(int(screen_width * 0.75), 1000), 1600)  # 75% screen width, min 1000, max 1600
        self.focus_height = min(max(int(screen_height * 0.75), 700), 1200)  # 75% screen height, min 700, max 1200
        
        # Set initial window size (normal mode)
        self.window.geometry(f"{self.normal_width}x{self.normal_height}")
        self.window.resizable(True, True)
        
        # Bind window close to cleanup
        self.window.protocol("WM_DELETE_WINDOW", self.cleanup_and_close)
        
        # Initialize dynamic layout values early
        self.dynamic_scroll_padding = 150  # Default values
        self.dynamic_card_spacing = 8
        self.dynamic_column_width = 280
        
        # Bind window resize events for dynamic layout updates
        self.window.bind('<Configure>', self.on_window_resize)
        self.last_window_width = self.normal_width
        self.last_window_height = self.normal_height
        
        # Initialize display settings
        self.display_settings = {
            'show_images': True,  # Re-enabled with safe synchronous loading
            'show_card_name': True,
            'show_rarity': True,
            'show_art_variant': True,
            'show_tcg_price': True,
            'show_tcg_market_price': True,
            'show_set_info': False,
            'show_timestamps': False
        }
        
        self.focus_mode = False
        self.setup_ui()
        
        # Update dynamic layout after UI setup
        self.window.after(100, self.update_dynamic_layout)
        
        # Use virtual scrolling for collections with more than 25 cards
        card_count = len(pack_session.cards)
        if card_count > 25:
            self.use_virtual_scrolling = True
            print(f"[PERFORMANCE] Using virtual scrolling for {card_count} cards")
            self.setup_virtual_scrolling()
        else:
            self.use_virtual_scrolling = False
            print(f"[PERFORMANCE] Using standard display for {card_count} cards")
            self.safe_update_cards_display()
        
        # Start bulk image preloading for better focus mode performance
        # Use longer delay to prevent immediate startup lag
        self.window.after(5000, self.bulk_preload_images_for_session)  # Start after 5 seconds delay
        
    def cleanup_and_close(self):
        """Cleanup resources and close window safely"""
        print("[SESSION TRACKER] Cleaning up resources before closing...")
        
        # Clear widget references
        self.card_widgets = []
        
        # Clear any image cache references
        if hasattr(self.image_manager, 'clear_memory_cache'):
            self.image_manager.clear_memory_cache()
        
        # Destroy window
        try:
            self.window.destroy()
        except:
            pass
    
    def on_window_resize(self, event):
        """Handle window resize events to update layout dynamically"""
        try:
            # Only handle window resize events, not widget resize events
            if event.widget != self.window:
                return
            
            current_width = self.window.winfo_width()
            current_height = self.window.winfo_height()
            
            # Check if this is a significant resize (more than 50 pixels)
            width_change = abs(current_width - self.last_window_width)
            height_change = abs(current_height - self.last_window_height)
            
            if width_change > 50 or height_change > 50:
                print(f"[RESIZE] Window resized to {current_width}x{current_height}")
                
                # Update stored dimensions
                self.last_window_width = current_width
                self.last_window_height = current_height
                
                # Update dynamic layout calculations
                self.update_dynamic_layout()
                
        except Exception as e:
            print(f"[RESIZE] Error in window resize handler: {e}")
    
    def update_dynamic_layout(self):
        """Update layout calculations based on current window size"""
        try:
            # Get current window dimensions
            current_width = self.window.winfo_width()
            current_height = self.window.winfo_height()
            
            # Calculate dynamic values based on window size
            self.dynamic_scroll_padding = max(int(current_height * 0.1), 50)  # 10% of height, min 50px
            self.dynamic_card_spacing = max(int(current_width * 0.005), 4)   # 0.5% of width, min 4px
            
            # Update focus mode column calculations
            if self.focus_mode:
                available_width = current_width - 40  # Account for padding/scrollbar
                self.dynamic_column_width = max(int(available_width / 4), 200)  # Min 200px per column
            else:
                self.dynamic_column_width = current_width - 40
            
            # Update scroll region with new padding
            self.update_scroll_region_dynamic()
            
            print(f"[DYNAMIC] Updated layout: padding={self.dynamic_scroll_padding}, "
                  f"spacing={self.dynamic_card_spacing}, col_width={self.dynamic_column_width}")
            
        except Exception as e:
            print(f"[DYNAMIC] Error updating dynamic layout: {e}")
    
    def cleanup_and_close(self):
        """Cleanup resources and close window safely"""
        
    def setup_ui(self):
        """Setup the session tracker UI"""
        # Main frame
        main_frame = ttk.Frame(self.window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Header with session info and controls
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Session info
        set_name = self.pack_session.current_set.get('set_name', 'Unknown Set') if self.pack_session.current_set else 'No Set'
        session_info = ttk.Label(header_frame, text=f"Pack Session: {set_name}", font=("Arial", 14, "bold"))
        session_info.pack(side=tk.LEFT)
        
        # Controls frame
        controls_frame = ttk.Frame(header_frame)
        controls_frame.pack(side=tk.RIGHT)
        
        # Focus mode toggle
        focus_btn = ttk.Button(controls_frame, text="Focus Mode", command=self.toggle_focus_mode)
        focus_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        # Display settings button
        settings_btn = ttk.Button(controls_frame, text="⚙️ Display", command=self.show_display_settings)
        settings_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        # Excel export button
        export_btn = ttk.Button(controls_frame, text="📊 Export Excel", command=self.export_session_to_excel)
        export_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        # Card count
        self.card_count_label = ttk.Label(controls_frame, text=f"Cards: {len(self.pack_session.cards)}")
        self.card_count_label.pack(side=tk.LEFT, padx=(5, 0))
        
        # Create scrollable area for cards
        self.create_scrollable_area(main_frame)
        
    def create_scrollable_area(self, parent):
        """Create robust scrollable area for card display using standard approach"""
        # Create main container
        container_frame = ttk.Frame(parent)
        container_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create canvas and scrollbar
        self.cards_canvas = tk.Canvas(container_frame, bg="white", highlightthickness=0)
        self.v_scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=self.cards_canvas.yview)
        
        # Create scrollable frame
        self.cards_scrollable_frame = ttk.Frame(self.cards_canvas)
        
        # Configure scrolling with simplified, robust approach
        def on_frame_configure(event):
            """Update scroll region when frame content changes with dynamic padding"""
            try:
                # Get the required canvas size
                bbox = self.cards_canvas.bbox("all")
                if bbox:
                    # Use dynamic padding based on window size
                    padding = getattr(self, 'dynamic_scroll_padding', 150)
                    x1, y1, x2, y2 = bbox
                    self.cards_canvas.configure(scrollregion=(x1, y1, x2, y2 + padding))
                else:
                    # Fallback calculation based on frame size with dynamic padding
                    self.cards_canvas.update_idletasks()
                    frame_height = self.cards_scrollable_frame.winfo_reqheight()
                    padding = getattr(self, 'dynamic_scroll_padding', 150)
                    self.cards_canvas.configure(scrollregion=(0, 0, 0, frame_height + padding))
                    
                print(f"[SCROLL DEBUG] Updated scrollregion with {padding}px padding: {self.cards_canvas.cget('scrollregion')}")
            except Exception as e:
                print(f"[SCROLL DEBUG] Error in frame configure: {e}")
        
        # Bind frame configuration changes
        self.cards_scrollable_frame.bind("<Configure>", on_frame_configure)
        
        # Create canvas window
        self.canvas_window = self.cards_canvas.create_window((0, 0), window=self.cards_scrollable_frame, anchor="nw")
        
        # Configure canvas scrolling
        self.cards_canvas.configure(yscrollcommand=self.v_scrollbar.set)
        
        # Pack elements
        self.cards_canvas.pack(side="left", fill="both", expand=True)
        self.v_scrollbar.pack(side="right", fill="y")
        
        # Configure canvas window width to match canvas
        def configure_canvas_width(event):
            """Make scrollable frame width match canvas width"""
            canvas_width = event.width
            self.cards_canvas.itemconfig(self.canvas_window, width=canvas_width)
        
        self.cards_canvas.bind('<Configure>', configure_canvas_width)
        
        # Enhanced mouse wheel scrolling with improved responsiveness
        def on_mousewheel(event):
            """Handle mouse wheel scrolling across platforms with improved sensitivity"""
            try:
                # Calculate scroll amount based on platform with improved sensitivity
                if hasattr(event, 'delta') and event.delta:
                    # Windows and MacOS - increase sensitivity
                    scroll_amount = int(-1 * (event.delta / 60))  # Changed from 120 to 60 for more sensitive scrolling
                elif hasattr(event, 'num'):
                    # Linux - increase scroll amount
                    if event.num == 4:
                        scroll_amount = -5  # Scroll up - increased from -3 to -5
                    elif event.num == 5:
                        scroll_amount = 5   # Scroll down - increased from 3 to 5
                    else:
                        return
                else:
                    return
                
                # Apply scrolling with bounds checking
                try:
                    self.cards_canvas.yview_scroll(scroll_amount, "units")
                    print(f"[SCROLL DEBUG] Mouse wheel scrolled: {scroll_amount} units")
                except tk.TclError:
                    # Fallback to page scrolling if units fail
                    page_amount = 1 if scroll_amount > 0 else -1
                    self.cards_canvas.yview_scroll(page_amount, "pages")
                    print(f"[SCROLL DEBUG] Mouse wheel page scroll: {page_amount}")
                    
            except Exception as e:
                print(f"[SCROLL DEBUG] Mouse wheel error: {e}")
        
        # Enhanced mouse wheel binding with additional events
        def bind_mousewheel(widget):
            """Bind mouse wheel events to widget with comprehensive coverage"""
            # Standard mouse wheel events
            widget.bind("<MouseWheel>", on_mousewheel)     # Windows/Mac
            widget.bind("<Button-4>", on_mousewheel)       # Linux scroll up
            widget.bind("<Button-5>", on_mousewheel)       # Linux scroll down
            
            # Additional wheel events for better compatibility
            widget.bind("<Shift-MouseWheel>", on_mousewheel)  # Horizontal scroll as vertical
            widget.bind("<Control-MouseWheel>", on_mousewheel) # Ctrl+wheel
        
        # Bind to multiple widgets for maximum coverage with focus management
        widgets_to_bind = [
            self.cards_canvas,
            self.cards_scrollable_frame,
            container_frame,
            self.window
        ]
        
        for widget in widgets_to_bind:
            bind_mousewheel(widget)
            # Also bind enter/leave events for better focus management
            widget.bind("<Enter>", lambda e: self.cards_canvas.focus_set())
        
        
        # Focus management for scroll events
        def set_focus_for_scroll(event):
            """Set focus to canvas for scrolling"""
            self.cards_canvas.focus_set()
        
        # Set focus when mouse enters any scroll area
        self.cards_canvas.bind("<Enter>", set_focus_for_scroll)
        container_frame.bind("<Enter>", set_focus_for_scroll)
        
        # Set initial focus
        self.window.after(100, lambda: self.cards_canvas.focus_set())
        
        print("[SCROLL DEBUG] Scrollable area created with enhanced mouse wheel support")
        
    def toggle_focus_mode(self):
        """Toggle between normal and focus mode with virtual scrolling support"""
        new_focus_mode = not self.focus_mode
        
        # Skip if mode doesn't change
        if new_focus_mode == self.focus_mode:
            print("[FOCUS MODE] Mode already matches target, skipping")
            return
            
        self.focus_mode = new_focus_mode
        
        # Adjust window size for mode using dynamic calculations
        if self.focus_mode:
            self.window.geometry(f"{self.focus_width}x{self.focus_height}")
            print(f"[FOCUS MODE] Enabled - Window enlarged to {self.focus_width}x{self.focus_height}, grid view activated")
        else:
            self.window.geometry(f"{self.normal_width}x{self.normal_height}")
            print(f"[FOCUS MODE] Disabled - Normal view restored to {self.normal_width}x{self.normal_height}")
        
        # Initialize dynamic layout values
        self.update_dynamic_layout()
        
        # Clear all existing widgets cleanly
        print("[FOCUS MODE] Clearing existing widgets...")
        self.clear_all_widgets()
        
        # Rebuild display for new mode with virtual scrolling awareness
        print(f"[FOCUS MODE] Rebuilding display for {'focus' if self.focus_mode else 'normal'} mode")
        if self.use_virtual_scrolling:
            # Reset virtual scrolling for new mode
            self.visible_start_index = 0
            self.visible_end_index = min(self.virtual_window_size, len(self.pack_session.cards))
            self.update_virtual_display_simple()
        else:
            self.rebuild_display_for_mode()
    
    @performance_timer("clear_all_widgets")
    def clear_all_widgets(self):
        """Clear all card widgets cleanly"""
        try:
            print(f"[CLEAR WIDGETS] Clearing {len(self.card_widgets)} widgets")
            
            # Destroy all widgets
            for widget in self.card_widgets:
                try:
                    if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                        widget.destroy()
                except Exception as e:
                    print(f"[CLEAR WIDGETS] Error destroying widget: {e}")
            
            # Clear the list
            self.card_widgets = []
            
            # Force update of scrollable frame
            if hasattr(self, 'cards_scrollable_frame') and self.cards_scrollable_frame:
                self.cards_scrollable_frame.update_idletasks()
                
            print("[CLEAR WIDGETS] All widgets cleared")
            
        except Exception as e:
            print(f"[CLEAR WIDGETS] Error clearing widgets: {e}")
            import traceback
            traceback.print_exc()
    
    @performance_timer("rebuild_display_for_mode")
    def rebuild_display_for_mode(self):
        """Rebuild the entire display for current mode"""
        try:
            print(f"[REBUILD] Rebuilding display for {'focus' if self.focus_mode else 'normal'} mode")
            print(f"[REBUILD] Cards to display: {len(self.pack_session.cards)}")
            
            # Create widgets for all cards
            for i, card in enumerate(self.pack_session.cards):
                try:
                    if self.focus_mode:
                        self.create_focus_mode_widget_simple(card, i)
                    else:
                        self.create_normal_mode_widget_simple(card, i)
                except Exception as e:
                    print(f"[REBUILD] Error creating widget {i}: {e}")
                    continue
            
            # Update the scroll region
            self.update_scroll_region_simple()
            
            # Auto-scroll to bottom to show newest cards
            self.window.after(100, self.auto_scroll_to_bottom)
            
            print(f"[REBUILD] Rebuild complete. Created {len(self.card_widgets)} widgets")
            
        except Exception as e:
            print(f"[REBUILD] Error rebuilding display: {e}")
            import traceback
            traceback.print_exc()
    
    def create_focus_mode_widget_simple(self, card, index):
        """Create focus mode widget with simplified, stable approach"""
        try:
            cards_per_row = 4
            row = index // cards_per_row
            col = index % cards_per_row
            
            # Configure grid columns for focus mode with dynamic sizing
            for c in range(cards_per_row):
                min_col_width = getattr(self, 'dynamic_column_width', 280)
                self.cards_scrollable_frame.columnconfigure(c, weight=1, minsize=min_col_width)
            
            # Create main card frame with dynamic spacing
            spacing = getattr(self, 'dynamic_card_spacing', 8)
            card_frame = ttk.Frame(self.cards_scrollable_frame, padding=spacing, relief="ridge")
            card_frame.grid(row=row, column=col, sticky=(tk.W, tk.E, tk.N), pady=spacing//2, padx=spacing//2)
            self.card_widgets.append(card_frame)
            
            # Configure internal grid
            card_frame.columnconfigure(0, weight=0)  # Image column
            card_frame.columnconfigure(1, weight=1)  # Content column
            
            # Add image if enabled
            if self.display_settings.get('show_images', True) and PIL_AVAILABLE:
                image_label = ttk.Label(card_frame, text="🖼️", font=("Arial", 8), justify="center")
                image_label.grid(row=0, column=0, rowspan=5, padx=(0, 8), pady=4, sticky="n")
                
                # Load image
                self.load_card_image_simple(card, image_label, focus_mode=True)
            
            # Content frame
            content_frame = ttk.Frame(card_frame)
            content_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N))
            content_frame.columnconfigure(0, weight=1)
            
            # Card name (truncated for focus mode)
            card_name = card.get('card_name', 'N/A')
            if len(card_name) > 25:
                card_name = card_name[:22] + "..."
            
            name_label = ttk.Label(content_frame, text=f"{index+1}. {card_name}", 
                                 font=("Arial", 10, "bold"), foreground="#2E86AB")
            name_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=1)
            
            # Rarity (truncated)
            rarity = card.get('card_rarity', 'N/A')
            if len(rarity) > 20:
                rarity = rarity[:17] + "..."
            rarity_label = ttk.Label(content_frame, text=f"💎 {rarity}", 
                                   font=("Arial", 9), foreground="#666")
            rarity_label.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=1)
            
            # Price info (compact)
            price_info = self.get_compact_price_info(card)
            price_label = ttk.Label(content_frame, text=f"💰 {price_info}", 
                                  font=("Arial", 9), foreground="#0066CC")
            price_label.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=1)
            
            # Quantity controls
            self.add_quantity_controls_focus(content_frame, card, index)
            
        except Exception as e:
            print(f"[FOCUS WIDGET] Error creating widget {index}: {e}")
            import traceback
            traceback.print_exc()
    
    def create_normal_mode_widget_simple(self, card, index):
        """Create normal mode widget with simplified, stable approach"""
        try:
            # Configure single column
            self.cards_scrollable_frame.columnconfigure(0, weight=1)
            
            # Create main card frame with dynamic spacing
            spacing = getattr(self, 'dynamic_card_spacing', 12)
            card_frame = ttk.Frame(self.cards_scrollable_frame, padding=spacing, relief="ridge")
            card_frame.grid(row=index, column=0, sticky=(tk.W, tk.E), pady=spacing//2, padx=spacing)
            self.card_widgets.append(card_frame)
            
            # Configure internal grid
            card_frame.columnconfigure(1, weight=1)
            
            current_row = 0
            
            # Add image if enabled
            if self.display_settings.get('show_images', True) and PIL_AVAILABLE:
                image_label = ttk.Label(card_frame, text="🖼️\nLoading...", 
                                      font=("Arial", 9), justify="center")
                image_label.grid(row=0, column=0, rowspan=7, padx=(0, 15), pady=5, sticky="n")
                
                # Load image
                self.load_card_image_simple(card, image_label, focus_mode=False)
            
            # Card details
            if self.display_settings.get('show_card_name', True):
                name_label = ttk.Label(card_frame, text=f"📋 {card.get('card_name', 'N/A')}", 
                                     font=("Arial", 12, "bold"))
                name_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            if self.display_settings.get('show_rarity', True):
                rarity_label = ttk.Label(card_frame, text=f"💎 Rarity: {card.get('card_rarity', 'N/A')}", 
                                       font=("Arial", 10))
                rarity_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            if self.display_settings.get('show_art_variant', True):
                variant_label = ttk.Label(card_frame, text=f"🎨 Art Variant: {card.get('art_variant', 'None')}", 
                                        font=("Arial", 10))
                variant_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            # Price information
            current_row = self.add_price_labels_simple(card_frame, card, current_row)
            
            # Set information if enabled
            if self.display_settings.get('show_set_info', False):
                set_name = card.get('set_name', 'N/A')
                set_code = card.get('set_code', 'N/A')
                set_label = ttk.Label(card_frame, text=f"📚 Set: {set_name} ({set_code})", 
                                    font=("Arial", 10))
                set_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            # Timestamp if enabled
            if self.display_settings.get('show_timestamps', False):
                timestamp = card.get('timestamp', datetime.now().strftime("%H:%M:%S"))
                time_label = ttk.Label(card_frame, text=f"🕒 Added: {timestamp}", 
                                     font=("Arial", 10))
                time_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            # Quantity controls
            self.add_quantity_controls_normal(card_frame, card, index, current_row)
            
        except Exception as e:
            print(f"[NORMAL WIDGET] Error creating widget {index}: {e}")
            import traceback
            traceback.print_exc()
    
    def load_card_image_simple(self, card, image_label, focus_mode):
        """Load card image with simplified, synchronous approach"""
        try:
            # Get image URL
            card_images = card.get('card_images', [])
            if not card_images:
                image_label.configure(text="🃏\nNo Image", font=("Arial", 8))
                return
            
            image_url = card_images[0].get('image_url')
            if not image_url:
                image_label.configure(text="🃏\nNo URL", font=("Arial", 8))
                return
            
            card_id = card.get('id', 'unknown')
            
            # Use smaller sizes in display for better performance
            if focus_mode:
                size = (60, 90)   # Reduced focus mode size
            else:
                size = (100, 145) # Reduced normal mode size
            
            # Load image
            photo = self.image_manager.load_image_for_display(card_id, image_url, size)
            
            if photo:
                image_label.configure(image=photo, text="")
                image_label.image = photo  # Keep reference
            else:
                image_label.configure(text="❌\nError", font=("Arial", 8))
                
        except Exception as e:
            print(f"[IMAGE LOAD] Error loading image: {e}")
            image_label.configure(text="❌\nError", font=("Arial", 8))
    
    def add_quantity_controls_focus(self, parent, card, index):
        """Add quantity controls for focus mode"""
        try:
            qty_frame = ttk.Frame(parent)
            qty_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=2)
            
            current_qty = card.get('quantity', 1)
            card_id = card.get('id', 'unknown')
            
            # Create stable quantity control functions
            def decrease_qty():
                self.update_card_quantity_by_id(card_id, -1)
            
            def increase_qty():
                self.update_card_quantity_by_id(card_id, 1)
            
            def remove_card():
                self.remove_card_by_id(card_id)
            
            # Layout controls
            ttk.Button(qty_frame, text="-", width=2, command=decrease_qty).pack(side=tk.LEFT)
            ttk.Label(qty_frame, text=str(current_qty), font=("Arial", 9, "bold"), width=2).pack(side=tk.LEFT, padx=2)
            ttk.Button(qty_frame, text="+", width=2, command=increase_qty).pack(side=tk.LEFT)
            ttk.Button(qty_frame, text="🗑️", width=3, command=remove_card).pack(side=tk.RIGHT)
            
        except Exception as e:
            print(f"[QTY CONTROLS] Error adding focus controls: {e}")
    
    def add_quantity_controls_normal(self, parent, card, index, row):
        """Add quantity controls for normal mode"""
        try:
            qty_frame = ttk.Frame(parent)
            qty_frame.grid(row=row, column=1, sticky=(tk.W), pady=4)
            
            current_qty = card.get('quantity', 1)
            card_id = card.get('id', 'unknown')
            
            # Create stable quantity control functions
            def decrease_qty():
                self.update_card_quantity_by_id(card_id, -1)
            
            def increase_qty():
                self.update_card_quantity_by_id(card_id, 1)
            
            def remove_card():
                self.remove_card_by_id(card_id)
            
            # Layout controls
            ttk.Label(qty_frame, text="📦 Quantity:", font=("Arial", 10)).pack(side=tk.LEFT)
            ttk.Button(qty_frame, text="-", width=3, command=decrease_qty).pack(side=tk.LEFT, padx=(5, 2))
            ttk.Label(qty_frame, text=str(current_qty), font=("Arial", 10, "bold"), width=3).pack(side=tk.LEFT, padx=2)
            ttk.Button(qty_frame, text="+", width=3, command=increase_qty).pack(side=tk.LEFT, padx=(2, 5))
            ttk.Button(qty_frame, text="🗑️ Remove", command=remove_card).pack(side=tk.LEFT, padx=(10, 0))
            
        except Exception as e:
            print(f"[QTY CONTROLS] Error adding normal controls: {e}")
    
    def add_price_labels_simple(self, parent, card, start_row):
        """Add price labels with simplified approach"""
        current_row = start_row
        
        # TCG Low price
        if self.display_settings.get('show_tcg_price', True):
            tcg_price = card.get('tcg_price')
            if tcg_price and tcg_price not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
                price_text = f"💰 TCG Low: ${tcg_price}"
            else:
                price_text = "💰 TCG Low: N/A"
            
            price_label = ttk.Label(parent, text=price_text, font=("Arial", 10))
            price_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            current_row += 1
        
        # TCG Market price
        if self.display_settings.get('show_tcg_market_price', True):
            tcg_market_price = card.get('tcg_market_price')
            if tcg_market_price and tcg_market_price not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
                price_text = f"📈 TCG Market: ${tcg_market_price}"
            else:
                price_text = "📈 TCG Market: N/A"
            
            price_label = ttk.Label(parent, text=price_text, font=("Arial", 10))
            price_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            current_row += 1
            
        return current_row
    
    def update_card_quantity_by_id(self, card_id, change):
        """Update card quantity by card ID with stable reference"""
        try:
            for i, card in enumerate(self.pack_session.cards):
                if card.get('id') == card_id:
                    current_qty = card.get('quantity', 1)
                    new_qty = max(1, current_qty + change)
                    card['quantity'] = new_qty
                    print(f"[QTY UPDATE] Updated card {card_id} quantity to {new_qty}")
                    
                    # Update the display for this card only
                    self.update_single_card_display(i)
                    break
        except Exception as e:
            print(f"[QTY UPDATE] Error updating quantity: {e}")
    
    def remove_card_by_id(self, card_id):
        """Remove card by card ID with efficient handling for both virtual and regular scrolling"""
        try:
            for i, card in enumerate(self.pack_session.cards):
                if card.get('id') == card_id:
                    del self.pack_session.cards[i]
                    print(f"[CARD REMOVE] Removed card {card_id}")
                    
                    # Handle removal based on scrolling mode
                    if self.use_virtual_scrolling:
                        self.handle_virtual_card_removal(i)
                    else:
                        # Use efficient incremental update for regular scrolling
                        self.remove_card_widget_efficiently(i)
                    break
        except Exception as e:
            print(f"[CARD REMOVE] Error removing card: {e}")
    
    def handle_virtual_card_removal(self, removed_index):
        """Handle card removal in virtual scrolling mode"""
        try:
            print(f"[VIRTUAL REMOVE] Handling removal of card at index {removed_index}")
            
            # Update virtual scroll region for new total
            self.setup_virtual_scroll_region()
            
            # If removed card was in visible range, update display
            if self.visible_start_index <= removed_index < self.visible_end_index:
                # Adjust visible range
                if self.visible_end_index > len(self.pack_session.cards):
                    self.visible_end_index = len(self.pack_session.cards)
                
                # Update virtual display to reflect removal
                self.update_virtual_display(self.visible_start_index, self.visible_end_index)
            
            # Update card count
            if hasattr(self, 'card_count_label') and self.card_count_label:
                self.card_count_label.config(text=f"Cards: {len(self.pack_session.cards)}")
            
            print(f"[VIRTUAL REMOVE] Virtual removal complete, {len(self.pack_session.cards)} cards remaining")
            
        except Exception as e:
            print(f"[VIRTUAL REMOVE] Error in virtual removal: {e}")
            # Fallback: refresh virtual display
            self.update_virtual_display(self.visible_start_index, 
                                       min(self.visible_end_index, len(self.pack_session.cards)))
    
    def remove_card_widget_efficiently(self, removed_index):
        """Efficiently remove a card widget and update display without full rebuild"""
        try:
            print(f"[EFFICIENT REMOVE] Removing widget at index {removed_index}")
            
            # Remove the specific widget
            if removed_index < len(self.card_widgets):
                widget_to_remove = self.card_widgets[removed_index]
                if hasattr(widget_to_remove, 'winfo_exists') and widget_to_remove.winfo_exists():
                    widget_to_remove.destroy()
                
                # Remove from widget list
                del self.card_widgets[removed_index]
            
            # Update grid positions for remaining widgets efficiently
            self.update_grid_positions_after_removal(removed_index)
            
            # Update card count
            if hasattr(self, 'card_count_label') and self.card_count_label:
                self.card_count_label.config(text=f"Cards: {len(self.pack_session.cards)}")
            
            # Update scroll region
            self.update_scroll_region_simple()
            
            print(f"[EFFICIENT REMOVE] Efficiently removed card, {len(self.card_widgets)} widgets remaining")
            
        except Exception as e:
            print(f"[EFFICIENT REMOVE] Error in efficient removal: {e}")
            # Fallback to full rebuild if needed
            self.clear_all_widgets()
            self.rebuild_display_for_mode()
    
    def update_grid_positions_after_removal(self, removed_index):
        """Update grid positions for widgets after a removal without recreating them"""
        try:
            cards_per_row = 4 if self.focus_mode else 1
            
            # Only update widgets that need new positions (those after the removed index)
            for i in range(removed_index, len(self.card_widgets)):
                widget = self.card_widgets[i]
                
                if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                    # Calculate new grid position
                    if self.focus_mode:
                        new_row = i // cards_per_row
                        new_col = i % cards_per_row
                    else:
                        new_row = i
                        new_col = 0
                    
                    # Update grid position
                    widget.grid_configure(row=new_row, column=new_col)
                    
                    # Update any index-based labels within the widget
                    self.update_widget_index_labels(widget, i)
            
            print(f"[GRID UPDATE] Updated positions for {len(self.card_widgets) - removed_index} widgets")
            
        except Exception as e:
            print(f"[GRID UPDATE] Error updating grid positions: {e}")
    
    def update_widget_index_labels(self, widget, new_index):
        """Update index-based labels within a widget (like "1. Card Name")"""
        try:
            # Recursively find and update labels with index numbers
            self.update_index_labels_recursive(widget, new_index)
        except Exception as e:
            print(f"[INDEX UPDATE] Error updating index labels: {e}")
    
    def update_index_labels_recursive(self, widget, new_index):
        """Recursively find and update index-based labels"""
        try:
            # Check if this is a label with index
            if hasattr(widget, 'cget') and hasattr(widget, 'configure'):
                try:
                    text = widget.cget('text')
                    if text and '. ' in text:
                        # Check if it starts with a number followed by '. '
                        parts = text.split('. ', 1)
                        if len(parts) == 2 and parts[0].isdigit():
                            # Update the index
                            new_text = f"{new_index + 1}. {parts[1]}"
                            widget.configure(text=new_text)
                            return
                except:
                    pass
            
            # Check children recursively
            try:
                for child in widget.winfo_children():
                    self.update_index_labels_recursive(child, new_index)
            except:
                pass
                
        except Exception as e:
            print(f"[INDEX LABELS] Error updating index labels: {e}")
    
    def update_single_card_display(self, card_index):
        """Update display for a single card (quantity change)"""
        try:
            if 0 <= card_index < len(self.card_widgets):
                widget = self.card_widgets[card_index]
                card = self.pack_session.cards[card_index]
                
                # Find and update quantity labels
                self.update_quantity_display_recursive(widget, card.get('quantity', 1))
                
        except Exception as e:
            print(f"[SINGLE UPDATE] Error updating card {card_index}: {e}")
    
    def update_quantity_display_recursive(self, widget, new_quantity):
        """Recursively find and update quantity displays"""
        try:
            # Check if this widget is a quantity label
            if hasattr(widget, 'cget') and hasattr(widget, 'configure'):
                try:
                    text = widget.cget('text')
                    if text and text.isdigit():
                        widget.configure(text=str(new_quantity))
                        return
                except:
                    pass
            
            # Check children recursively
            try:
                for child in widget.winfo_children():
                    self.update_quantity_display_recursive(child, new_quantity)
            except:
                pass
                
        except Exception as e:
            print(f"[QTY DISPLAY] Error updating quantity display: {e}")
    
    def update_scroll_region_dynamic(self):
        """Update scroll region with dynamic padding calculations"""
        try:
            if hasattr(self, 'cards_canvas') and self.cards_canvas:
                self.window.after_idle(lambda: self._finalize_scroll_region_dynamic())
        except Exception as e:
            print(f"[SCROLL DYNAMIC] Error updating scroll region: {e}")
    
    def _finalize_scroll_region_dynamic(self):
        """Finalize scroll region calculation with dynamic values"""
        try:
            # Get canvas bounds with multiple attempts for accuracy
            bbox = None
            for attempt in range(3):
                bbox = self.cards_canvas.bbox("all")
                if bbox:
                    break
                self.cards_scrollable_frame.update_idletasks()
            
            # Use dynamic padding
            padding = getattr(self, 'dynamic_scroll_padding', 150)
            
            if bbox:
                x1, y1, x2, y2 = bbox
                self.cards_canvas.configure(scrollregion=(x1, y1, x2, y2 + padding))
                print(f"[SCROLL DYNAMIC] Updated scrollregion with {padding}px dynamic padding: {self.cards_canvas.cget('scrollregion')}")
            else:
                # Enhanced fallback with dynamic calculations
                widget_count = len(self.card_widgets)
                if widget_count > 0:
                    current_height = self.window.winfo_height()
                    if self.focus_mode:
                        cards_per_row = 4
                        rows = (widget_count + cards_per_row - 1) // cards_per_row
                        # Dynamic row height based on window size
                        row_height = max(int(current_height * 0.25), 180)  # 25% of window height, min 180px
                        height = rows * row_height + padding
                    else:
                        # Dynamic card height based on window size  
                        card_height = max(int(current_height * 0.35), 220)  # 35% of window height, min 220px
                        height = widget_count * card_height + padding
                    
                    self.cards_canvas.configure(scrollregion=(0, 0, 0, height))
                    print(f"[SCROLL DYNAMIC] Used dynamic fallback height: {height} (padding: {padding})")
                    
        except Exception as e:
            print(f"[SCROLL DYNAMIC] Error finalizing scroll region: {e}")
    
    def update_scroll_region_simple(self):
        """Simple, robust scroll region update with enhanced bottom padding"""
        try:
            if hasattr(self, 'cards_canvas') and self.cards_canvas:
                # Force layout update multiple times to ensure accurate measurements
                self.cards_scrollable_frame.update_idletasks()
                self.window.update_idletasks()
                
                # Use dynamic scroll region update
                self.update_scroll_region_dynamic()
                    
        except Exception as e:
            print(f"[SCROLL SIMPLE] Error updating scroll region: {e}")
    
    def _finalize_scroll_region(self):
        """Finalize scroll region calculation after layout stabilizes"""
        try:
            # Get canvas bounds with multiple attempts for accuracy
            bbox = None
            for attempt in range(3):
                bbox = self.cards_canvas.bbox("all")
                if bbox:
                    break
                self.cards_scrollable_frame.update_idletasks()
            
            # Use dynamic padding
            padding = getattr(self, 'dynamic_scroll_padding', 150)
            
            if bbox:
                x1, y1, x2, y2 = bbox
                self.cards_canvas.configure(scrollregion=(x1, y1, x2, y2 + padding))
                print(f"[SCROLL FINALIZE] Updated scrollregion with {padding}px padding: {self.cards_canvas.cget('scrollregion')}")
            else:
                # Enhanced fallback with dynamic height estimation
                widget_count = len(self.card_widgets)
                if widget_count > 0:
                    current_height = self.window.winfo_height()
                    if self.focus_mode:
                        cards_per_row = 4
                        rows = (widget_count + cards_per_row - 1) // cards_per_row
                        # Dynamic row height based on window size
                        row_height = max(int(current_height * 0.25), 180)
                        height = rows * row_height + padding
                    else:
                        # Dynamic card height based on window size
                        card_height = max(int(current_height * 0.35), 220)
                        height = widget_count * card_height + padding
                    
                    self.cards_canvas.configure(scrollregion=(0, 0, 0, height))
                    print(f"[SCROLL FINALIZE] Used enhanced dynamic fallback height: {height}")
                    
        except Exception as e:
            print(f"[SCROLL FINALIZE] Error finalizing scroll region: {e}")
    
    
    def auto_scroll_to_bottom(self):
        """Auto-scroll to bottom to show newest cards"""
        try:
            if hasattr(self, 'cards_canvas') and self.cards_canvas:
                self.cards_canvas.yview_moveto(1.0)
                print("[SCROLL SIMPLE] Auto-scrolled to bottom")
        except Exception as e:
            print(f"[SCROLL SIMPLE] Error auto-scrolling: {e}")
    
    def setup_virtual_scrolling(self):
        """Set up simple virtual scrolling for performance"""
        try:
            print("[VIRTUAL SCROLL] Setting up simplified virtual scrolling")
            
            # Calculate initial visible range
            self.visible_start_index = 0
            self.visible_end_index = min(self.virtual_window_size, len(self.pack_session.cards))
            
            # Create initial display
            self.update_virtual_display_simple()
            
            # Set up scroll event handler
            def on_scroll_event(event):
                """Handle scroll events for virtual scrolling"""
                self.handle_virtual_scroll_event()
            
            # Bind scroll events
            self.cards_canvas.bind("<MouseWheel>", on_scroll_event)
            self.cards_canvas.bind("<Button-4>", on_scroll_event)
            self.cards_canvas.bind("<Button-5>", on_scroll_event)
            
            # Update virtual display periodically
            self.schedule_virtual_update()
            
        except Exception as e:
            print(f"[VIRTUAL SCROLL] Error setting up virtual scrolling: {e}")
            # Fallback to regular display
            self.use_virtual_scrolling = False
            self.safe_update_cards_display()
    
    def schedule_virtual_update(self):
        """Schedule periodic virtual display updates"""
        try:
            self.handle_virtual_scroll_event()
            # Schedule next update
            self.window.after(200, self.schedule_virtual_update)
        except Exception as e:
            print(f"[VIRTUAL SCROLL] Error in scheduled update: {e}")
    
    def handle_virtual_scroll_event(self):
        """Handle virtual scroll events"""
        try:
            if not hasattr(self, 'cards_canvas') or not self.cards_canvas:
                return
                
            # Get current scroll position
            try:
                scroll_top, scroll_bottom = self.cards_canvas.yview()
            except:
                return
                
            total_cards = len(self.pack_session.cards)
            if total_cards == 0:
                return
            
            # Calculate which cards should be visible
            # Simple approach: divide scroll position by card ranges
            buffer_size = 5  # Extra cards to render for smoother scrolling
            
            if self.focus_mode:
                cards_per_row = 4
                total_rows = (total_cards + cards_per_row - 1) // cards_per_row
                current_row = int(scroll_top * total_rows)
                visible_rows = max(5, self.virtual_window_size // cards_per_row)
                
                new_start_index = max(0, (current_row - buffer_size) * cards_per_row)
                new_end_index = min(total_cards, (current_row + visible_rows + buffer_size) * cards_per_row)
            else:
                current_card = int(scroll_top * total_cards)
                visible_cards = self.virtual_window_size
                
                new_start_index = max(0, current_card - buffer_size)
                new_end_index = min(total_cards, current_card + visible_cards + buffer_size * 2)
            
            # Update display if range changed significantly
            if (abs(new_start_index - self.visible_start_index) > 3 or 
                abs(new_end_index - self.visible_end_index) > 3):
                
                self.visible_start_index = new_start_index
                self.visible_end_index = new_end_index
                self.update_virtual_display_simple()
                
        except Exception as e:
            print(f"[VIRTUAL SCROLL] Error handling scroll event: {e}")
    
    @performance_timer("update_virtual_display_simple")
    def update_virtual_display_simple(self):
        """Simple virtual display update"""
        try:
            if not hasattr(self, 'pack_session') or not self.pack_session:
                return
                
            # Clear existing widgets efficiently
            self.clear_all_widgets()
            
            # Create spacer for virtual scrolling offset
            if self.visible_start_index > 0:
                spacer_height = self.calculate_spacer_height(0, self.visible_start_index)
                spacer = tk.Frame(self.cards_scrollable_frame, height=spacer_height)
                spacer.grid(row=0, column=0, columnspan=4 if self.focus_mode else 1, sticky=(tk.W, tk.E))
                self.card_widgets.append(spacer)
            
            # Create widgets for visible cards
            widget_row_offset = 1 if self.visible_start_index > 0 else 0
            
            for i in range(self.visible_start_index, min(self.visible_end_index, len(self.pack_session.cards))):
                card = self.pack_session.cards[i]
                
                try:
                    if self.focus_mode:
                        self.create_focus_mode_widget_virtual(card, i, widget_row_offset)
                    else:
                        self.create_normal_mode_widget_virtual(card, i, widget_row_offset)
                except Exception as e:
                    print(f"[VIRTUAL SCROLL] Error creating widget {i}: {e}")
            
            # Create bottom spacer if needed
            if self.visible_end_index < len(self.pack_session.cards):
                spacer_height = self.calculate_spacer_height(self.visible_end_index, len(self.pack_session.cards))
                bottom_row = widget_row_offset + self.get_widget_rows_used()
                spacer = tk.Frame(self.cards_scrollable_frame, height=spacer_height)
                spacer.grid(row=bottom_row, column=0, columnspan=4 if self.focus_mode else 1, sticky=(tk.W, tk.E))
                self.card_widgets.append(spacer)
            
            # Update scroll region
            self.update_scroll_region_simple()
            
            print(f"[VIRTUAL SCROLL] Updated virtual display: {self.visible_start_index}-{self.visible_end_index}")
            
        except Exception as e:
            print(f"[VIRTUAL SCROLL] Error in simple virtual update: {e}")
    
    def calculate_spacer_height(self, start_card, end_card):
        """Calculate height needed for spacer representing hidden cards with dynamic sizing"""
        card_count = end_card - start_card
        if card_count <= 0:
            return 0
        
        # Get current window height for dynamic calculations
        current_height = self.window.winfo_height()
        
        if self.focus_mode:
            cards_per_row = 4
            rows = (card_count + cards_per_row - 1) // cards_per_row
            # Dynamic row height based on window size
            row_height = max(int(current_height * 0.25), 150)  # 25% of window height, min 150px
            return rows * row_height
        else:
            # Dynamic card height based on window size
            card_height = max(int(current_height * 0.3), 200)  # 30% of window height, min 200px
            return card_count * card_height
    
    def get_widget_rows_used(self):
        """Get number of rows used by currently visible widgets"""
        visible_cards = self.visible_end_index - self.visible_start_index
        if visible_cards <= 0:
            return 0
            
        if self.focus_mode:
            cards_per_row = 4
            return (visible_cards + cards_per_row - 1) // cards_per_row
        else:
            return visible_cards
    
    def create_focus_mode_widget_virtual(self, card, index, row_offset):
        """Create focus mode widget for virtual scrolling"""
        try:
            cards_per_row = 4
            virtual_row = (index - self.visible_start_index) // cards_per_row + row_offset
            col = index % cards_per_row
            
            # Configure grid with dynamic column sizing
            for c in range(cards_per_row):
                min_col_width = getattr(self, 'dynamic_column_width', 220)
                self.cards_scrollable_frame.columnconfigure(c, weight=1, minsize=min_col_width)
            
            # Create simplified card frame
            card_frame = ttk.Frame(self.cards_scrollable_frame, padding="6", relief="solid")
            card_frame.grid(row=virtual_row, column=col, sticky=(tk.W, tk.E, tk.N), pady=3, padx=3)
            self.card_widgets.append(card_frame)
            
            # Compact layout for performance
            self.create_compact_card_content(card_frame, card, index, focus_mode=True)
            
        except Exception as e:
            print(f"[VIRTUAL SCROLL] Error creating focus widget {index}: {e}")
    
    def create_normal_mode_widget_virtual(self, card, index, row_offset):
        """Create normal mode widget for virtual scrolling"""
        try:
            virtual_row = (index - self.visible_start_index) + row_offset
            
            # Configure grid
            self.cards_scrollable_frame.columnconfigure(0, weight=1)
            
            # Create simplified card frame
            card_frame = ttk.Frame(self.cards_scrollable_frame, padding="8", relief="solid")
            card_frame.grid(row=virtual_row, column=0, sticky=(tk.W, tk.E), pady=4, padx=6)
            self.card_widgets.append(card_frame)
            
            # Compact layout for performance
            self.create_compact_card_content(card_frame, card, index, focus_mode=False)
            
        except Exception as e:
            print(f"[VIRTUAL SCROLL] Error creating normal widget {index}: {e}")
    
    def create_compact_card_content(self, parent, card, index, focus_mode):
        """Create compact card content for better performance"""
        try:
            # Configure parent grid
            parent.columnconfigure(1, weight=1)
            
            # Add image if enabled (smaller, simpler)
            if self.display_settings.get('show_images', True) and PIL_AVAILABLE:
                image_label = ttk.Label(parent, text="🖼️", font=("Arial", 8))
                image_label.grid(row=0, column=0, rowspan=3, padx=(0, 8), pady=2, sticky="n")
                
                # Load image asynchronously for better performance
                self.load_card_image_async(card, image_label, focus_mode)
            
            # Essential card info only
            card_name = card.get('card_name', 'N/A')
            if focus_mode and len(card_name) > 20:
                card_name = card_name[:17] + "..."
                
            name_label = ttk.Label(parent, text=f"{index+1}. {card_name}", 
                                 font=("Arial", 9 if focus_mode else 11, "bold"), 
                                 foreground="#2E86AB")
            name_label.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=1)
            
            # Price info (most important)
            price_info = self.get_compact_price_info(card)
            price_label = ttk.Label(parent, text=f"💰 {price_info}", 
                                  font=("Arial", 8 if focus_mode else 10), 
                                  foreground="#0066CC")
            price_label.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=1)
            
            # Quantity controls (essential)
            self.add_compact_quantity_controls(parent, card, index, focus_mode)
            
        except Exception as e:
            print(f"[COMPACT CONTENT] Error creating content for card {index}: {e}")
    
    def load_card_image_async(self, card, image_label, focus_mode):
        """Load card image asynchronously for better performance"""
        try:
            card_images = card.get('card_images', [])
            if not card_images:
                image_label.configure(text="🃏")
                return
            
            image_url = card_images[0].get('image_url')
            if not image_url:
                image_label.configure(text="🃏")
                return
            
            card_id = card.get('id', 'unknown')
            
            # Use smaller sizes for better performance
            size = (60, 90) if focus_mode else (100, 145)
            
            # Try to get cached image first
            photo = self.image_manager.load_image_for_display(card_id, image_url, size)
            
            if photo and hasattr(image_label, 'winfo_exists') and image_label.winfo_exists():
                image_label.configure(image=photo, text="")
                image_label.image = photo
            else:
                image_label.configure(text="❌")
                
        except Exception as e:
            print(f"[ASYNC IMAGE] Error loading image: {e}")
            image_label.configure(text="❌")
    
    def add_compact_quantity_controls(self, parent, card, index, focus_mode):
        """Add compact quantity controls"""
        try:
            qty_frame = ttk.Frame(parent)
            qty_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=2)
            
            current_qty = card.get('quantity', 1)
            card_id = card.get('id', 'unknown')
            
            # Simplified controls
            def decrease_qty():
                self.update_card_quantity_by_id(card_id, -1)
            
            def increase_qty():
                self.update_card_quantity_by_id(card_id, 1)
            
            def remove_card():
                self.remove_card_by_id(card_id)
            
            # Compact layout
            ttk.Button(qty_frame, text="-", width=2, command=decrease_qty).pack(side=tk.LEFT)
            ttk.Label(qty_frame, text=str(current_qty), font=("Arial", 8, "bold"), width=2).pack(side=tk.LEFT, padx=2)
            ttk.Button(qty_frame, text="+", width=2, command=increase_qty).pack(side=tk.LEFT)
            
            if not focus_mode:
                ttk.Button(qty_frame, text="Remove", command=remove_card).pack(side=tk.LEFT, padx=(10, 0))
                
        except Exception as e:
            print(f"[COMPACT QTY] Error creating quantity controls: {e}")
    
    
    def modify_existing_widgets_in_place(self):
        """Modify existing widgets to switch between focus and normal mode without recreation"""
        try:
            cards_per_row = 4 if self.focus_mode else 1
            
            # Pre-load images for new mode if needed - this will use cached resized images
            if PIL_AVAILABLE and self.display_settings.get('show_images', True):
                print(f"[FOCUS MODE] Ensuring images are available for {'focus' if self.focus_mode else 'normal'} mode")
                for card in self.pack_session.cards:
                    card_id = card.get('id', 'unknown')
                    # Check if image is already cached for target mode
                    cached_image = self.image_manager.get_cached_image_for_mode(card_id, self.focus_mode)
                    if not cached_image:
                        # Only load if not already cached - this should be very fast now
                        self.image_manager.get_image(card, focus_mode=self.focus_mode)
            
            # Modify grid layout and content of existing widgets
            for i, widget in enumerate(self.card_widgets):
                try:
                    if not widget.winfo_exists():
                        continue
                        
                    card = list(self.pack_session.cards)[i] if i < len(self.pack_session.cards) else None
                    if not card:
                        continue
                    
                    # Update grid position
                    if self.focus_mode:
                        row = i // cards_per_row
                        col = i % cards_per_row
                    else:
                        row = i
                        col = 0
                    
                    widget.grid_configure(row=row, column=col)
                    
                    # Update widget content for new mode - now using cached images
                    self.update_widget_content_for_mode(widget, card, i)
                    
                except Exception as e:
                    print(f"[FOCUS MODE] Error updating widget {i}: {e}")
                    continue
            
            # Configure column weights for new layout
            max_cols = cards_per_row if self.focus_mode else 1
            for c in range(max_cols):
                self.cards_scrollable_frame.columnconfigure(c, weight=1)
            
            # Clear unused columns if switching to normal mode
            if not self.focus_mode:
                for c in range(1, 4):
                    self.cards_scrollable_frame.columnconfigure(c, weight=0)
            
            # Update scroll region
            self.update_scroll_region()
            
            print(f"[FOCUS MODE] Successfully modified {len(self.card_widgets)} widgets in place")
            
        except Exception as e:
            print(f"[FOCUS MODE] Error in modify_existing_widgets_in_place: {e}")
            raise
    
    def bulk_preload_images_for_session(self):
        """Pre-load all images for the current session in both modes for maximum performance"""
        if not PIL_AVAILABLE or not self.display_settings.get('show_images', True):
            return
        
        # Prevent multiple runs
        if self.bulk_preload_started:
            print("[IMAGE PRELOAD] Bulk preload already running, skipping")
            return
            
        try:
            card_count = len(self.pack_session.cards)
            if card_count == 0:
                return
            
            self.bulk_preload_started = True
            print(f"[IMAGE PRELOAD] Starting optimized bulk preload for {card_count} cards")
            
            def preload_worker():
                """Background worker to preload images with throttling"""
                preloaded = 0
                skipped = 0
                
                for i, card in enumerate(self.pack_session.cards):
                    try:
                        card_images = card.get('card_images', [])
                        if card_images and len(card_images) > 0:
                            image_url = card_images[0].get('image_url')
                            card_id = card.get('id', 'unknown')
                            
                            if image_url and card_id:
                                # More efficient check - look for both disk cache and memory cache
                                focus_cached = self.image_manager.is_image_fully_cached(card_id, True)
                                normal_cached = self.image_manager.is_image_fully_cached(card_id, False)
                                
                                if focus_cached and normal_cached:
                                    skipped += 1
                                else:
                                    # Pre-load only missing modes
                                    if not focus_cached:
                                        self.image_manager.load_image_for_display(card_id, image_url, self.image_manager.focus_mode_size)
                                    if not normal_cached:
                                        self.image_manager.load_image_for_display(card_id, image_url, self.image_manager.normal_mode_size)
                                    preloaded += 1
                                    
                                    # Throttle to avoid overwhelming the system
                                    if preloaded % 5 == 0:
                                        time.sleep(0.01)  # Brief pause every 5 images
                                    
                                    # Show progress less frequently to reduce log spam
                                    if preloaded % 25 == 0:
                                        print(f"[IMAGE PRELOAD] Processed {preloaded} cards, skipped {skipped} already cached...")
                                        
                    except Exception as e:
                        print(f"[IMAGE PRELOAD] Error preloading card {i}: {e}")
                        continue
                
                print(f"[IMAGE PRELOAD] Bulk preload completed - processed {preloaded} cards, skipped {skipped} already cached")
                self.bulk_preload_started = False  # Reset flag when done
            
            # Run preload in background with lower priority
            preload_thread = threading.Thread(target=preload_worker, daemon=True)
            preload_thread.start()
            
        except Exception as e:
            print(f"[IMAGE PRELOAD] Error in bulk preload: {e}")
            self.bulk_preload_started = False  # Reset flag on error
    
    def update_widget_content_for_mode(self, widget, card, index):
        """Update widget content to match current focus mode using cached images"""
        try:
            # Find image and content elements within the widget
            for child in widget.winfo_children():
                if hasattr(child, '_image_content') and child._image_content:
                    # Update image for new mode - use cached image if available
                    if PIL_AVAILABLE and self.display_settings.get('show_images', True):
                        card_id = card.get('id', 'unknown')
                        
                        # Try to get cached image first for maximum speed
                        cached_image = self.image_manager.get_cached_image_for_mode(card_id, self.focus_mode)
                        if cached_image:
                            print(f"[FOCUS MODE] Using cached image for card {card_id}")
                            child.configure(image=cached_image)
                            child.image = cached_image  # Keep a reference
                        else:
                            # Fallback to get_image which should be fast with disk caching
                            image = self.image_manager.get_image(card, focus_mode=self.focus_mode)
                            if image:
                                print(f"[FOCUS MODE] Loaded image for card {card_id}")
                                child.configure(image=image)
                                child.image = image  # Keep a reference
                elif isinstance(child, ttk.Frame):
                    # Update text content in the content frame
                    self.update_text_content_for_mode(child, card, index)
                    
        except Exception as e:
            print(f"[FOCUS MODE] Error updating widget content: {e}")
            import traceback
            traceback.print_exc()
    
    def update_text_content_for_mode(self, content_frame, card, index):
        """Update text content based on current mode"""
        try:
            card_name = card.get('card_name', 'N/A')
            
            # Truncate name based on mode
            if self.focus_mode and len(card_name) > 22:
                card_name = card_name[:19] + "..."
            elif not self.focus_mode and len(card_name) > 40:
                card_name = card_name[:37] + "..."
            
            # Find and update labels within content frame
            for child in content_frame.winfo_children():
                if isinstance(child, ttk.Label):
                    text = child.cget('text')
                    if text.startswith(f"{index+1}."):
                        # Update card name label
                        if self.focus_mode:
                            child.configure(text=f"{index+1}. {card_name}", font=("Arial", 9, "bold"))
                        else:
                            child.configure(text=f"{index+1}. {card_name}", font=("Arial", 12, "bold"))
                        break
                        
        except Exception as e:
            print(f"[FOCUS MODE] Error updating text content: {e}")
    
    def ensure_all_images_preloaded_for_mode(self):
        """Pre-load ALL images for target mode to eliminate loading time during rebuild"""
        if not PIL_AVAILABLE or not self.display_settings.get('show_images', True):
            return
            
        try:
            print(f"[FOCUS MODE] Pre-loading ALL images for {'focus' if self.focus_mode else 'normal'} mode")
            
            # Pre-load images for ALL cards in the target mode
            for i, card in enumerate(self.pack_session.cards):
                card_images = card.get('card_images', [])
                if card_images and len(card_images) > 0:
                    image_url = card_images[0].get('image_url')
                    card_id = card.get('id', 'unknown')
                    
                    if image_url and card_id:
                        # Load image for target mode if not already cached
                        cached_image = self.image_manager.get_cached_image_for_mode(card_id, self.focus_mode)
                        
                        if not cached_image:
                            # Load and cache the image synchronously
                            if self.focus_mode:
                                self.image_manager.load_image_for_display(card_id, image_url, self.image_manager.focus_mode_size)
                            else:
                                self.image_manager.load_image_for_display(card_id, image_url, self.image_manager.normal_mode_size)
            
            print(f"[FOCUS MODE] All images pre-loaded for {'focus' if self.focus_mode else 'normal'} mode")
                            
        except Exception as e:
            print(f"[FOCUS MODE] Error pre-loading images: {e}")
    
    def fast_rebuild_with_preloaded_resources(self):
        """Super-fast widget rebuild using pre-loaded resources"""
        try:
            # Store reference to cards to avoid lookups
            cards_to_rebuild = list(self.pack_session.cards)
            
            # Quick cleanup of existing widgets
            self.quick_cleanup_widgets()
            
            # Super-fast widget creation with pre-loaded resources
            print(f"[FOCUS MODE] Creating {len(cards_to_rebuild)} widgets with pre-loaded resources")
            
            for i, card in enumerate(cards_to_rebuild):
                if self.focus_mode:
                    self.create_focus_mode_widget_fast(card, i)
                else:
                    self.create_normal_mode_widget_fast(card, i)
                
            # Quick scroll region update
            self.update_scroll_region()
                
        except Exception as e:
            print(f"[FOCUS MODE] Error in fast rebuild: {e}")
            import traceback
            traceback.print_exc()
    
    def quick_cleanup_widgets(self):
        """Quick cleanup of existing widgets without expensive operations"""
        try:
            widgets_to_destroy = list(self.card_widgets)
            self.card_widgets = []
            
            # Quick destruction without complex cleanup
            for widget in widgets_to_destroy:
                try:
                    if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                        widget.destroy()
                except:
                    pass  # Ignore errors during cleanup for speed
                    
        except Exception as e:
            print(f"[FOCUS MODE] Error in quick cleanup: {e}")
    
    def create_focus_mode_widget_fast(self, card, index):
        """Create focus mode widget with pre-loaded resources (very fast)"""
        try:
            cards_per_row = 4
            row = index // cards_per_row
            col = index % cards_per_row
            
            # Configure columns
            for c in range(cards_per_row):
                self.cards_scrollable_frame.columnconfigure(c, weight=1)
            
            # Create compact card frame
            card_frame = ttk.Frame(self.cards_scrollable_frame, padding="5", relief="ridge")
            card_frame.grid(row=row, column=col, sticky=(tk.W, tk.E, tk.N), pady=2, padx=2)
            self.card_widgets.append(card_frame)
            
            # Configure grid
            card_frame.columnconfigure(0, weight=0)
            card_frame.columnconfigure(1, weight=1)
            
            # Add image display (fast with pre-loaded images)
            if self.display_settings.get('show_images', True) and PIL_AVAILABLE:
                image_label = ttk.Label(card_frame, text="🖼️", 
                                      font=("Arial", 6), justify="center", foreground="#666666")
                image_label.grid(row=0, column=0, rowspan=4, padx=(0, 5), pady=2, sticky="n")
                image_label._image_content = True
                
                # Load pre-cached image (should be instant)
                self.load_precached_image_fast(card, image_label)
            
            # Content column (same as before)
            content_frame = ttk.Frame(card_frame)
            content_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N))
            content_frame.columnconfigure(0, weight=1)
            
            # Rest of the focus mode widget creation (unchanged)
            card_name = card.get('card_name', 'N/A')
            if len(card_name) > 22:
                card_name = card_name[:19] + "..."
            
            name_label = ttk.Label(content_frame, text=f"{index+1}. {card_name}", 
                                 font=("Arial", 9, "bold"), foreground="#2E86AB")
            name_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=1)
            
            rarity = card.get('card_rarity', 'N/A')
            if len(rarity) > 18:
                rarity = rarity[:15] + "..."
            rarity_label = ttk.Label(content_frame, text=f"💎 {rarity}", 
                                   font=("Arial", 8), foreground="#666666")
            rarity_label.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=1)
            
            price_info = self.get_compact_price_info(card)
            price_label = ttk.Label(content_frame, text=f"💰 {price_info}", 
                                  font=("Arial", 8), foreground="#0066CC")
            price_label.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=1)
            
            # Quantity controls with stable card references
            qty_frame = ttk.Frame(content_frame)
            qty_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=1)
            
            current_qty = card.get('quantity', 1)
            card_id = card.get('id', 'unknown')  # Use card ID for stable reference
            
            def decrease_qty():
                # Find card by ID instead of using potentially stale index
                for i, session_card in enumerate(self.pack_session.cards):
                    if session_card.get('id') == card_id:
                        current_card_qty = session_card.get('quantity', 1)
                        new_qty = max(1, current_card_qty - 1)
                        self.update_card_quantity(i, new_qty)
                        break
            
            def increase_qty():
                # Find card by ID instead of using potentially stale index
                for i, session_card in enumerate(self.pack_session.cards):
                    if session_card.get('id') == card_id:
                        current_card_qty = session_card.get('quantity', 1)
                        new_qty = current_card_qty + 1
                        self.update_card_quantity(i, new_qty)
                        break
            
            def remove_card():
                # Find card by ID for removal
                for i, session_card in enumerate(self.pack_session.cards):
                    if session_card.get('id') == card_id:
                        self.remove_card_from_session(i)
                        break
            
            ttk.Button(qty_frame, text="-", width=2, command=decrease_qty).pack(side=tk.LEFT)
            ttk.Label(qty_frame, text=str(current_qty), font=("Arial", 8, "bold"), width=2).pack(side=tk.LEFT)
            ttk.Button(qty_frame, text="+", width=2, command=increase_qty).pack(side=tk.LEFT)
            ttk.Button(qty_frame, text="🗑️", width=3, command=remove_card).pack(side=tk.RIGHT)
            
        except Exception as e:
            print(f"[FOCUS MODE] Error creating fast focus widget {index}: {e}")
    
    def create_normal_mode_widget_fast(self, card, index):
        """Create normal mode widget with pre-loaded resources (very fast)"""
        try:
            # Configure single column
            self.cards_scrollable_frame.columnconfigure(0, weight=1)
            
            # Create card frame
            card_frame = ttk.Frame(self.cards_scrollable_frame, padding="10", relief="ridge")
            card_frame.grid(row=index, column=0, sticky=(tk.W, tk.E), pady=5, padx=5)
            self.card_widgets.append(card_frame)
            
            # Configure grid
            card_frame.columnconfigure(1, weight=1)
            
            current_row = 0
            
            # Image display with pre-loaded images (fast)
            if self.display_settings.get('show_images', True) and PIL_AVAILABLE:
                image_label = ttk.Label(card_frame, text="🖼️\nLoading...", 
                                      font=("Arial", 8), justify="center", foreground="#666666")
                image_label.grid(row=0, column=0, rowspan=6, padx=(0, 15), pady=5)
                image_label._image_content = True
                
                # Load pre-cached image (should be instant)
                self.load_precached_image_fast(card, image_label)
                
            # Rest of normal mode widget creation (unchanged from original)
            if self.display_settings.get('show_card_name', True):
                name_label = ttk.Label(card_frame, text=f"📋 {card.get('card_name', 'N/A')}", 
                                     font=("Arial", 12, "bold"))
                name_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            if self.display_settings.get('show_rarity', True):
                rarity_label = ttk.Label(card_frame, text=f"💎 Rarity: {card.get('card_rarity', 'N/A')}", 
                                       font=("Arial", 10))
                rarity_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            if self.display_settings.get('show_art_variant', True):
                variant_label = ttk.Label(card_frame, text=f"🎨 Art Variant: {card.get('art_variant', 'None')}", 
                                        font=("Arial", 10))
                variant_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            # Add price information
            current_row = self.add_price_labels(card_frame, card, current_row)
            
            # Add other fields as before...
            if self.display_settings.get('show_set_info', False):
                set_name = card.get('set_name', 'N/A')
                set_code = card.get('set_code', 'N/A')
                set_label = ttk.Label(card_frame, text=f"📚 Set: {set_name} ({set_code})", 
                                    font=("Arial", 10))
                set_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            if self.display_settings.get('show_timestamps', False):
                timestamp = card.get('timestamp', datetime.now().strftime("%H:%M:%S"))
                time_label = ttk.Label(card_frame, text=f"🕒 Added: {timestamp}", 
                                     font=("Arial", 10))
                time_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
                current_row += 1
            
            # Quantity controls with stable card references
            qty_frame = ttk.Frame(card_frame)
            qty_frame.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            
            ttk.Label(qty_frame, text="📦 Quantity:", font=("Arial", 10)).pack(side=tk.LEFT)
            
            current_qty = card.get('quantity', 1)
            card_id = card.get('id', 'unknown')  # Use card ID for stable reference
            
            def decrease_qty():
                # Find card by ID instead of using potentially stale index
                for i, session_card in enumerate(self.pack_session.cards):
                    if session_card.get('id') == card_id:
                        current_card_qty = session_card.get('quantity', 1)
                        new_qty = max(1, current_card_qty - 1)
                        self.update_card_quantity(i, new_qty)
                        break
            
            def increase_qty():
                # Find card by ID instead of using potentially stale index
                for i, session_card in enumerate(self.pack_session.cards):
                    if session_card.get('id') == card_id:
                        current_card_qty = session_card.get('quantity', 1)
                        new_qty = current_card_qty + 1
                        self.update_card_quantity(i, new_qty)
                        break
            
            def remove_card():
                # Find card by ID for removal
                for i, session_card in enumerate(self.pack_session.cards):
                    if session_card.get('id') == card_id:
                        self.remove_card_from_session(i)
                        break
            
            ttk.Button(qty_frame, text="-", width=3, command=decrease_qty).pack(side=tk.LEFT, padx=2)
            ttk.Label(qty_frame, text=str(current_qty), font=("Arial", 10, "bold"), width=3).pack(side=tk.LEFT)
            ttk.Button(qty_frame, text="+", width=3, command=increase_qty).pack(side=tk.LEFT, padx=2)
            ttk.Button(qty_frame, text="Remove", command=remove_card).pack(side=tk.LEFT, padx=10)
            
        except Exception as e:
            print(f"[FOCUS MODE] Error creating fast normal widget {index}: {e}")
    
    def load_precached_image_fast(self, card, image_label):
        """Load pre-cached image (should be instant)"""
        try:
            # Get card image info
            card_images = card.get('card_images', [])
            if not card_images or len(card_images) == 0:
                self.set_image_placeholder(image_label, "🃏\nNo Image")
                return
            
            card_id = card.get('id', 'unknown')
            
            # Get pre-cached image (should be instant)
            photo = self.image_manager.get_cached_image_for_mode(card_id, self.focus_mode)
            
            if photo:
                # Image is pre-cached, just set it
                if hasattr(image_label, 'winfo_exists') and image_label.winfo_exists():
                    image_label.configure(image=photo, text="")
                    image_label.image = photo
            else:
                # Fallback if somehow not cached
                self.set_image_placeholder(image_label, "🖼️\nFallback")
                
        except Exception as e:
            print(f"[FOCUS MODE] Error loading pre-cached image: {e}")
            self.set_image_placeholder(image_label, "❌\nError")
    
    def ensure_images_cached_for_mode(self):
        """Legacy method - now redirects to ensure_all_images_preloaded_for_mode"""
        return self.ensure_all_images_preloaded_for_mode()
    


        
    def safe_update_cards_display(self):
        """Thread-safe wrapper for updating cards display with simplified approach"""
        print("[SESSION UPDATE DEBUG] safe_update_cards_display called")
        
        with self.ui_update_lock:
            if self.is_updating_display:
                print("[SESSION TRACKER] Update already in progress, skipping...")
                return
            self.is_updating_display = True
        
        try:
            print("[SESSION UPDATE DEBUG] Starting display update")
            self.update_cards_display_simple()
        finally:
            with self.ui_update_lock:
                self.is_updating_display = False
    
    def update_cards_display_simple(self):
        """Simple, robust display update that handles all cases with virtual scrolling optimization"""
        try:
            print(f"[SESSION TRACKER] Simple update with {len(self.pack_session.cards)} cards")
            
            # Safety check - ensure window still exists
            if not (hasattr(self, 'window') and self.window):
                print("[SESSION TRACKER] Window no longer exists, skipping display update")
                return
            
            # Update card count
            if hasattr(self, 'card_count_label') and self.card_count_label:
                try:
                    self.card_count_label.config(text=f"Cards: {len(self.pack_session.cards)}")
                except Exception as e:
                    print(f"[SESSION TRACKER] Error updating card count: {e}")
            
            # Check if we should switch to/from virtual scrolling based on card count
            should_use_virtual = len(self.pack_session.cards) > 25
            if should_use_virtual != self.use_virtual_scrolling:
                print(f"[PERFORMANCE] Switching virtual scrolling: {self.use_virtual_scrolling} -> {should_use_virtual}")
                self.use_virtual_scrolling = should_use_virtual
                # Force full rebuild when switching modes
                self.clear_all_widgets()
                if self.use_virtual_scrolling:
                    self.setup_virtual_scrolling()
                else:
                    self.rebuild_display_for_mode()
                return
            
            # Use virtual scrolling logic or regular logic
            if self.use_virtual_scrolling:
                # For virtual scrolling, just update the virtual display
                self.update_virtual_display_simple()
            else:
                # Regular update logic
                current_widget_count = len(self.card_widgets)
                required_widget_count = len(self.pack_session.cards)
                
                if required_widget_count == current_widget_count:
                    # Just update existing content (for price updates)
                    print("[SESSION TRACKER] Updating existing card content")
                    self.update_existing_cards_content()
                elif required_widget_count > current_widget_count:
                    # Add new cards
                    print(f"[SESSION TRACKER] Adding {required_widget_count - current_widget_count} new cards")
                    self.add_new_cards_simple(current_widget_count)
                else:
                    # Cards were removed - do full rebuild
                    print("[SESSION TRACKER] Cards removed - doing full rebuild")
                    self.clear_all_widgets()
                    self.rebuild_display_for_mode()
                
        except Exception as e:
            print(f"[SESSION TRACKER] Error in simple display update: {e}")
            import traceback
            traceback.print_exc()
    
    def update_virtual_display_for_new_cards(self):
        """Update virtual display when new cards are added"""
        try:
            # For virtual scrolling, we may need to update the visible range
            # if new cards were added within the current view
            total_cards = len(self.pack_session.cards)
            
            # Update virtual scroll region to account for new cards
            self.setup_virtual_scroll_region()
            
            # If new cards are within current visible range, update display
            if total_cards > self.visible_end_index:
                # Extend visible range if we're near the bottom
                current_scroll_pos = self.cards_canvas.yview()[1]  # Bottom position
                if current_scroll_pos > 0.8:  # If we're near the bottom
                    # Extend visible range to show new cards
                    new_end = min(total_cards, self.visible_end_index + 10)
                    self.update_virtual_display(self.visible_start_index, new_end)
                    # Auto-scroll to show new cards
                    self.window.after(100, self.auto_scroll_to_bottom)
            
            print(f"[VIRTUAL UPDATE] Updated virtual display for {total_cards} total cards")
            
        except Exception as e:
            print(f"[VIRTUAL UPDATE] Error updating virtual display: {e}")
    
    def add_new_cards_simple(self, start_index):
        """Add new cards starting from start_index"""
        try:
            for i in range(start_index, len(self.pack_session.cards)):
                card = self.pack_session.cards[i]
                
                if self.focus_mode:
                    self.create_focus_mode_widget_simple(card, i)
                else:
                    self.create_normal_mode_widget_simple(card, i)
            
            # Update scroll region
            self.update_scroll_region_simple()
            
            # Auto-scroll to show new cards
            self.window.after(100, self.auto_scroll_to_bottom)
            
        except Exception as e:
            print(f"[ADD NEW CARDS] Error adding new cards: {e}")
            import traceback
            traceback.print_exc()
    
    def update_existing_cards_content(self):
        """Update content of existing cards (for price updates)"""
        try:
            for i, card in enumerate(self.pack_session.cards):
                if i < len(self.card_widgets):
                    # Update price information in existing widgets
                    self.update_price_info_in_widget(self.card_widgets[i], card)
                    
        except Exception as e:
            print(f"[UPDATE CONTENT] Error updating existing content: {e}")
    
    def update_price_info_in_widget(self, widget, card):
        """Update price information in a widget"""
        try:
            # Recursively find and update price labels
            self.update_price_labels_recursive(widget, card)
        except Exception as e:
            print(f"[PRICE UPDATE] Error updating price in widget: {e}")
    
    def update_price_labels_recursive(self, widget, card):
        """Recursively find and update price labels"""
        try:
            # Check if this is a price label
            if hasattr(widget, 'cget') and hasattr(widget, 'configure'):
                try:
                    text = widget.cget('text')
                    if text:
                        # Update TCG Low price
                        if text.startswith('💰'):
                            tcg_price = card.get('tcg_price')
                            if tcg_price and tcg_price not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
                                new_text = f"💰 TCG Low: ${tcg_price}" if not self.focus_mode else f"💰 L:${tcg_price}"
                                if text != new_text:
                                    widget.configure(text=new_text)
                        # Update TCG Market price
                        elif text.startswith('📈'):
                            tcg_market = card.get('tcg_market_price')
                            if tcg_market and tcg_market not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
                                new_text = f"📈 TCG Market: ${tcg_market}" if not self.focus_mode else f"📈 M:${tcg_market}"
                                if text != new_text:
                                    widget.configure(text=new_text)
                except:
                    pass
            
            # Check children recursively
            try:
                for child in widget.winfo_children():
                    self.update_price_labels_recursive(child, card)
            except:
                pass
                
        except Exception as e:
            print(f"[PRICE RECURSIVE] Error in recursive price update: {e}")
    
    def update_cards_display(self):
        """Legacy method - redirects to simplified update"""
        self.update_cards_display_simple()
    
    def update_cards_display_incremental(self):
        """Update the cards display incrementally to prevent segmentation faults"""
        try:
            print(f"[SESSION TRACKER] Incremental update with {len(self.pack_session.cards)} cards")
            
            # Safety check - ensure window still exists
            if not (hasattr(self, 'window') and self.window):
                print("[SESSION TRACKER] Window no longer exists, skipping display update")
                return
            
            # Update card count safely
            if hasattr(self, 'card_count_label') and hasattr(self.card_count_label, 'winfo_exists') and self.card_count_label.winfo_exists():
                try:
                    self.card_count_label.config(text=f"Cards: {len(self.pack_session.cards)}")
                except Exception as e:
                    print(f"[SESSION TRACKER] Error updating card count: {e}")
            
            # Check if we need to add new widgets only (avoid destruction/recreation)
            current_widget_count = len(self.card_widgets)
            required_widget_count = len(self.pack_session.cards)
            
            if required_widget_count > current_widget_count:
                # Add new widgets for new cards only
                print(f"[SESSION TRACKER] Adding {required_widget_count - current_widget_count} new card widgets")
                self.add_new_card_widgets(current_widget_count)
                # Also update existing widgets in case their data changed
                print("[SESSION TRACKER] Also updating existing widget content")
                self.update_existing_widgets_range(0, current_widget_count)
            elif required_widget_count < current_widget_count:
                # Remove excess widgets (only if cards were removed)
                print(f"[SESSION TRACKER] Removing {current_widget_count - required_widget_count} excess widgets")
                self.remove_excess_widgets(required_widget_count)
                # Update remaining widgets in case their data changed
                print("[SESSION TRACKER] Also updating remaining widget content")
                self.update_existing_widgets()
            else:
                # Update existing widgets content (for price updates, etc.)
                print("[SESSION TRACKER] Updating existing widget content")
                self.update_existing_widgets()
                    
        except Exception as e:
            print(f"[SESSION TRACKER] Error in incremental update: {e}")
            import traceback
            traceback.print_exc()
    
    def add_new_card_widgets(self, start_index):
        """Add widgets for new cards starting from start_index with optimized image preloading"""
        print(f"[ADD WIDGETS DEBUG] Adding new card widgets from index {start_index}")
        print(f"[ADD WIDGETS DEBUG] Total cards in session: {len(self.pack_session.cards)}")
        print(f"[ADD WIDGETS DEBUG] Current widget count: {len(self.card_widgets)}")
        print(f"[ADD WIDGETS DEBUG] Focus mode: {self.focus_mode}")
        
        try:
            # Double-check window still exists
            if not (hasattr(self, 'window') and self.window and hasattr(self.window, 'winfo_exists') and self.window.winfo_exists()):
                print("[ADD WIDGETS DEBUG] ERROR: Window closed during widget creation")
                return
            
            # Pre-load images for both modes for new cards in background
            if PIL_AVAILABLE and self.display_settings.get('show_images', True):
                print(f"[SESSION TRACKER] Pre-loading images for {len(self.pack_session.cards) - start_index} new cards")
                for i in range(start_index, len(self.pack_session.cards)):
                    card = self.pack_session.cards[i]
                    card_images = card.get('card_images', [])
                    if card_images and len(card_images) > 0:
                        image_url = card_images[0].get('image_url')
                        card_id = card.get('id', 'unknown')
                        if image_url and card_id:
                            # Pre-load for both modes in background
                            threading.Thread(
                                target=self.image_manager.preload_image_for_both_modes,
                                args=(card_id, image_url),
                                daemon=True
                            ).start()
            
            # Add widgets for new cards only
            print(f"[ADD WIDGETS DEBUG] Starting widget creation loop from {start_index} to {len(self.pack_session.cards)}")
            for i in range(start_index, len(self.pack_session.cards)):
                card = self.pack_session.cards[i]
                print(f"[ADD WIDGETS DEBUG] Creating widget {i} for card: {card.get('card_name', 'Unknown')}")
                
                if self.focus_mode:
                    print(f"[ADD WIDGETS DEBUG] Creating focus mode widget for card {i}")
                    self.create_focus_mode_widget(card, i)
                else:
                    print(f"[ADD WIDGETS DEBUG] Creating normal mode widget for card {i}")
                    self.create_normal_mode_widget(card, i)
                
                print(f"[ADD WIDGETS DEBUG] Widget {i} creation completed")
                
            print(f"[ADD WIDGETS DEBUG] Widget creation loop completed. Total widgets now: {len(self.card_widgets)}")
            
            # Update scroll region safely
            print(f"[ADD WIDGETS DEBUG] Calling update_scroll_region...")
            self.update_scroll_region()
                
        except Exception as e:
            print(f"[SESSION TRACKER] Error adding new widgets: {e}")
            import traceback
            traceback.print_exc()
    
    def remove_excess_widgets(self, target_count):
        """Remove excess widgets when cards are removed"""
        try:
            widgets_to_remove = self.card_widgets[target_count:]
            self.card_widgets = self.card_widgets[:target_count]
            
            for widget in widgets_to_remove:
                try:
                    if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                        self.clear_widget_image_references(widget)
                        widget.destroy()
                except Exception as e:
                    print(f"[SESSION TRACKER] Error removing widget: {e}")
            
            # Update scroll region safely
            self.update_scroll_region()
                    
        except Exception as e:
            print(f"[SESSION TRACKER] Error removing excess widgets: {e}")
    
    def update_existing_widgets(self):
        """Update content of existing widgets (for price updates, etc.)"""
        try:
            print("[SESSION TRACKER] Updating existing widget content with new data")
            
            # Update each widget with current card data
            for i, widget in enumerate(self.card_widgets):
                if i < len(self.pack_session.cards):
                    card = self.pack_session.cards[i]
                    self.update_widget_content(widget, card, i)
            
            # Update scroll region after content updates
            self.update_scroll_region()
        except Exception as e:
            print(f"[SESSION TRACKER] Error updating existing widgets: {e}")
            import traceback
            traceback.print_exc()
    
    def update_existing_widgets_range(self, start_index, end_index):
        """Update content of existing widgets in a specific range"""
        try:
            print(f"[SESSION TRACKER] Updating widget content for range {start_index}-{end_index}")
            
            # Update each widget in the range with current card data
            for i in range(start_index, min(end_index, len(self.card_widgets), len(self.pack_session.cards))):
                widget = self.card_widgets[i]
                card = self.pack_session.cards[i]
                self.update_widget_content(widget, card, i)
                
        except Exception as e:
            print(f"[SESSION TRACKER] Error updating existing widgets in range: {e}")
            import traceback
            traceback.print_exc()
    
    def update_widget_content(self, widget, card, index):
        """Update individual widget content with current card data"""
        try:
            if not (hasattr(widget, 'winfo_exists') and widget.winfo_exists()):
                print(f"[SESSION TRACKER] Widget {index} no longer exists, skipping update")
                return
            
            # Update different widgets based on focus mode
            if self.focus_mode:
                self.update_focus_mode_widget_content(widget, card, index)
            else:
                self.update_normal_mode_widget_content(widget, card, index)
                
        except Exception as e:
            print(f"[SESSION TRACKER] Error updating widget {index} content: {e}")
            import traceback
            traceback.print_exc()
    
    def update_focus_mode_widget_content(self, widget, card, index):
        """Update focus mode widget content"""
        try:
            # Find price label in focus mode widget and update it
            for child in widget.winfo_children():
                if hasattr(child, 'cget') and hasattr(child, 'configure'):
                    try:
                        text = child.cget('text')
                        # Update price label (starts with 💰)
                        if text and text.startswith('💰'):
                            price_info = self.get_compact_price_info(card)
                            new_text = f"💰 {price_info}"
                            if text != new_text:
                                child.configure(text=new_text)
                                print(f"[SESSION TRACKER] Updated focus mode price for card {index}: {new_text}")
                        # Update quantity label (check if it's numeric)
                        elif text and text.isdigit():
                            current_qty = card.get('quantity', 1)
                            if text != str(current_qty):
                                child.configure(text=str(current_qty))
                                print(f"[SESSION TRACKER] Updated focus mode quantity for card {index}: {current_qty}")
                    except:
                        pass
        except Exception as e:
            print(f"[SESSION TRACKER] Error updating focus mode widget {index}: {e}")
    
    def update_normal_mode_widget_content(self, widget, card, index):
        """Update normal mode widget content"""
        try:
            # Recursively find and update labels in normal mode
            self.update_labels_recursive(widget, card, index)
        except Exception as e:
            print(f"[SESSION TRACKER] Error updating normal mode widget {index}: {e}")
    
    def update_labels_recursive(self, widget, card, index):
        """Recursively find and update labels in widget tree"""
        try:
            # Check current widget
            if hasattr(widget, 'cget') and hasattr(widget, 'configure'):
                try:
                    text = widget.cget('text')
                    if text:
                        updated_text = self.get_updated_label_text(text, card)
                        if updated_text and updated_text != text:
                            widget.configure(text=updated_text)
                            print(f"[SESSION TRACKER] Updated label for card {index}: {updated_text}")
                except:
                    pass
            
            # Recursively check children
            if hasattr(widget, 'winfo_children'):
                for child in widget.winfo_children():
                    self.update_labels_recursive(child, card, index)
                    
        except Exception as e:
            print(f"[SESSION TRACKER] Error in recursive label update: {e}")
    
    def get_updated_label_text(self, current_text, card):
        """Get updated text for a label based on current card data"""
        try:
            # Update TCG Low price labels
            if current_text.startswith('💰') and ('TCG Low' in current_text or '⏳ Loading TCG price' in current_text):
                tcg_price = card.get('tcg_price')
                if tcg_price and tcg_price not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
                    return f"💰 TCG Low: ${tcg_price}"
                elif tcg_price == '⏳ Loading...':
                    return "💰 ⏳ Loading TCG price..."
                else:
                    return "💰 TCG Low: N/A"
            
            # Update TCG Market price labels
            elif current_text.startswith('📈') and ('TCG Market' in current_text or '⏳ Loading market price' in current_text):
                tcg_market_price = card.get('tcg_market_price')
                if tcg_market_price and tcg_market_price not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
                    return f"📈 TCG Market: ${tcg_market_price}"
                elif tcg_market_price == '⏳ Loading...':
                    return "📈 ⏳ Loading market price..."
                else:
                    return "📈 TCG Market: N/A"
            
            # No update needed
            return None
            
        except Exception as e:
            print(f"[SESSION TRACKER] Error getting updated label text: {e}")
            return None
    
    def update_scroll_region(self):
        """Update scroll region - redirects to simple method"""
        self.update_scroll_region_simple()
    
    def create_focus_mode_widget(self, card, index):
        """Create a single widget in focus mode"""
        print(f"[FOCUS MODE DEBUG] Creating widget for card {index}: {card.get('card_name', 'Unknown')}")
        
        cards_per_row = 4
        row = index // cards_per_row
        col = index % cards_per_row
        
        print(f"[FOCUS MODE DEBUG] Widget {index} positioned at row={row}, col={col}")
        
        # Configure columns with better weight distribution and spacing
        for c in range(cards_per_row):
            self.cards_scrollable_frame.columnconfigure(c, weight=1, minsize=250, pad=5)
        
        print(f"[FOCUS MODE DEBUG] Configured {cards_per_row} columns with weight=1, minsize=250, pad=5")
        
        # Create compact card frame with improved sizing
        card_frame = ttk.Frame(self.cards_scrollable_frame, padding="8", relief="ridge")
        card_frame.grid(row=row, column=col, sticky=(tk.W, tk.E, tk.N, tk.S), pady=3, padx=3, ipadx=3, ipady=3)
        self.card_widgets.append(card_frame)
        
        print(f"[FOCUS MODE DEBUG] Created and gridded card_frame for widget {index}")
        
        # Configure internal grid with better proportions
        card_frame.columnconfigure(0, weight=0, minsize=80)  # Image column - increased from 60
        card_frame.columnconfigure(1, weight=1, minsize=150)  # Content column - ensure minimum width
        
        # Configure rows to prevent content truncation
        for r in range(5):  # Increased to handle all content rows
            card_frame.rowconfigure(r, weight=0, pad=1)
        
        # Add image display for focus mode
        if self.display_settings.get('show_images', True) and PIL_AVAILABLE:
            # Create compact image label
            image_label = ttk.Label(card_frame, text="🖼️", 
                                  font=("Arial", 6), justify="center", foreground="#666666")
            image_label.grid(row=0, column=0, rowspan=4, padx=(0, 5), pady=2, sticky="n")
            
            # Mark this as an image label for focus mode switching
            image_label._image_content = True
            
            # Load image synchronously 
            self.load_card_image_safe(card, image_label)
        
        # Content column for text with improved layout
        content_frame = ttk.Frame(card_frame)
        content_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=2)
        content_frame.columnconfigure(0, weight=1)
        
        # Configure content frame rows properly
        for r in range(5):
            content_frame.rowconfigure(r, weight=0, minsize=20)
        
        print(f"[FOCUS MODE DEBUG] Created content_frame for widget {index}")
        
        # Improved text display with better truncation limits
        card_name = card.get('card_name', 'N/A')
        if len(card_name) > 30:  # Increased limit for better readability
            card_name = card_name[:27] + "..."
        
        # Card name with index
        name_label = ttk.Label(content_frame, text=f"{index+1}. {card_name}", 
                             font=("Arial", 9, "bold"), foreground="#2E86AB", anchor="w")
        name_label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 1))
        
        # Rarity with improved truncation
        rarity = card.get('card_rarity', 'N/A')
        if len(rarity) > 25:  # Increased limit
            rarity = rarity[:22] + "..."
        rarity_label = ttk.Label(content_frame, text=f"💎 {rarity}", 
                               font=("Arial", 8), foreground="#666666", anchor="w")
        rarity_label.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 1))
        
        # Price info
        price_info = self.get_compact_price_info(card)
        price_label = ttk.Label(content_frame, text=f"💰 {price_info}", 
                              font=("Arial", 8), foreground="#0066CC", anchor="w")
        price_label.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 1))
        
        # Quantity controls with stable references
        qty_frame = ttk.Frame(content_frame)
        qty_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(2, 0))
        
        current_qty = card.get('quantity', 1)
        card_id = card.get('id', 'unknown')  # Use card ID for stable reference
        
        def decrease_qty():
            # Find card by ID instead of using potentially stale index
            for i, session_card in enumerate(self.pack_session.cards):
                if session_card.get('id') == card_id:
                    current_card_qty = session_card.get('quantity', 1)
                    new_qty = max(1, current_card_qty - 1)
                    self.update_card_quantity(i, new_qty)
                    break
        
        def increase_qty():
            # Find card by ID instead of using potentially stale index
            for i, session_card in enumerate(self.pack_session.cards):
                if session_card.get('id') == card_id:
                    current_card_qty = session_card.get('quantity', 1)
                    new_qty = current_card_qty + 1
                    self.update_card_quantity(i, new_qty)
                    break
        
        def remove_card():
            # Find card by ID for removal
            for i, session_card in enumerate(self.pack_session.cards):
                if session_card.get('id') == card_id:
                    self.remove_card_from_session(i)
                    break
        
        # Layout: [- 2 +] [Remove]
        ttk.Button(qty_frame, text="-", width=2, command=decrease_qty).pack(side=tk.LEFT, padx=(0, 1))
        ttk.Label(qty_frame, text=str(current_qty), font=("Arial", 8, "bold"), width=2).pack(side=tk.LEFT, padx=1)
        ttk.Button(qty_frame, text="+", width=2, command=increase_qty).pack(side=tk.LEFT, padx=(1, 5))
        ttk.Button(qty_frame, text="🗑️", width=3, command=remove_card).pack(side=tk.RIGHT)
    
    def create_normal_mode_widget(self, card, index):
        """Create a single widget in normal mode"""
        print(f"[NORMAL MODE DEBUG] Creating widget for card {index}: {card.get('card_name', 'Unknown')}")
        
        # Verify scrollable frame exists
        if not hasattr(self, 'cards_scrollable_frame') or not self.cards_scrollable_frame:
            print(f"[NORMAL MODE DEBUG] ERROR: cards_scrollable_frame is None!")
            return
            
        print(f"[NORMAL MODE DEBUG] Scrollable frame exists: {self.cards_scrollable_frame}")
        
        # Configure single column
        self.cards_scrollable_frame.columnconfigure(0, weight=1)
        print(f"[NORMAL MODE DEBUG] Column 0 configured with weight=1")
        
        # Create card frame
        card_frame = ttk.Frame(self.cards_scrollable_frame, padding="10", relief="ridge")
        print(f"[NORMAL MODE DEBUG] Created card_frame: {card_frame}")
        
        # Grid the frame
        try:
            card_frame.grid(row=index, column=0, sticky=(tk.W, tk.E), pady=5, padx=5)
            print(f"[NORMAL MODE DEBUG] Successfully gridded card_frame at row={index}, col=0")
        except Exception as e:
            print(f"[NORMAL MODE DEBUG] ERROR gridding card_frame: {e}")
            return
            
        # Add to widgets list
        self.card_widgets.append(card_frame)
        print(f"[NORMAL MODE DEBUG] Added to card_widgets list. Total widgets: {len(self.card_widgets)}")
        print(f"[NORMAL MODE DEBUG] Card frame winfo_exists: {card_frame.winfo_exists()}")
        print(f"[NORMAL MODE DEBUG] Card frame winfo_ismapped: {card_frame.winfo_ismapped()}")
        
        # Configure grid
        card_frame.columnconfigure(1, weight=1)
        
        current_row = 0
        
        # Image display with safe synchronous loading
        if self.display_settings.get('show_images', True) and PIL_AVAILABLE:
            # Create image label
            image_label = ttk.Label(card_frame, text="🖼️\nLoading...", 
                                  font=("Arial", 8), justify="center", foreground="#666666")
            image_label.grid(row=0, column=0, rowspan=6, padx=(0, 15), pady=5)
            
            # Mark this as an image label for focus mode switching
            image_label._image_content = True
            
            # Load image synchronously to prevent segmentation faults
            self.load_card_image_safe(card, image_label)
            
        elif self.display_settings.get('show_images', True) and not PIL_AVAILABLE:
            # Show placeholder when PIL not available
            placeholder_label = ttk.Label(card_frame, text="🖼️\nNo PIL", font=("Arial", 10), justify="center")
            placeholder_label.grid(row=0, column=0, rowspan=6, padx=(0, 15), pady=5)
        elif self.display_settings.get('show_images', True):
            # Fallback for any other case
            placeholder_label = ttk.Label(card_frame, text="🃏\nCard", font=("Arial", 10), justify="center")
            placeholder_label.grid(row=0, column=0, rowspan=6, padx=(0, 15), pady=5)
        
        # Card details
        if self.display_settings.get('show_card_name', True):
            name_label = ttk.Label(card_frame, text=f"📋 {card.get('card_name', 'N/A')}", 
                                 font=("Arial", 12, "bold"))
            name_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            current_row += 1
        
        if self.display_settings.get('show_rarity', True):
            rarity_label = ttk.Label(card_frame, text=f"💎 Rarity: {card.get('card_rarity', 'N/A')}", 
                                   font=("Arial", 10))
            rarity_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            current_row += 1
        
        if self.display_settings.get('show_art_variant', True):
            variant_label = ttk.Label(card_frame, text=f"🎨 Art Variant: {card.get('art_variant', 'None')}", 
                                    font=("Arial", 10))
            variant_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            current_row += 1
        
        # Add price information
        current_row = self.add_price_labels(card_frame, card, current_row)
        
        # Add set information if enabled
        if self.display_settings.get('show_set_info', False):
            set_name = card.get('set_name', 'N/A')
            set_code = card.get('set_code', 'N/A')
            set_label = ttk.Label(card_frame, text=f"📚 Set: {set_name} ({set_code})", 
                                font=("Arial", 10))
            set_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            current_row += 1
        
        # Add timestamp if enabled
        if self.display_settings.get('show_timestamps', False):
            timestamp = card.get('timestamp', datetime.now().strftime("%H:%M:%S"))
            time_label = ttk.Label(card_frame, text=f"🕒 Added: {timestamp}", 
                                 font=("Arial", 10))
            time_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            current_row += 1
        
        # Quantity controls
        qty_frame = ttk.Frame(card_frame)
        qty_frame.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
        current_row += 1
        
        ttk.Label(qty_frame, text="📦 Quantity:", font=("Arial", 10)).pack(side=tk.LEFT)
        
        # Get current quantity (default to 1 if not set)
        current_qty = card.get('quantity', 1)
        
        # Quantity controls
        def decrease_qty():
            # Get current quantity from card data at execution time
            current_card_qty = self.pack_session.cards[index].get('quantity', 1) if index < len(self.pack_session.cards) else 1
            new_qty = max(1, current_card_qty - 1)  # Minimum quantity is 1
            self.update_card_quantity(index, new_qty)
        
        def increase_qty():
            # Get current quantity from card data at execution time
            current_card_qty = self.pack_session.cards[index].get('quantity', 1) if index < len(self.pack_session.cards) else 1
            new_qty = current_card_qty + 1
            self.update_card_quantity(index, new_qty)
        
        # Decrease button
        ttk.Button(qty_frame, text="-", width=3, command=decrease_qty).pack(side=tk.LEFT, padx=(5, 2))
        
        # Quantity display
        qty_label = ttk.Label(qty_frame, text=str(current_qty), font=("Arial", 10, "bold"), width=3)
        qty_label.pack(side=tk.LEFT, padx=2)
        
        # Increase button  
        ttk.Button(qty_frame, text="+", width=3, command=increase_qty).pack(side=tk.LEFT, padx=(2, 5))
        
        # Remove button
        def remove_card():
            self.remove_card_from_session(index)
        
        ttk.Button(qty_frame, text="🗑️ Remove", command=remove_card).pack(side=tk.LEFT, padx=(10, 0))
        
    def clear_widget_image_references(self, widget):
        """Recursively clear image references in widget tree"""
        try:
            # Clear image reference if it's a label with an image
            if hasattr(widget, 'cget') and hasattr(widget, 'configure'):
                try:
                    if widget.cget('image'):
                        widget.configure(image='')
                    if hasattr(widget, 'image'):
                        delattr(widget, 'image')
                except:
                    pass
            
            # Recursively clear children
            if hasattr(widget, 'winfo_children'):
                for child in widget.winfo_children():
                    self.clear_widget_image_references(child)
        except:
            pass
            
    def load_card_image_safe(self, card, image_label):
        """Load card image synchronously to prevent segmentation faults"""
        try:
            # Get card image URL
            card_images = card.get('card_images', [])
            if not card_images or len(card_images) == 0:
                print(f"[IMAGE DEBUG] No images available for {card.get('card_name', 'Unknown')}")
                self.set_image_placeholder(image_label, "🃏\nNo Image")
                return
            
            image_url = card_images[0].get('image_url')
            if not image_url:
                print(f"[IMAGE DEBUG] No image URL for {card.get('card_name', 'Unknown')}")
                self.set_image_placeholder(image_label, "🃏\nNo URL")
                return
            
            card_id = card.get('id', 'unknown')
            card_name = card.get('card_name', 'Unknown')
            
            # Use appropriate size based on current mode
            if self.focus_mode:
                display_size = self.image_manager.focus_mode_size
                print(f"[IMAGE DEBUG] Loading focus mode image for {card_name} (ID: {card_id})")
            else:
                display_size = self.image_manager.normal_mode_size
                print(f"[IMAGE DEBUG] Loading normal mode image for {card_name} (ID: {card_id})")
            
            # Try to get cached image first
            photo = self.image_manager.get_cached_image_for_mode(card_id, self.focus_mode)
            
            if not photo:
                # Load image with appropriate size
                photo = self.image_manager.load_image_for_display(card_id, image_url, display_size)
            
            if photo:
                # Verify widget still exists before updating
                if hasattr(image_label, 'winfo_exists') and image_label.winfo_exists():
                    print(f"[IMAGE DEBUG] Setting image for {card_name}")
                    image_label.configure(image=photo, text="")
                    image_label.image = photo  # Keep reference to prevent garbage collection
                else:
                    print(f"[IMAGE DEBUG] Widget no longer exists for {card_name}")
            else:
                print(f"[IMAGE DEBUG] Failed to load photo for {card_name}")
                self.set_image_placeholder(image_label, "🖼️\nError")
                
        except Exception as e:
            print(f"[IMAGE DEBUG] Error loading image for {card.get('card_name', 'Unknown')}: {e}")
            import traceback
            traceback.print_exc()
            self.set_image_placeholder(image_label, "❌\nError")
    
    def set_image_placeholder(self, image_label, text):
        """Safely set placeholder text on image label"""
        try:
            if hasattr(image_label, 'winfo_exists') and image_label.winfo_exists():
                image_label.configure(text=text, font=("Arial", 10), justify="center")
                # Clear any previous image reference
                if hasattr(image_label, 'image'):
                    delattr(image_label, 'image')
        except Exception as e:
            print(f"[IMAGE DEBUG] Error setting placeholder: {e}")
        
    def add_price_labels(self, parent, card, start_row):
        """Add price labels to card display"""
        current_row = start_row
        
        # TCG Low price
        if self.display_settings.get('show_tcg_price', True):
            tcg_price = card.get('tcg_price')
            if tcg_price and tcg_price not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
                price_text = f"💰 TCG Low: ${tcg_price}"
            elif tcg_price == '⏳ Loading...':
                price_text = "💰 ⏳ Loading TCG price..."
            else:
                price_text = "💰 TCG Low: N/A"
            
            price_label = ttk.Label(parent, text=price_text, font=("Arial", 10))
            price_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            current_row += 1
        
        # TCG Market price
        if self.display_settings.get('show_tcg_market_price', True):
            tcg_market_price = card.get('tcg_market_price')
            if tcg_market_price and tcg_market_price not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
                price_text = f"📈 TCG Market: ${tcg_market_price}"
            elif tcg_market_price == '⏳ Loading...':
                price_text = "📈 ⏳ Loading market price..."
            else:
                price_text = "📈 TCG Market: N/A"
            
            price_label = ttk.Label(parent, text=price_text, font=("Arial", 10))
            price_label.grid(row=current_row, column=1, sticky=(tk.W), pady=2)
            current_row += 1
            
        return current_row
            
    def get_compact_price_info(self, card):
        """Get compact price information for focus mode"""
        if card.get('tcg_price') and card.get('tcg_price') not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
            price_low = card['tcg_price']
            if card.get('tcg_market_price') and card.get('tcg_market_price') not in ['⏳ Loading...', 'Price unavailable', '❌ Error']:
                price_market = card['tcg_market_price']
                return f"L:${price_low} M:${price_market}"
            else:
                return f"Low: ${price_low}"
        elif card.get('tcg_price') == '⏳ Loading...':
            return "⏳ Loading..."
        else:
            return "No Price"
    
    def update_card_quantity(self, card_index, new_quantity):
        """Update the quantity of a card in the session with optimized UI updates"""
        if 0 <= card_index < len(self.pack_session.cards):
            self.pack_session.cards[card_index]['quantity'] = new_quantity
            print(f"[SESSION TRACKER] Updated card {card_index} quantity to {new_quantity}")
            
            # Try to update just the quantity label instead of rebuilding everything
            if self.update_quantity_label_only(card_index, new_quantity):
                print(f"[SESSION TRACKER] Updated quantity label in-place for card {card_index}")
            else:
                # Fallback to full refresh if in-place update fails
                print(f"[SESSION TRACKER] Falling back to full refresh for card {card_index}")
                self.safe_update_cards_display()
    
    def update_quantity_label_only(self, card_index, new_quantity):
        """Try to update only the quantity label for a specific card"""
        try:
            if card_index < len(self.card_widgets):
                widget = self.card_widgets[card_index]
                if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                    # Find the quantity label recursively
                    quantity_label = self.find_quantity_label_recursive(widget)
                    if quantity_label:
                        quantity_label.configure(text=str(new_quantity))
                        return True
            return False
        except Exception as e:
            print(f"[SESSION TRACKER] Error updating quantity label: {e}")
            return False
    
    def find_quantity_label_recursive(self, widget):
        """Recursively find the quantity label in a widget"""
        try:
            # Check if current widget is a quantity label (shows only a number)
            if hasattr(widget, 'cget') and hasattr(widget, 'configure'):
                text = widget.cget('text')
                if text and text.isdigit():
                    return widget
            
            # Check children recursively
            if hasattr(widget, 'winfo_children'):
                for child in widget.winfo_children():
                    result = self.find_quantity_label_recursive(child)
                    if result:
                        return result
            return None
        except:
            return None
    
    def remove_card_from_session(self, card_index):
        """Remove a card from the session"""
        if 0 <= card_index < len(self.pack_session.cards):
            removed_card = self.pack_session.cards.pop(card_index)
            print(f"[SESSION TRACKER] Removed card: {removed_card.get('card_name', 'Unknown')}")
            # Refresh display
            self.safe_update_cards_display()
    
    def auto_scroll_to_bottom(self):
        """Automatically scroll to the bottom of the card display with enhanced reliability"""
        print(f"[AUTO SCROLL DEBUG] auto_scroll_to_bottom called")
        try:
            if hasattr(self, 'cards_canvas') and self.cards_canvas:
                print(f"[AUTO SCROLL DEBUG] Canvas exists, scheduling scroll operation")
                # Schedule the scroll operation to happen after UI updates complete
                def scroll_to_bottom():
                    try:
                        print(f"[AUTO SCROLL DEBUG] Executing scroll_to_bottom")
                        
                        # Force final updates before scrolling
                        self.cards_canvas.update_idletasks()
                        self.cards_scrollable_frame.update_idletasks()
                        
                        # Check current scroll region
                        scroll_region = self.cards_canvas.cget('scrollregion')
                        print(f"[AUTO SCROLL DEBUG] Current scrollregion: {scroll_region}")
                        
                        # Verify we have a valid scroll region
                        if scroll_region and scroll_region != "0 0 0 0":
                            # Scroll to the very bottom
                            self.cards_canvas.yview_moveto(1.0)
                            print("[AUTO SCROLL DEBUG] Auto-scrolled to bottom (yview_moveto 1.0)")
                            
                            # Verify scroll position
                            current_view = self.cards_canvas.yview()
                            print(f"[AUTO SCROLL DEBUG] Current yview after scroll: {current_view}")
                            
                            # If we're not at the bottom, try alternative scroll methods
                            if current_view[1] < 0.99:  # Allow for small floating point differences
                                print(f"[AUTO SCROLL DEBUG] Not at bottom, trying yview_scroll to end")
                                self.cards_canvas.yview_scroll(1000, "units")  # Large scroll to ensure we reach bottom
                                
                                # Check again
                                final_view = self.cards_canvas.yview()
                                print(f"[AUTO SCROLL DEBUG] Final yview after scroll: {final_view}")
                        else:
                            print(f"[AUTO SCROLL DEBUG] Invalid scroll region, trying to recalculate")
                            # Try to recalculate scroll region
                            bbox = self.cards_canvas.bbox("all")
                            if bbox:
                                x1, y1, x2, y2 = bbox
                                expanded_bbox = (x1, y1, x2, y2 + 50)
                                self.cards_canvas.configure(scrollregion=expanded_bbox)
                                self.cards_canvas.yview_moveto(1.0)
                                print(f"[AUTO SCROLL DEBUG] Recalculated scroll region and scrolled to bottom")
                        
                    except Exception as e:
                        print(f"[AUTO SCROLL DEBUG] Error during auto-scroll: {e}")
                        import traceback
                        traceback.print_exc()
                
                # Schedule after a delay to ensure UI updates are complete
                print(f"[AUTO SCROLL DEBUG] Scheduling scroll_to_bottom after 200ms")
                self.window.after(200, scroll_to_bottom)  # Increased delay for better reliability
            else:
                print(f"[AUTO SCROLL DEBUG] No canvas available for scrolling")
        except Exception as e:
            print(f"[AUTO SCROLL DEBUG] Error in auto_scroll_to_bottom: {e}")
    
    def show_display_settings(self):
        """Show display settings dialog"""
        settings_window = tk.Toplevel(self.window)
        settings_window.title("Display Settings")
        settings_window.geometry("300x400")
        settings_window.resizable(False, False)
        
        # Make it modal
        settings_window.transient(self.window)
        settings_window.grab_set()
        
        main_frame = ttk.Frame(settings_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="Customize Card Display", font=("Arial", 12, "bold"))
        title_label.pack(pady=(0, 15))
        
        # Create checkboxes for each display setting
        setting_vars = {}
        
        settings_info = [
            ('show_images', '🖼️ Show Card Images'),
            ('show_card_name', '📋 Show Card Names'),
            ('show_rarity', '💎 Show Rarity'),
            ('show_art_variant', '🎨 Show Art Variant'),
            ('show_tcg_price', '💰 Show TCG Low Price'),
            ('show_tcg_market_price', '📈 Show TCG Market Price'),
            ('show_set_info', '📚 Show Set Information'),
            ('show_timestamps', '🕒 Show Timestamps')
        ]
        
        for setting_key, setting_label in settings_info:
            var = tk.BooleanVar(value=self.display_settings.get(setting_key, True))
            setting_vars[setting_key] = var
            
            checkbox = ttk.Checkbutton(main_frame, text=setting_label, variable=var)
            checkbox.pack(anchor=tk.W, pady=2)
        
        # Buttons frame
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill=tk.X, pady=(20, 0))
        
        def apply_settings():
            # Update display settings
            for setting_key, var in setting_vars.items():
                self.display_settings[setting_key] = var.get()
            
            # Refresh display
            self.safe_update_cards_display()
            settings_window.destroy()
        
        def cancel_settings():
            settings_window.destroy()
        
        # Apply and Cancel buttons
        ttk.Button(buttons_frame, text="Apply", command=apply_settings).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(buttons_frame, text="Cancel", command=cancel_settings).pack(side=tk.RIGHT)
        
        # Center the dialog
        settings_window.update_idletasks()
        x = (settings_window.winfo_screenwidth() // 2) - (300 // 2)
        y = (settings_window.winfo_screenheight() // 2) - (400 // 2)
        settings_window.geometry(f"300x400+{x}+{y}")
    
    def export_session_to_excel(self):
        """Export session cards to Excel file"""
        if not self.pack_session.cards:
            messagebox.showwarning("No Data", "No cards in current session to export!")
            return
        
        # Show field selection dialog
        self.show_session_export_dialog()
    
    def show_session_export_dialog(self):
        """Show dialog for selecting fields to export from session"""
        def export_selected():
            selected_fields = []
            for field, var in field_vars.items():
                if var.get():
                    selected_fields.append(field)
            
            if not selected_fields:
                messagebox.showwarning("No Fields", "Please select at least one field to export.")
                return
            
            dialog.destroy()
            self.perform_session_excel_export(selected_fields)
        
        # Create export dialog
        dialog = tk.Toplevel(self.window)
        dialog.title("Select Export Fields")
        dialog.geometry("350x450")
        dialog.transient(self.window)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Select fields to export:", font=("Arial", 12, "bold")).pack(pady=10)
        
        # Available fields for session export
        available_fields = {
            'card_name': 'Card Name',
            'card_rarity': 'Card Rarity',
            'quantity': 'Quantity',  # Add quantity field
            'art_variant': 'Art Variant',
            'tcg_price': 'TCGPlayer Low Price',
            'tcg_market_price': 'TCGPlayer Market Price',
            'set_code': 'Set Code',
            'booster_set_name': 'Set Name',
            'card_number': 'Card Number',
            'card_art_variant': 'Card Art Variant',
            'scrape_success': 'Scrape Success',
            'source_url': 'Source URL',
            'last_price_updt': 'Last Updated',
            'timestamp': 'Added Timestamp',
            'error_message': 'Error Message (if any)'
        }
        
        field_vars = {}
        for field, label in available_fields.items():
            var = tk.BooleanVar(value=True)  # Default to selected
            field_vars[field] = var
            ttk.Checkbutton(dialog, text=label, variable=var).pack(anchor=tk.W, padx=20, pady=2)
        
        # Buttons
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="📊 Export", command=export_selected).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)
    
    def perform_session_excel_export(self, selected_fields):
        """Perform the actual Excel export for session"""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Export Error", "openpyxl library is not available. Please install it to use Excel export functionality.")
            return
        
        # Create ExcelExports directory if it doesn't exist
        export_dir = os.path.join(os.getcwd(), "ExcelExports")
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
            print(f"[EXPORT] Created ExcelExports directory: {export_dir}")
        
        # Generate default filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        set_name = self.pack_session.current_set.get('set_name', 'Unknown_Set') if self.pack_session.current_set else 'Session'
        # Clean set name for filename
        safe_set_name = re.sub(r'[^\w\-_\.]', '_', set_name)
        default_filename = f"{safe_set_name}_export_{timestamp}.xlsx"
        default_path = os.path.join(export_dir, default_filename)
        
        # Ask for file save location, defaulting to ExcelExports folder
        file_path = filedialog.asksaveasfilename(
            initialdir=export_dir,
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Save Excel Export"
        )
        
        if not file_path:
            return
        
        try:
            # Create workbook
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Pack Session Cards"
            
            # Field labels
            field_labels = {
                'card_name': 'Card Name',
                'card_rarity': 'Card Rarity',
                'quantity': 'Quantity',  # Add quantity field
                'art_variant': 'Art Variant',
                'tcg_price': 'TCGPlayer Low Price',
                'tcg_market_price': 'TCGPlayer Market Price',
                'set_code': 'Set Code',
                'booster_set_name': 'Set Name',
                'card_number': 'Card Number',
                'card_art_variant': 'Card Art Variant',
                'scrape_success': 'Scrape Success',
                'source_url': 'Source URL',
                'last_price_updt': 'Last Updated',
                'timestamp': 'Added Timestamp',
                'error_message': 'Error Message (if any)'
            }
            
            # Write headers
            headers = [field_labels.get(field, field) for field in selected_fields]
            worksheet.append(headers)
            
            # Style headers
            for cell in worksheet["1:1"]:
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            
            # Write data
            print(f"[SESSION EXPORT DEBUG] Starting export of {len(self.pack_session.cards)} cards from session")
            row_count = 0
            for card in self.pack_session.cards:
                row_data = []
                for field in selected_fields:
                    value = card.get(field, "N/A")
                    if value is None:
                        value = "N/A"
                    row_data.append(value)
                worksheet.append(row_data)
                row_count += 1
                if row_count <= 5:  # Debug first 5 cards
                    print(f"[SESSION EXPORT DEBUG] Exported card {row_count}: {card.get("card_name", "Unknown")} - Qty: {card.get("quantity", "N/A")} - Rarity: {card.get("card_rarity", "N/A")}")
            
            print(f"[SESSION EXPORT DEBUG] Total cards exported: {row_count}")
            print(f"[SESSION EXPORT DEBUG] Session cards count: {len(self.pack_session.cards)}")
            
            # Auto-size columns
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(file_path)
            
            # Calculate total quantity for better messaging
            total_quantity = sum(card.get('quantity', 1) for card in self.pack_session.cards)
            
            messagebox.showinfo("Export Success", 
                               f"Excel file created successfully!\n\nFile: {file_path}\nCards exported: {len(self.pack_session.cards)} records\nTotal quantity: {total_quantity} cards")
            print(f"[SESSION EXPORT] Session exported to: {file_path}")
            print(f"[SESSION EXPORT] Records: {len(self.pack_session.cards)}, Total Quantity: {total_quantity}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to create Excel file:\n{str(e)}")
            print(f"[EXPORT] Export failed: {str(e)}")
            
    def close_window(self):
        """Close the session tracker window"""
        self.window.destroy()

class ImageManager:
    """Handles card image caching and display with enhanced safety features and performance optimizations"""
    def __init__(self):
        self.image_cache_dir = os.path.join(os.path.expanduser("~"), ".ygo_ripper_cache", "images")
        self.create_cache_directory()
        
        # Enhanced caching system
        self.image_cache = {}  # In-memory cache for loaded images
        self.focus_mode_cache = {}  # Separate cache for focus mode images
        self.normal_mode_cache = {}  # Separate cache for normal mode images
        
        self.threaded_loading_enabled = True  # Safety switch for threading
        self.max_cache_size = 1000  # Increased cache size for better performance
        self.loading_lock = threading.Lock()  # Thread safety for cache operations
        
        # Define standard sizes for different modes - reduced for better performance
        self.focus_mode_size = (60, 90)    # Smaller for focus mode - reduced by 25%
        self.normal_mode_size = (100, 145)  # Standard for normal mode - reduced by 33%
    
    def create_cache_directory(self):
        """Create image cache directory if it doesn't exist"""
        os.makedirs(self.image_cache_dir, exist_ok=True)
    
    def get_image_filename(self, card_id, image_url, size_suffix=""):
        """Generate a unique filename for caching based on card ID, URL and size"""
        # Create a hash of the URL to handle different image variants
        url_hash = hashlib.md5(image_url.encode()).hexdigest()[:8]
        if size_suffix:
            return f"card_{card_id}_{url_hash}_{size_suffix}.jpg"
        else:
            return f"card_{card_id}_{url_hash}.jpg"
    
    def get_cached_image_path(self, card_id, image_url, size_suffix=""):
        """Get the local path for a cached image with optional size suffix"""
        filename = self.get_image_filename(card_id, image_url, size_suffix)
        return os.path.join(self.image_cache_dir, filename)
    
    def download_and_cache_image(self, card_id, image_url, max_size=(200, 300)):
        """Download and cache a card image, return local path with enhanced error handling"""
        try:
            cache_path = self.get_cached_image_path(card_id, image_url)
            
            # If already cached, return the path
            if os.path.exists(cache_path):
                print(f"[IMAGE CACHE] Using existing cached image for card {card_id}")
                return cache_path
            
            # Download the image
            print(f"[IMAGE CACHE] Downloading image for card {card_id}: {image_url}")
            try:
                response = requests.get(image_url, timeout=10, stream=True)
                response.raise_for_status()
            except Exception as download_error:
                print(f"[IMAGE CACHE] Download failed for card {card_id}: {download_error}")
                return None
            
            # Save the original image
            temp_path = cache_path + ".tmp"
            try:
                with open(temp_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                print(f"[IMAGE CACHE] Downloaded image data for card {card_id}")
            except Exception as save_error:
                print(f"[IMAGE CACHE] Failed to save image data for card {card_id}: {save_error}")
                return None
            
            # Resize image to save space and standardize display
            if PIL_AVAILABLE:
                try:
                    print(f"[IMAGE CACHE] Processing image for card {card_id}")
                    with Image.open(temp_path) as img:
                        print(f"[IMAGE CACHE] Original image mode: {img.mode}, size: {img.size} for card {card_id}")
                        
                        # Convert to RGB if necessary
                        if img.mode in ('RGBA', 'LA', 'P'):
                            print(f"[IMAGE CACHE] Converting {img.mode} to RGB for card {card_id}")
                            img = img.convert('RGB')
                        
                        # Resize while maintaining aspect ratio
                        img.thumbnail(max_size, Image.Resampling.LANCZOS)
                        print(f"[IMAGE CACHE] Resized to {img.size} for card {card_id}")
                        
                        img.save(cache_path, 'JPEG', quality=85, optimize=True)
                        print(f"[IMAGE CACHE] Saved processed image for card {card_id}")
                except Exception as pil_error:
                    print(f"[IMAGE CACHE] PIL processing failed for card {card_id}: {pil_error}")
                    # Fallback: just copy the file
                    import shutil
                    try:
                        shutil.copy(temp_path, cache_path)
                        print(f"[IMAGE CACHE] Used fallback copy for card {card_id}")
                    except Exception as copy_error:
                        print(f"[IMAGE CACHE] Fallback copy failed for card {card_id}: {copy_error}")
                        return None
            else:
                # If PIL not available, just copy the file
                import shutil
                try:
                    shutil.copy(temp_path, cache_path)
                    print(f"[IMAGE CACHE] Copied without PIL for card {card_id}")
                except Exception as copy_error:
                    print(f"[IMAGE CACHE] Copy failed for card {card_id}: {copy_error}")
                    return None
            
            # Remove temporary file
            try:
                os.remove(temp_path)
            except Exception as cleanup_error:
                print(f"[IMAGE CACHE] Failed to cleanup temp file for card {card_id}: {cleanup_error}")
            
            print(f"[IMAGE CACHE] Successfully cached image: {cache_path}")
            return cache_path
            
        except Exception as e:
            print(f"[IMAGE CACHE] Error downloading image for card {card_id}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def load_image_for_display(self, card_id, image_url, display_size=(150, 220)):
        """Load image for Tkinter display, using pre-cached resized images to eliminate repeated resizing"""
        if not PIL_AVAILABLE:
            print(f"[IMAGE CACHE] PIL not available for card {card_id}")
            return None
            
        try:
            # Determine which cache to use based on size
            is_focus_mode = display_size == self.focus_mode_size
            is_normal_mode = display_size == self.normal_mode_size
            
            with self.loading_lock:  # Thread-safe cache operations
                # Check appropriate mode-specific cache first
                if is_focus_mode and card_id in self.focus_mode_cache:
                    # Reduce verbose logging for cached images
                    return self.focus_mode_cache[card_id]
                elif is_normal_mode and card_id in self.normal_mode_cache:
                    # Reduce verbose logging for cached images
                    return self.normal_mode_cache[card_id]
                
                # Fallback to general cache
                cache_key = f"{card_id}_{display_size}"
                if cache_key in self.image_cache:
                    # Reduce verbose logging - only log every 50th cache hit to avoid spam
                    if hasattr(self, '_cache_hit_counter'):
                        self._cache_hit_counter += 1
                    else:
                        self._cache_hit_counter = 1
                    
                    if self._cache_hit_counter % 50 == 0:
                        print(f"[IMAGE CACHE] Using general cached image for card {card_id} (hit #{self._cache_hit_counter})")
                    return self.image_cache[cache_key]
            
            # Generate size suffix for filename
            size_suffix = f"{display_size[0]}x{display_size[1]}"
            
            # Check if pre-resized image exists on disk
            resized_cache_path = self.get_cached_image_path(card_id, image_url, size_suffix)
            
            if os.path.exists(resized_cache_path):
                # Reduce verbose logging for disk cache hits
                try:
                    # Load pre-resized image directly
                    with Image.open(resized_cache_path) as img:
                        img_copy = img.copy()
                        if img_copy.mode not in ('RGB', 'RGBA'):
                            img_copy = img_copy.convert('RGB')
                        
                        photo = ImageTk.PhotoImage(img_copy)
                        
                        # Cache in memory
                        self.cache_image_in_memory(card_id, photo, display_size, is_focus_mode, is_normal_mode)
                        return photo
                        
                except Exception as e:
                    print(f"[IMAGE CACHE] Error loading pre-resized image for card {card_id}: {e}")
                    # Fall through to create resized version
            
            # Get original cached image path
            original_cache_path = self.get_cached_image_path(card_id, image_url)
            print(f"[IMAGE CACHE] Original cache path for card {card_id}: {original_cache_path}")
            
            # Download if not cached
            if not os.path.exists(original_cache_path):
                print(f"[IMAGE CACHE] Downloading image for card {card_id}")
                original_cache_path = self.download_and_cache_image(card_id, image_url)
                if not original_cache_path:
                    print(f"[IMAGE CACHE] Failed to download image for card {card_id}")
                    return None
            
            # Create and cache resized version
            print(f"[IMAGE CACHE] Creating resized version for card {card_id}")
            try:
                with Image.open(original_cache_path) as img:
                    print(f"[IMAGE CACHE] Image opened successfully for card {card_id}, mode: {img.mode}, size: {img.size}")
                    
                    # Create a copy to avoid issues with the context manager
                    img_copy = img.copy()
                    
                    # Ensure RGB mode for consistent handling
                    if img_copy.mode not in ('RGB', 'RGBA'):
                        print(f"[IMAGE CACHE] Converting image mode from {img_copy.mode} to RGB for card {card_id}")
                        img_copy = img_copy.convert('RGB')
                    
                    # Resize for display
                    img_copy.thumbnail(display_size, Image.Resampling.LANCZOS)
                    print(f"[IMAGE CACHE] Image resized to {img_copy.size} for card {card_id}")
                    
                    # Save the resized version to disk for future use
                    try:
                        img_copy.save(resized_cache_path, 'JPEG', quality=85, optimize=True)
                        print(f"[IMAGE CACHE] Saved resized image to {resized_cache_path}")
                    except Exception as save_error:
                        print(f"[IMAGE CACHE] Failed to save resized image for card {card_id}: {save_error}")
                    
                    # Convert to PhotoImage for Tkinter
                    print(f"[IMAGE CACHE] Converting to PhotoImage for card {card_id}")
                    photo = ImageTk.PhotoImage(img_copy)
                    print(f"[IMAGE CACHE] PhotoImage created successfully for card {card_id}")
                    
                    # Cache in memory
                    self.cache_image_in_memory(card_id, photo, display_size, is_focus_mode, is_normal_mode)
                    return photo
                    
            except Exception as img_error:
                print(f"[IMAGE CACHE] PIL operation failed for card {card_id}: {img_error}")
                import traceback
                traceback.print_exc()
                return None
                
        except Exception as e:
            print(f"[IMAGE CACHE] Error loading image for display card {card_id}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def cache_image_in_memory(self, card_id, photo, display_size, is_focus_mode, is_normal_mode):
        """Helper method to cache image in appropriate memory caches"""
        with self.loading_lock:
            # Store in mode-specific cache (reduce logging verbosity)
            if is_focus_mode:
                self.focus_mode_cache[card_id] = photo
            elif is_normal_mode:
                self.normal_mode_cache[card_id] = photo
            
            # Also store in general cache with size limit check
            cache_key = f"{card_id}_{display_size}"
            if len(self.image_cache) < self.max_cache_size:
                self.image_cache[cache_key] = photo
            else:
                # Only log cache full message occasionally to avoid spam
                if not hasattr(self, '_cache_full_logged') or not self._cache_full_logged:
                    print(f"[IMAGE CACHE] General cache full ({len(self.image_cache)}/{self.max_cache_size}), not caching image for card {card_id}")
                    self._cache_full_logged = True
                
                # Optionally clear some cache to make room
                if len(self.image_cache) >= self.max_cache_size * 1.2:  # 20% over limit
                    print(f"[IMAGE CACHE] Cache size exceeded limit significantly, clearing general memory cache")
                    self.clear_memory_cache()
                    self._cache_full_logged = False  # Reset logging flag after cleanup
    
    def get_cache_size(self):
        """Get the total size of the image cache in MB"""
        total_size = 0
        try:
            for filename in os.listdir(self.image_cache_dir):
                file_path = os.path.join(self.image_cache_dir, filename)
                if os.path.isfile(file_path):
                    total_size += os.path.getsize(file_path)
        except:
            pass
        return total_size / (1024 * 1024)  # Convert to MB
    
    def clear_cache(self):
        """Clear the image cache directory"""
        try:
            for filename in os.listdir(self.image_cache_dir):
                file_path = os.path.join(self.image_cache_dir, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
            self.image_cache.clear()
            print("[IMAGE CACHE] Cache cleared")
        except Exception as e:
            print(f"[IMAGE CACHE] Error clearing cache: {e}")
    
    def clear_memory_cache(self):
        """Clear only the in-memory image cache to reduce memory pressure"""
        try:
            cache_size = len(self.image_cache)
            focus_cache_size = len(self.focus_mode_cache)
            normal_cache_size = len(self.normal_mode_cache)
            
            self.image_cache.clear()
            # Don't clear mode-specific caches as they're more valuable for performance
            # self.focus_mode_cache.clear()
            # self.normal_mode_cache.clear()
            
            print(f"[IMAGE CACHE] General memory cache cleared - freed {cache_size} images")
            print(f"[IMAGE CACHE] Mode-specific caches preserved (focus: {focus_cache_size}, normal: {normal_cache_size})")
        except Exception as e:
            print(f"[IMAGE CACHE] Error clearing memory cache: {e}")
    
    def clear_mode_specific_caches(self):
        """Clear mode-specific caches when needed (e.g., low memory)"""
        try:
            focus_size = len(self.focus_mode_cache)
            normal_size = len(self.normal_mode_cache)
            
            self.focus_mode_cache.clear()
            self.normal_mode_cache.clear()
            
            print(f"[IMAGE CACHE] Mode-specific caches cleared - freed {focus_size + normal_size} images")
        except Exception as e:
            print(f"[IMAGE CACHE] Error clearing mode-specific caches: {e}")
    
    def get_cache_stats(self):
        """Get cache statistics for monitoring"""
        try:
            return {
                'general_cache_size': len(self.image_cache),
                'focus_cache_size': len(self.focus_mode_cache),
                'normal_cache_size': len(self.normal_mode_cache),
                'total_cached_images': len(self.image_cache) + len(self.focus_mode_cache) + len(self.normal_mode_cache)
            }
        except Exception:
            return {
                'general_cache_size': 0,
                'focus_cache_size': 0,
                'normal_cache_size': 0,
                'total_cached_images': 0
            }
    
    def preload_image_for_both_modes(self, card_id, image_url):
        """Pre-load image in both focus and normal mode sizes for fast switching - optimized version"""
        try:
            # Pre-load focus mode image (this will create resized cache on disk if needed)
            focus_photo = self.load_image_for_display(card_id, image_url, self.focus_mode_size)
            
            # Pre-load normal mode image (this will create resized cache on disk if needed)
            normal_photo = self.load_image_for_display(card_id, image_url, self.normal_mode_size)
                
            return focus_photo, normal_photo
        except Exception as e:
            print(f"[IMAGE CACHE] Error pre-loading images for card {card_id}: {e}")
            return None, None
    
    def get_image(self, card, focus_mode=False):
        """Main method to get card image for display with proper mode handling"""
        try:
            card_id = card.get('id', 'unknown')
            card_images = card.get('card_images', [])
            
            if not card_images or len(card_images) == 0:
                print(f"[IMAGE CACHE] No images available for card {card_id}")
                return None
                
            image_url = card_images[0].get('image_url')
            if not image_url:
                print(f"[IMAGE CACHE] No image URL for card {card_id}")
                return None
            
            # Get appropriate display size based on mode
            display_size = self.focus_mode_size if focus_mode else self.normal_mode_size
            
            # Try to get from mode-specific cache first
            cached_image = self.get_cached_image_for_mode(card_id, focus_mode)
            if cached_image:
                return cached_image
            
            # Load image for display with proper size
            return self.load_image_for_display(card_id, image_url, display_size)
            
        except Exception as e:
            print(f"[IMAGE CACHE] Error getting image for card {card.get('id', 'unknown')}: {e}")
            return None

    def is_image_fully_cached(self, card_id, focus_mode=False):
        """Check if image is fully cached (both in memory and on disk) for specified mode"""
        try:
            # Check memory cache first
            with self.loading_lock:
                if focus_mode and card_id in self.focus_mode_cache:
                    return True
                elif not focus_mode and card_id in self.normal_mode_cache:
                    return True
            
            # Check general memory cache
            display_size = self.focus_mode_size if focus_mode else self.normal_mode_size
            cache_key = f"{card_id}_{display_size}"
            with self.loading_lock:
                if cache_key in self.image_cache:
                    return True
            
            # For disk cache, we need to check the cache directory for any files matching the card_id and size pattern
            size_suffix = f"{display_size[0]}x{display_size[1]}"
            cache_pattern = f"card_{card_id}_*_{size_suffix}.jpg"
            
            if hasattr(self, 'cache_dir') and os.path.exists(self.cache_dir):
                import glob
                matching_files = glob.glob(os.path.join(self.cache_dir, "images", cache_pattern))
                if matching_files:
                    return True
            
            return False
            
        except Exception as e:
            print(f"[IMAGE CACHE] Error checking cache status for card {card_id}: {e}")
            return False

    def get_cached_image_for_mode(self, card_id, focus_mode=False):
        """Get cached image for specific mode without loading"""
        try:
            with self.loading_lock:
                if focus_mode and card_id in self.focus_mode_cache:
                    return self.focus_mode_cache[card_id]
                elif not focus_mode and card_id in self.normal_mode_cache:
                    return self.normal_mode_cache[card_id]
            return None
        except Exception as e:
            print(f"[IMAGE CACHE] Error getting cached image for card {card_id}: {e}")
            return None

def log_api_call(method, url, headers=None, body=None, response=None, error=None):
    """Comprehensive API call logging"""
    logger.info("=" * 80)
    logger.info(f"🌐 API CALL: {method.upper()} {url}")
    logger.info("=" * 80)
    
    # Log request details
    if headers:
        logger.info("📤 REQUEST HEADERS:")
        for key, value in headers.items():
            logger.info(f"  {key}: {value}")
    
    if body:
        logger.info("📤 REQUEST BODY:")
        if isinstance(body, dict):
            logger.info(json.dumps(body, indent=2))
        else:
            logger.info(str(body))
    
    # Log response details
    if response:
        logger.info(f"📥 RESPONSE STATUS: {response.status_code}")
        logger.info("📥 RESPONSE HEADERS:")
        for key, value in response.headers.items():
            logger.info(f"  {key}: {value}")
        
        try:
            response_data = response.json()
            logger.info("📥 RESPONSE BODY:")
            logger.info(json.dumps(response_data, indent=2))
        except:
            logger.info("📥 RESPONSE BODY (non-JSON):")
            logger.info(response.text[:1000] + "..." if len(response.text) > 1000 else response.text)
    
    # Log errors
    if error:
        logger.error(f"❌ ERROR: {str(error)}")
    
    logger.info("=" * 80)

class PackRipperSession:
    """Manages pack ripping session data"""
    def __init__(self):
        self.cards = []
        self.current_set = None
        self.set_cards = []
        self.session_file = "pack_session.json"
        self.auto_save_interval = 30  # seconds
        
    def add_card(self, card_data):
        """Add a card to the session"""
        self.cards.append({
            **card_data,
            'timestamp': datetime.now().isoformat()
        })
        
    def save_session(self):
        """Save current session to file"""
        try:
            session_data = {
                'cards': self.cards,
                'current_set': self.current_set,
                'set_cards': self.set_cards,
                'last_saved': datetime.now().isoformat()
            }
            with open(self.session_file, 'w') as f:
                json.dump(session_data, f, indent=2)
            return True
        except Exception as e:
            print(f"Error saving session: {e}")
            return False
            
    def load_session(self):
        """Load session from file"""
        try:
            if os.path.exists(self.session_file):
                with open(self.session_file, 'r') as f:
                    session_data = json.load(f)
                self.cards = session_data.get('cards', [])
                self.current_set = session_data.get('current_set')
                self.set_cards = session_data.get('set_cards', [])
                return True
        except Exception as e:
            print(f"Error loading session: {e}")
        return False

class VoiceRecognizer:
    """Handles voice recognition functionality"""
    def __init__(self):
        if not SPEECH_RECOGNITION_AVAILABLE:
            self.enabled = False
            return
        self.enabled = True
        self.recognizer = sr.Recognizer()
        self.microphone = sr.Microphone()
        self.is_listening = False
        self.audio_lock = threading.Lock()  # Prevent concurrent microphone access
        
        # Adjust for ambient noise
        with self.microphone as source:
            self.recognizer.adjust_for_ambient_noise(source)
        
        # Yu-Gi-Oh specific recognition improvements
        self.recognizer.energy_threshold = 300  # Lower threshold for better pickup
        self.recognizer.dynamic_energy_threshold = True
        self.recognizer.pause_threshold = 0.8  # Shorter pause threshold
    
    def listen_once(self, timeout=5):
        """Listen for a single voice command with YGO-specific improvements"""
        # Try to acquire audio lock - if we can't, another recognition is in progress
        if not self.audio_lock.acquire(blocking=False):
            print("[VOICE DEBUG] Audio source busy, skipping recognition")
            return None
            
        try:
            with self.microphone as source:
                # Increased phrase_time_limit and improved audio capture
                audio = self.recognizer.listen(source, timeout=timeout, phrase_time_limit=15)
            
            # Try multiple recognition approaches for better YGO card name detection
            recognized_texts = []
            
            # Primary: Google Speech Recognition
            try:
                text = self.recognizer.recognize_google(audio)
                recognized_texts.append(("google", text.lower().strip()))
                print(f"[VOICE DEBUG] Google recognized: '{text.lower().strip()}'")
            except:
                pass
            
            # Secondary: Google with language hints (sometimes helps with fantasy words)
            try:
                text = self.recognizer.recognize_google(audio, language='en-US', show_all=False)
                recognized_texts.append(("google_en", text.lower().strip()))
                print(f"[VOICE DEBUG] Google EN recognized: '{text.lower().strip()}'")
            except:
                pass
            
            # Return the longest/most complete recognition
            if recognized_texts:
                # Prefer the longest recognition as it's often more complete
                best_recognition = max(recognized_texts, key=lambda x: len(x[1]))
                print(f"[VOICE DEBUG] Best recognition: '{best_recognition[1]}' (method: {best_recognition[0]})")
                return best_recognition[1]
            
            return None
            
        except sr.WaitTimeoutError:
            return None
        except sr.UnknownValueError:
            return None
        except sr.RequestError as e:
            print(f"Speech recognition error: {e}")
            return None
        except Exception as e:
            print(f"[VOICE DEBUG] Unexpected audio error: {e}")
            return None
        finally:
            # Always release the lock
            self.audio_lock.release()

class YGORipperUI:
    def __init__(self, root):
        self.root = root
        self.root.title("YGORipperUI - Card Price Checker & Pack Ripper")
        
        # Get screen dimensions for dynamic main window sizing
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Calculate dynamic main window size (70% of screen)
        main_width = min(max(int(screen_width * 0.7), 1000), 1600)  # 70% screen width, min 1000, max 1600
        main_height = min(max(int(screen_height * 0.7), 800), 1200)  # 70% screen height, min 800, max 1200
        
        self.root.geometry(f"{main_width}x{main_height}")
        self.root.resizable(True, True)
        
        # Get API URL from environment variable
        self.api_url = os.getenv('API_URL', 'http://127.0.0.1:8081')
        
        # Initialize components
        self.pack_session = PackRipperSession()
        self.voice_recognizer = VoiceRecognizer()
        self.image_manager = ImageManager()  # Add image manager
        self.available_sets = []
        self.filtered_sets = []
        
        # Voice confirmation state
        self.pending_voice_confirmation = False
        self.pending_card_options = []
        self.pending_voice_data = {}
        self.voice_status_locked = False  # Prevent voice loop from overriding status messages
        
        # Focus mode and display settings
        self.focus_mode = False
        self.auto_rarity_enabled = True  # Enable auto rarity extraction by default
        self.auto_art_rarity_enabled = True  # Enable auto art rarity extraction by default
        
        # Session tracker window
        self.session_tracker = None
        
        self.create_widgets()
        self.setup_auto_save()
        
        # Auto-load card sets from cache on startup
        self.auto_load_card_sets()
        
    def auto_load_session(self):
        """Automatically load session on startup if it exists"""
        if self.pack_session.load_session():
            print(f"[SESSION] Auto-loaded session with {len(self.pack_session.cards)} cards")
            self.update_session_display()
        else:
            print("[SESSION] No saved session found or failed to load")
        
    def setup_auto_save(self):
        """Setup automatic session saving"""
        def auto_save():
            if hasattr(self, 'pack_session') and self.pack_session.cards:
                self.pack_session.save_session()
            self.root.after(30000, auto_save)  # Save every 30 seconds
        
        self.root.after(30000, auto_save)
        
    def create_widgets(self):
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create tabs
        self.create_price_checker_tab()
        self.create_pack_ripper_tab()
        
    def create_price_checker_tab(self):
        """Create the original price checker tab"""
        # Price Checker Tab
        price_frame = ttk.Frame(self.notebook)
        self.notebook.add(price_frame, text="🔍 Price Checker")
        
        # Main frame
        main_frame = ttk.Frame(price_frame, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        price_frame.columnconfigure(0, weight=1)
        price_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="YGORipperUI Price Checker", 
                               font=("Arial", 20, "bold"), foreground="#2E86AB")
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 5))
        
        subtitle_label = ttk.Label(main_frame, text="Yu-Gi-Oh Card Price Checker", 
                                  font=("Arial", 12), foreground="#666666")
        subtitle_label.grid(row=1, column=0, columnspan=2, pady=(0, 20))
        
        # API URL display
        api_frame = ttk.Frame(main_frame)
        api_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        ttk.Label(api_frame, text="API URL:", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky=tk.W)
        api_url_label = ttk.Label(api_frame, text=self.api_url, foreground="#0066CC", font=("Arial", 9))
        api_url_label.grid(row=0, column=1, sticky=tk.W, padx=(5, 0))
        
        # Input fields frame
        input_frame = ttk.LabelFrame(main_frame, text="Card Information", padding="15")
        input_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 20))
        input_frame.columnconfigure(1, weight=1)
        
        # Card Number (Required)
        ttk.Label(input_frame, text="Card Number *:", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.card_number_var = tk.StringVar()
        card_number_entry = ttk.Entry(input_frame, textvariable=self.card_number_var, width=35, font=("Arial", 10))
        card_number_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # Card Name
        ttk.Label(input_frame, text="Card Name:", font=("Arial", 10)).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.card_name_var = tk.StringVar()
        card_name_entry = ttk.Entry(input_frame, textvariable=self.card_name_var, width=35, font=("Arial", 10))
        card_name_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # Card Rarity (Required)
        ttk.Label(input_frame, text="Card Rarity *:", font=("Arial", 10)).grid(row=2, column=0, sticky=tk.W, pady=5)
        self.card_rarity_var = tk.StringVar()
        card_rarity_entry = ttk.Entry(input_frame, textvariable=self.card_rarity_var, width=35, font=("Arial", 10))
        card_rarity_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # Art Variant
        ttk.Label(input_frame, text="Art Variant:", font=("Arial", 10)).grid(row=3, column=0, sticky=tk.W, pady=5)
        self.art_variant_var = tk.StringVar()
        art_variant_entry = ttk.Entry(input_frame, textvariable=self.art_variant_var, width=35, font=("Arial", 10))
        art_variant_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # Force Refresh (Required)
        force_refresh_frame = ttk.Frame(input_frame)
        force_refresh_frame.grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=10)
        
        ttk.Label(force_refresh_frame, text="Force Refresh *:", font=("Arial", 10)).pack(side=tk.LEFT)
        self.force_refresh_var = tk.BooleanVar()
        force_refresh_check = ttk.Checkbutton(force_refresh_frame, variable=self.force_refresh_var)
        force_refresh_check.pack(side=tk.LEFT, padx=(10, 0))
        
        # Required fields note
        ttk.Label(input_frame, text="* Required fields", foreground="red", font=("Arial", 9, "italic")).grid(
            row=5, column=0, columnspan=2, sticky=tk.W, pady=(15, 0))
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=(0, 15))
        
        self.check_button = ttk.Button(button_frame, text="🔍 Check Price", command=self.check_price)
        self.check_button.pack(side=tk.LEFT, padx=(0, 15))
        
        clear_button = ttk.Button(button_frame, text="🗑 Clear", command=self.clear_fields)
        clear_button.pack(side=tk.LEFT)
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Results frame
        results_frame = ttk.LabelFrame(main_frame, text="Price Information", padding="15")
        results_frame.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(6, weight=1)
        
        # Results text area
        self.results_text = scrolledtext.ScrolledText(results_frame, height=18, width=80, wrap=tk.WORD, font=("Consolas", 10))
        self.results_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Initial message
        welcome_msg = """Welcome to YGORipperUI! 🃏

Enter the card information above and click "Check Price" to get pricing data.

Required fields:
• Card Number: The card's identification number
• Card Rarity: The rarity of the card
• Force Refresh: Check this to force a fresh data scrape

Optional fields:
• Card Name: The name of the card
• Art Variant: Specific art variant if applicable

The application will fetch pricing data from TCGPlayer and display:
• Card details (name, rarity, set)
• TCGPlayer Low and Market prices
• Last update timestamp
"""
        self.results_text.insert(tk.END, welcome_msg)
        
    def validate_inputs(self):
        """Validate required fields"""
        if not self.card_number_var.get().strip():
            messagebox.showerror("Validation Error", "Card Number is required!")
            return False
        if not self.card_rarity_var.get().strip():
            messagebox.showerror("Validation Error", "Card Rarity is required!")
            return False
        return True
    
    def prepare_payload(self):
        """Prepare the JSON payload for the API request"""
        return {
            "card_number": self.card_number_var.get().strip(),
            "card_name": self.card_name_var.get().strip(),
            "card_rarity": self.card_rarity_var.get().strip(),
            "art_variant": self.art_variant_var.get().strip(),
            "force_refresh": self.force_refresh_var.get()
        }
    
    def format_price_data(self, data):
        """Format the API response data for display"""
        if not data.get('success'):
            return f"❌ Error: {data.get('message', 'Unknown error occurred')}"
        
        card_data = data.get('data', {})
        
        # Header information
        result = "=" * 80 + "\n"
        result += "🃏 YGORIPPERUI - CARD PRICE INFORMATION\n"
        result += "=" * 80 + "\n\n"
        
        # Basic card info
        result += "📋 CARD DETAILS:\n"
        result += "-" * 40 + "\n"
        result += f"Name: {card_data.get('card_name', 'N/A')}\n"
        result += f"Number: {card_data.get('card_number', 'N/A')}\n"
        result += f"Rarity: {card_data.get('card_rarity', 'N/A')}\n"
        result += f"Set: {card_data.get('booster_set_name', 'N/A')}\n"
        result += f"Art Variant: {card_data.get('card_art_variant', 'N/A')}\n"
        result += f"Set Code: {card_data.get('set_code', 'N/A')}\n"
        result += f"Last Updated: {card_data.get('last_price_updt', 'N/A')}\n\n"
        
        # Price information
        result += "💰 PRICING INFORMATION:\n"
        result += "-" * 40 + "\n"
        
        # Collect all non-null prices
        prices = []
        
        # TCGPlayer prices (new v2 API format)
        if card_data.get('tcg_price') is not None:
            prices.append(f"🎯 TCGPlayer Low: ${card_data['tcg_price']}")
        if card_data.get('tcg_market_price') is not None:
            prices.append(f"📈 TCGPlayer Market: ${card_data['tcg_market_price']}")
        
        if prices:
            for price in prices:
                result += f"  {price}\n"
        else:
            result += "  ❌ No pricing data available\n"
        
        result += "\n"
        
        # Additional info
        result += "ℹ️  ADDITIONAL INFORMATION:\n"
        result += "-" * 40 + "\n"
        result += f"Scrape Success: {'✅ Yes' if card_data.get('scrape_success') else '❌ No'}\n"
        
        if card_data.get('source_url'):
            result += f"Source URL: {card_data['source_url']}\n"
        
        result += f"\nAPI Message: {data.get('message', 'N/A')}\n"
        result += f"Query Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        result += "\n" + "=" * 80
        
        return result
    
    def check_price_thread(self):
        """Run the API request in a separate thread"""
        url = None  # Initialize url variable
        try:
            payload = self.prepare_payload()
            url = f"{self.api_url}/cards/price"
            
            # Update UI to show loading
            self.root.after(0, lambda: self.check_button.config(state='disabled'))
            self.root.after(0, lambda: self.progress.start())
            self.root.after(0, lambda: self.results_text.delete(1.0, tk.END))
            self.root.after(0, lambda: self.results_text.insert(tk.END, "🔄 Fetching price data...\n\nPlease wait while we retrieve the latest pricing information.\n"))
            
            # Make API request
            response = requests.post(url, json=payload, timeout=30)
            
            # Log the API call
            log_api_call("POST", url, body=payload, response=response)
            
            response.raise_for_status()
            
            data = response.json()
            formatted_result = self.format_price_data(data)
            
            # Update UI with results
            self.root.after(0, lambda: self.results_text.delete(1.0, tk.END))
            self.root.after(0, lambda: self.results_text.insert(tk.END, formatted_result))
            
        except requests.exceptions.ConnectionError as e:
            # Log the error
            if url:
                log_api_call("POST", url, body=payload, error=e)
            error_msg = f"🚫 CONNECTION ERROR\n\nCould not connect to: {url}\n\nPossible solutions:\n• Check if the API server is running\n• Verify the API_URL environment variable is correct\n• Check your internet connection\n• Ensure the server is accessible"
            self.root.after(0, lambda: self.results_text.delete(1.0, tk.END))
            self.root.after(0, lambda: self.results_text.insert(tk.END, error_msg))
        except requests.exceptions.Timeout as e:
            # Log the error
            if url:
                log_api_call("POST", url, body=payload, error=e)
            error_msg = "⏰ TIMEOUT ERROR\n\nThe request took too long to complete.\nThe server might be experiencing high load.\nPlease try again in a few moments."
            self.root.after(0, lambda: self.results_text.delete(1.0, tk.END))
            self.root.after(0, lambda: self.results_text.insert(tk.END, error_msg))
        except requests.exceptions.HTTPError as e:
            # Log the error
            if url:
                log_api_call("POST", url, body=payload, response=response, error=e)
            error_msg = f"🔴 HTTP ERROR\n\nServer Error: {e}\nStatus Code: {response.status_code}\n\nThe server returned an error. Please check your input data and try again."
            self.root.after(0, lambda: self.results_text.delete(1.0, tk.END))
            self.root.after(0, lambda: self.results_text.insert(tk.END, error_msg))
        except Exception as e:
            # Log the error
            if url:
                log_api_call("POST", url, body=payload, error=e)
            error_msg = f"❌ UNEXPECTED ERROR\n\nAn unexpected error occurred:\n{str(e)}\n\nPlease try again or contact support if the problem persists."
            self.root.after(0, lambda: self.results_text.delete(1.0, tk.END))
            self.root.after(0, lambda: self.results_text.insert(tk.END, error_msg))
        finally:
            # Re-enable UI
            self.root.after(0, lambda: self.progress.stop())
            self.root.after(0, lambda: self.check_button.config(state='normal'))
    
    def check_price(self):
        """Handle the check price button click"""
        if not self.validate_inputs():
            return
        
        # Run API request in separate thread to prevent UI freezing
        thread = threading.Thread(target=self.check_price_thread)
        thread.daemon = True
        thread.start()
    
    def clear_fields(self):
        """Clear all input fields and results"""
        self.card_number_var.set("")
        self.card_name_var.set("")
        self.card_rarity_var.set("")
        self.art_variant_var.set("")
        self.force_refresh_var.set(False)
        self.results_text.delete(1.0, tk.END)
        
        # Show welcome message again
        welcome_msg = """Welcome back! 🃏

Enter new card information above and click "Check Price" to get pricing data.
"""
        self.results_text.insert(tk.END, welcome_msg)
    
    def create_pack_ripper_tab(self):
        """Create the new pack ripper tab"""
        # Pack Ripper Tab
        pack_frame = ttk.Frame(self.notebook)
        self.notebook.add(pack_frame, text="🎤 Pack Ripper")
        
        # Main frame
        main_frame = ttk.Frame(pack_frame, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        pack_frame.columnconfigure(0, weight=1)
        pack_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="YGORipperUI Pack Ripper", 
                               font=("Arial", 20, "bold"), foreground="#D2691E")
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 5))
        
        subtitle_label = ttk.Label(main_frame, text="Voice-Enabled Pack Opening Experience", 
                                  font=("Arial", 12), foreground="#666666")
        subtitle_label.grid(row=1, column=0, columnspan=2, pady=(0, 20))
        
        # Set Selection Frame
        set_frame = ttk.LabelFrame(main_frame, text="Set Selection", padding="15")
        set_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        set_frame.columnconfigure(1, weight=1)
        
        # Set search
        ttk.Label(set_frame, text="Search Set:", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.set_search_var = tk.StringVar()
        self.set_search_var.trace_add('write', self.on_set_search_change)
        set_search_entry = ttk.Entry(set_frame, textvariable=self.set_search_var, width=40, font=("Arial", 10))
        set_search_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        
        # Set dropdown
        ttk.Label(set_frame, text="Select Set:", font=("Arial", 10)).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.set_selection_var = tk.StringVar()
        self.set_combobox = ttk.Combobox(set_frame, textvariable=self.set_selection_var, 
                                        state="readonly", width=37, font=("Arial", 10))
        self.set_combobox.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
        self.set_combobox.bind('<<ComboboxSelected>>', self.on_set_selected)
        
        # Load sets button
        load_sets_btn = ttk.Button(set_frame, text="🔄 Load Sets", command=self.load_card_sets)
        load_sets_btn.grid(row=0, column=2, padx=(10, 0), pady=5)
        
        # Load cards button
        load_cards_btn = ttk.Button(set_frame, text="📋 Load Cards", command=self.load_set_cards)
        load_cards_btn.grid(row=1, column=2, padx=(10, 0), pady=5)
        
        # Session Controls Frame
        session_frame = ttk.LabelFrame(main_frame, text="Session Controls", padding="15")
        session_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Session buttons row 1
        session_row1 = ttk.Frame(session_frame)
        session_row1.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.start_session_btn = ttk.Button(session_row1, text="▶️ Start Pack Session", command=self.start_pack_session)
        self.start_session_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.end_session_btn = ttk.Button(session_row1, text="⏹️ End Session", command=self.end_pack_session, state=tk.DISABLED)
        self.end_session_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        save_session_btn = ttk.Button(session_row1, text="💾 Save Session", command=self.save_session_manual)
        save_session_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        load_session_btn = ttk.Button(session_row1, text="📂 Load Session", command=self.load_session_manual)
        load_session_btn.pack(side=tk.LEFT)
        
        # Session buttons row 2
        session_row2 = ttk.Frame(session_frame)
        session_row2.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E))
        
        # Auto-confirm checkbox and threshold
        self.auto_confirm_var = tk.BooleanVar()
        auto_confirm_check = ttk.Checkbutton(session_row2, text="Auto-confirm matches above", variable=self.auto_confirm_var)
        auto_confirm_check.pack(side=tk.LEFT, padx=(0, 5))
        
        # Auto-confirm threshold spinner
        self.auto_confirm_threshold_var = tk.StringVar(value="85")
        threshold_spinner = ttk.Spinbox(session_row2, from_=70, to=95, increment=5, width=5, 
                                       textvariable=self.auto_confirm_threshold_var)
        threshold_spinner.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(session_row2, text="% confidence").pack(side=tk.LEFT, padx=(0, 20))
        
        # Session status
        self.session_status_var = tk.StringVar(value="No active session")
        status_label = ttk.Label(session_row2, textvariable=self.session_status_var, foreground="#666666")
        status_label.pack(side=tk.LEFT)
        
        # Voice Controls Frame
        voice_frame = ttk.LabelFrame(main_frame, text="Voice Controls", padding="15")
        voice_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Voice controls row 1 - listening controls
        voice_row1 = ttk.Frame(voice_frame)
        voice_row1.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.listen_btn = ttk.Button(voice_row1, text="🎤 Start Listening", command=self.toggle_voice_listening, state=tk.DISABLED)
        self.listen_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.voice_status_var = tk.StringVar(value="Voice recognition ready")
        voice_status_label = ttk.Label(voice_row1, textvariable=self.voice_status_var, foreground="#666666")
        voice_status_label.pack(side=tk.LEFT)
        
        # Voice controls row 2 - auto extraction settings
        voice_row2 = ttk.Frame(voice_frame)
        voice_row2.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Auto rarity extraction checkbox
        self.auto_rarity_var = tk.BooleanVar(value=self.auto_rarity_enabled)
        auto_rarity_check = ttk.Checkbutton(voice_row2, text="Auto-extract rarity from voice", 
                                           variable=self.auto_rarity_var,
                                           command=self.update_auto_rarity_setting)
        auto_rarity_check.pack(side=tk.LEFT, padx=(0, 15))
        
        # Auto art rarity extraction checkbox
        self.auto_art_rarity_var = tk.BooleanVar(value=self.auto_art_rarity_enabled)
        auto_art_rarity_check = ttk.Checkbutton(voice_row2, text="Auto-extract art rarity from voice", 
                                               variable=self.auto_art_rarity_var,
                                               command=self.update_auto_art_rarity_setting)
        auto_art_rarity_check.pack(side=tk.LEFT, padx=(0, 15))
        
        # Voice controls row 3 - focus mode
        voice_row3 = ttk.Frame(voice_frame)
        voice_row3.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E))
        
        # Session Information Frame
        session_info_frame = ttk.LabelFrame(main_frame, text="Session Information", padding="15")
        session_info_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        session_info_frame.columnconfigure(0, weight=1)
        
        # Info text
        info_text = ("When you start a pack session, a dedicated session tracker window will open to display "
                    "your collected cards with images, prices, and details. The tracker window includes:\n\n"
                    "• Card images with automatic caching\n"
                    "• TCGPlayer pricing information\n"
                    "• Focus mode for compact card viewing\n"
                    "• Real-time updates as you add cards\n"
                    "• Export functionality for session data")
        
        info_label = ttk.Label(session_info_frame, text=info_text, wraplength=800, justify=tk.LEFT)
        info_label.pack(fill=tk.X)
        
        # Display Settings Frame (updated to include focus mode info)
        display_frame = ttk.LabelFrame(main_frame, text="Display Settings", padding="15")
        display_frame.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Display info row
        display_row = ttk.Frame(display_frame)
        display_row.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E))
        
        settings_btn = ttk.Button(display_row, text="⚙️ Customize Display", command=self.show_display_settings)
        settings_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        cache_info_btn = ttk.Button(display_row, text="🖼️ Image Cache Info", command=self.show_cache_info)
        cache_info_btn.pack(side=tk.LEFT, padx=(0, 10))

        # Export and Settings Frame
        export_frame = ttk.LabelFrame(main_frame, text="Export & Settings", padding="15")
        export_frame.grid(row=7, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        export_btn = ttk.Button(export_frame, text="📊 Export to Excel", command=self.export_to_excel)
        export_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Pack session state
        self.pack_session_active = False
        self.voice_listening = False

    def update_auto_rarity_setting(self):
        """Update auto rarity extraction setting"""
        self.auto_rarity_enabled = self.auto_rarity_var.get()
        print(f"[VOICE SETTINGS] Auto rarity extraction: {'Enabled' if self.auto_rarity_enabled else 'Disabled'}")
    
    def update_auto_art_rarity_setting(self):
        """Update auto art rarity extraction setting"""
        self.auto_art_rarity_enabled = self.auto_art_rarity_var.get()
        print(f"[VOICE SETTINGS] Auto art rarity extraction: {'Enabled' if self.auto_art_rarity_enabled else 'Disabled'}")
    
    # Pack Ripper Methods
    def on_set_search_change(self, *args):
        """Handle set search text change"""
        search_text = self.set_search_var.get().strip().lower()
        if not search_text:
            self.filtered_sets = self.available_sets
        else:
            self.filtered_sets = [s for s in self.available_sets if search_text in s['set_name'].lower()]
        
        self.update_set_combobox()
    
    def update_set_combobox(self):
        """Update the set selection combobox"""
        set_names = [s['set_name'] for s in self.filtered_sets]
        self.set_combobox['values'] = set_names
        if set_names:
            self.set_combobox.current(0)
        else:
            self.set_selection_var.set("")
    
    def on_set_selected(self, event):
        """Handle set selection from combobox"""
        selected_set = self.set_combobox.get()
        if not selected_set:
            self.pack_session.current_set = None
            self.pack_session.set_cards = []
            self.session_status_var.set("No active session")
            return
        
        # Find the set in the available sets
        set_info = next((s for s in self.filtered_sets if s['set_name'] == selected_set), None)
        if set_info:
            self.pack_session.current_set = set_info
    
    def load_card_sets(self):
        """Load card sets from the API"""
        def load_sets_thread():
            try:
                # Check if there's a search term
                search_term = self.set_search_var.get().strip()
                
                if search_term:
                    # Use search endpoint with search term
                    url = f"{self.api_url}/card-sets/search/{search_term}"
                else:
                    # Use cache endpoint to get all sets
                    url = f"{self.api_url}/card-sets/from-cache"
                
                response = requests.get(url, timeout=30)
                
                # Log the API call
                log_api_call("GET", url, response=response)
                
                response.raise_for_status()
                
                data = response.json()
                if data.get('success'):
                    sets_data = data.get('data', [])
                    self.available_sets = sets_data
                    self.filtered_sets = sets_data
                    
                    # Update UI on main thread
                    self.root.after(0, self.update_set_combobox)
                    if search_term:
                        self.root.after(0, lambda: messagebox.showinfo("Success", f"Found {len(sets_data)} card sets matching '{search_term}'!"))
                    else:
                        self.root.after(0, lambda: messagebox.showinfo("Success", f"Loaded {len(sets_data)} card sets from cache!"))
                else:
                    error_message = data.get('message', 'Failed to load sets')
                    self.root.after(0, lambda msg=error_message: messagebox.showerror("Error", msg))
            except requests.exceptions.RequestException as e:
                # Log the error
                log_api_call("GET", url, error=e)
                error_message = f"Failed to load sets:\n{str(e)}"
                self.root.after(0, lambda msg=error_message: messagebox.showerror("Request Error", msg))
        
        # Run in separate thread
        thread = threading.Thread(target=load_sets_thread)
        thread.daemon = True
        thread.start()
    
    def load_set_cards(self):
        """Load cards for the selected set"""
        if not self.pack_session.current_set:
            messagebox.showwarning("Warning", "Please select a set first!")
            return
        
        def load_cards_thread():
            try:
                set_name = self.pack_session.current_set.get('set_name')
                url = f"{self.api_url}/card-sets/{set_name}/cards"
                response = requests.get(url, timeout=30)
                
                # Log the API call
                log_api_call("GET", url, response=response)
                
                response.raise_for_status()
                
                data = response.json()
                if data.get('success'):
                    cards = data.get('data', [])
                    self.pack_session.set_cards = cards
                    
                    # Update UI on main thread
                    self.root.after(0, lambda: messagebox.showinfo("Success", f"Loaded {len(cards)} cards for {set_name}!"))
                else:
                    error_message = data.get('message', 'Failed to load cards')
                    self.root.after(0, lambda msg=error_message: messagebox.showerror("Error", msg))
            except requests.exceptions.RequestException as e:
                # Log the error
                log_api_call("GET", url, error=e)
                error_message = f"Failed to load cards:\n{str(e)}"
                self.root.after(0, lambda msg=error_message: messagebox.showerror("Request Error", msg))
        
        # Run in separate thread
        thread = threading.Thread(target=load_cards_thread)
        thread.daemon = True
        thread.start()
    
    def update_session_display(self):
        """Update the session tracker display if it exists"""
        print("[SESSION UPDATE DEBUG] update_session_display called")
        print(f"[SESSION UPDATE DEBUG] Called from thread: {threading.current_thread().name}")
        print(f"[SESSION UPDATE DEBUG] Session tracker exists: {self.session_tracker is not None}")
        
        if self.session_tracker:
            print("[SESSION UPDATE DEBUG] Session tracker exists, calling safe_update_cards_display")
            print(f"[SESSION UPDATE DEBUG] Session has {len(self.pack_session.cards)} cards")
            
            # CRITICAL FIX: Always schedule on main thread with multiple fallback mechanisms
            def do_update():
                try:
                    print("[SESSION UPDATE DEBUG] Executing UI update on main thread")
                    self.session_tracker.safe_update_cards_display()
                    print("[SESSION UPDATE DEBUG] UI update completed successfully")
                except Exception as e:
                    print(f"[SESSION UPDATE DEBUG] ERROR during UI update: {e}")
                    import traceback
                    traceback.print_exc()
            
            if threading.current_thread() == threading.main_thread():
                print("[SESSION UPDATE DEBUG] Already on main thread, calling directly")
                do_update()
            else:
                print("[SESSION UPDATE DEBUG] Not on main thread, scheduling with after()")
                # Use both after(0) and after_idle as fallbacks
                self.root.after(0, do_update)
                # Also schedule with after_idle as a backup
                self.root.after_idle(lambda: print("[SESSION UPDATE DEBUG] after_idle backup triggered"))
        else:
            print("[SESSION UPDATE DEBUG] No session tracker exists")
    
    def update_cards_display(self):
        """Legacy method - now redirects to session tracker"""
        self.update_session_display()
    
    def start_pack_session(self):
        """Start a new pack ripping session with dedicated tracker window"""
        if self.pack_session_active:
            messagebox.showwarning("Warning", "Pack session is already active!")
            return
        
        if not self.pack_session.current_set:
            messagebox.showerror("Error", "Please select a set first!")
            return
        
        if not self.pack_session.set_cards:
            messagebox.showerror("Error", "Please load cards for the selected set first!")
            return
        
        # Load existing session if available
        self.pack_session.load_session()
        
        # Create and show session tracker window
        self.session_tracker = SessionTrackerWindow(self.root, self.pack_session, self.image_manager)
        
        # Set up window close event
        def on_session_window_closing():
            if self.pack_session_active:
                if messagebox.askokcancel("Close Session", "Do you want to end the current pack session?"):
                    self.end_pack_session()
                else:
                    return  # Don't close the window
            else:
                self.session_tracker = None
                
        self.session_tracker.window.protocol("WM_DELETE_WINDOW", on_session_window_closing)
        
        self.pack_session_active = True
        self.session_status_var.set(f"Active session: {self.pack_session.current_set['set_name']}")
        self.start_session_btn.config(state=tk.DISABLED)
        self.end_session_btn.config(state=tk.NORMAL)
        self.listen_btn.config(state=tk.NORMAL)
        
        messagebox.showinfo("Session Started", 
                           f"Pack ripping session started for:\n{self.pack_session.current_set['set_name']}\n\n"
                           "A session tracker window has opened to display your cards.\n\n"
                           "Click 'Start Listening' to begin voice recognition.\n\n"
                           "Voice Tips:\n"
                           "• Speak card names clearly\n"
                           "• You can say 'Art Variant [number]' or 'Art Variant [name]'\n"
                           "• You can say 'Rarity [rarity name]'\n"
                           "• Say 'none' for no art variant")
    
    def end_pack_session(self):
        """End the pack ripping session and close tracker window"""
        if not self.pack_session_active:
            return
        
        self.pack_session_active = False
        self.voice_listening = False
        self.session_status_var.set("No active session")
        self.start_session_btn.config(state=tk.NORMAL)
        self.end_session_btn.config(state=tk.DISABLED)
        self.listen_btn.config(state=tk.DISABLED, text="🎤 Start Listening")
        self.voice_status_var.set("Voice recognition ready")
        
        # Save final session
        session_saved = self.pack_session.save_session()
        
        # Close session tracker window
        if self.session_tracker:
            self.session_tracker.close_window()
            self.session_tracker = None
        
        if session_saved:
            messagebox.showinfo("Session Ended", 
                               f"Pack session ended successfully!\n\n"
                               f"Cards collected: {len(self.pack_session.cards)}\n"
                               f"Session saved to: {self.pack_session.session_file}")
        else:
            messagebox.showwarning("Session Ended", "Session ended but failed to save data.")
    
    def toggle_voice_listening(self):
        """Toggle voice listening on/off"""
        if self.voice_listening:
            self.stop_voice_listening()
        else:
            self.start_voice_listening()
    
    def start_voice_listening(self):
        """Start voice listening in a separate thread"""
        if not self.pack_session_active:
            messagebox.showwarning("Warning", "Please start a pack session first!")
            return
        
        self.voice_listening = True
        self.listen_btn.config(text="🛑 Stop Listening")
        self.voice_status_var.set("🎤 Listening for card names...")
        
        # Start voice recognition in separate thread
        thread = threading.Thread(target=self.voice_recognition_loop)
        thread.daemon = True
        thread.start()
    
    def stop_voice_listening(self):
        """Stop voice listening"""
        self.voice_listening = False
        self.voice_status_locked = False  # Unlock status when stopping voice recognition
        self.listen_btn.config(text="🎤 Start Listening")
        self.voice_status_var.set("Voice recognition ready")
    
    def voice_recognition_loop(self):
        """Main voice recognition loop"""
        while self.voice_listening and self.pack_session_active:
            try:
                # Skip voice recognition if a dialog is pending confirmation
                # This prevents microphone context manager conflicts
                if self.pending_voice_confirmation:
                    if not self.voice_status_locked:  # Only update status if not locked
                        self.root.after(0, lambda: self.voice_status_var.set("🎤 Waiting for selection..."))
                    time.sleep(0.1)  # Short sleep while waiting for dialog response
                    continue
                
                # Skip voice recognition if status is locked (e.g., showing "Added:" message)
                if self.voice_status_locked:
                    time.sleep(0.1)  # Short sleep while status is locked
                    continue
                
                self.root.after(0, lambda: self.voice_status_var.set("🎤 Listening..."))
                
                # Listen for voice input
                voice_text = self.voice_recognizer.listen_once(timeout=3)
                
                if voice_text:
                    # Debug: Print what was heard to console and log
                    print(f"[VOICE DEBUG] Raw heard text: '{voice_text}'")
                    logger.info(f"Voice recognition captured: '{voice_text}'")
                    
                    # Lock status during processing to prevent overriding
                    self.voice_status_locked = True
                    
                    # Update UI with debug info and schedule status clear
                    self.root.after(0, lambda text=voice_text: self.voice_status_var.set(f"Heard: '{text}' (len: {len(text)})"))
                    # Clear the "Heard:" status after 2 seconds with high priority to prevent UI appearing stuck
                    self.root.after(2000, lambda: self.root.after_idle(self.clear_voice_status_if_heard))
                    
                    self.process_voice_input(voice_text)
                    time.sleep(0.5)  # Brief pause between recognitions
                else:
                    # Debug: Show when nothing was heard
                    print("[VOICE DEBUG] No speech detected or recognition failed")
                    logger.info("Voice recognition: No speech detected")
                    if not self.voice_status_locked:  # Only update if not locked
                        self.root.after(0, lambda: self.voice_status_var.set("🎤 Listening... (no speech detected)"))
                
            except Exception as e:
                print(f"[VOICE DEBUG] Voice recognition error: {e}")
                logger.error(f"Voice recognition error: {e}")
                if not self.voice_status_locked:  # Only update if not locked
                    self.root.after(0, lambda: self.voice_status_var.set("Voice recognition error"))
                time.sleep(1)
    
    def clear_voice_status_if_heard(self):
        """Clear voice status if it shows 'Heard:' to prevent UI appearing stuck"""
        try:
            current_status = self.voice_status_var.get()
            if current_status.startswith("Heard:"):
                print("[VOICE DEBUG] Clearing 'Heard:' status and unlocking voice status")
                if self.pending_voice_confirmation:
                    self.voice_status_var.set("🎤 Waiting for selection...")
                elif self.voice_listening:
                    self.voice_status_var.set("🎤 Listening...")
                else:
                    self.voice_status_var.set("Voice recognition ready")
                print("[VOICE DEBUG] Cleared stuck 'Heard:' status message")
            
            # Unlock status after clearing "Heard:" message
            self.voice_status_locked = False
            print("[VOICE DEBUG] Voice status unlocked after 'Heard:' clear - ready for new input")
            
        except Exception as e:
            print(f"[VOICE DEBUG] Error clearing voice status: {e}")
            # Always unlock status even if there's an error
            self.voice_status_locked = False
    
    def process_voice_input(self, voice_text):
        """Process voice input and extract card information"""
        voice_text = voice_text.lower().strip()
        
        # Check if we're waiting for a numbered selection
        if self.pending_voice_confirmation:
            # CRITICAL FIX: Schedule voice selection handling on main thread to fix UI issues
            self.root.after(0, lambda: self.handle_voice_selection(voice_text))
            return
        
        # Parse art variant and rarity from voice input (only if enabled)
        art_variant = None
        rarity = None
        card_name = voice_text
        
        # Extract art variant (if auto art rarity is enabled)
        if self.auto_art_rarity_enabled:
            art_patterns = [
                r'art variant (\w+)',
                r'art (\w+)',
                r'variant (\w+)',
                r'artwork (\w+)',
                r'art rarity (.+?)(?:\s|$)',
                r'art variant (.+?)(?:\s|$)'
            ]
            for pattern in art_patterns:
                match = re.search(pattern, voice_text)
                if match:
                    art_variant = match.group(1)
                    card_name = re.sub(pattern, '', card_name).strip()
                    print(f"[VOICE DEBUG] Auto-extracted art variant: '{art_variant}'")
                    break
        
        # Extract rarity (if auto rarity is enabled)
        if self.auto_rarity_enabled:
            # Enhanced rarity patterns to catch more YGO rarity types
            rarity_patterns = [
                r'quarter century secret rare',
                r'quarter century secret',
                r'prismatic secret rare',
                r'prismatic secret',
                r'starlight rare',
                r'collector.*?rare',
                r'ghost rare',
                r'secret rare',
                r'ultra rare',
                r'super rare',
                r'rare',
                r'common',
                r'rarity (.+?)(?:\s|$)',
                r'rare (.+?)(?:\s|$)',
                r'(.+?) rare(?:\s|$)',
                r'(.+?) rarity(?:\s|$)'
            ]
            for pattern in rarity_patterns:
                match = re.search(pattern, voice_text)
                if match:
                    if pattern in [r'quarter century secret rare', r'quarter century secret', 
                                 r'prismatic secret rare', r'prismatic secret', r'starlight rare',
                                 r'collector.*?rare', r'ghost rare', r'secret rare', 
                                 r'ultra rare', r'super rare', r'rare', r'common']:
                        # Direct match patterns
                        rarity = match.group(0).strip()
                    else:
                        # Group-based patterns
                        rarity = match.group(1).strip()
                    
                    card_name = re.sub(pattern, '', card_name).strip()
                    print(f"[VOICE DEBUG] Auto-extracted rarity: '{rarity}'")
                    break
        
        # Clean up card name
        card_name = re.sub(r'\s+', ' ', card_name).strip()
        
        print(f"[VOICE DEBUG] Processed - Card: '{card_name}', Rarity: '{rarity}', Art: '{art_variant}'")
        
        # Find matching cards
        self.find_and_present_card_options(card_name, art_variant, rarity)
    
    def handle_voice_selection(self, voice_text):
        """Handle numbered selection from voice input"""
        # Check for numbered selections (e.g., "1", "option 1", "select 1", etc.)
        number_patterns = [
            r'^\s*(\d+)\s*$',
            r'option\s+(\d+)',
            r'select\s+(\d+)',
            r'choose\s+(\d+)',
            r'number\s+(\d+)'
        ]
        
        for pattern in number_patterns:
            match = re.search(pattern, voice_text)
            if match:
                selection_index = int(match.group(1)) - 1  # Convert to 0-based index
                if 0 <= selection_index < len(self.pending_card_options):
                    # Get selected option
                    selected_option = self.pending_card_options[selection_index]
                    
                    # Clear pending state
                    self.pending_voice_confirmation = False
                    self.pending_card_options = []
                    
                    # Add the selected card using the SPECIFIC rarity from the selected option
                    self.add_card_to_session(
                        selected_option['card_data'],
                        self.pending_voice_data.get('art_variant'),
                        selected_option['rarity']  # Use the specific rarity from the selected option
                    )
                    self.pending_voice_data = {}
                    return
                else:
                    self.root.after(0, lambda: self.voice_status_var.set(f"Invalid selection. Please say 1-{len(self.pending_card_options)}"))
                    return
        
        # Check for reject/cancel commands
        reject_patterns = [
            r'reject', r'cancel', r'no', r'none', r'skip'
        ]
        
        for pattern in reject_patterns:
            if re.search(pattern, voice_text):
                self.pending_voice_confirmation = False
                self.pending_card_options = []
                self.pending_voice_data = {}
                self.root.after(0, lambda: self.voice_status_var.set("Selection cancelled, continue speaking..."))
                return
        
        # If we get here, the voice input wasn't recognized
        self.root.after(0, lambda: self.voice_status_var.set(f"Say a number 1-{len(self.pending_card_options)} or 'cancel'"))
    
    def find_and_present_card_options(self, card_name, art_variant=None, rarity=None):
        """Find multiple matching cards and present options to user"""
        if not self.pack_session.set_cards:
            return
        
        # Debug: Print search parameters
        print(f"[CARD SEARCH DEBUG] Searching for: '{card_name}', rarity: '{rarity}', art_variant: '{art_variant}'")
        print(f"[CARD SEARCH DEBUG] Total cards in set: {len(self.pack_session.set_cards)}")
        
        # Enhanced matching for YGO fantasy names with improved lenient search
        card_matches = []
        
        # Implement both normal and lenient search modes
        def get_card_name_variants(name):
            """Generate alternative spellings and pronunciations for YGO cards"""
            variants = [name]
            
            # Common YGO card name patterns and substitutions
            substitutions = {
                'yu': ['you', 'u'],
                'gi': ['gee', 'ji'],
                'oh': ['o'],
                'elemental': ['elemental', 'element'],
                'hero': ['hiro', 'heero', 'hero'],
                'evil': ['evil', 'evel'],
                'dark': ['dark', 'drak'],
                'gaia': ['gaia', 'gaya', 'guy', 'gya'],
                'cyber': ['siber', 'cyber'],
                'dragon': ['drago', 'drag'],
                'magician': ['magic', 'mage'],
                'warrior': ['war', 'warrior'],
                'machine': ['mach', 'machin'],
                'beast': ['best', 'beast'],
                'fiend': ['fend', 'fiend'],
                'spellcaster': ['spell', 'caster'],
                'aqua': ['agua', 'aqua'],
                'winged': ['wing', 'winged'],
                'thunder': ['under', 'thunder'],
                'zombie': ['zomb', 'zombie'],
                'plant': ['plan', 'plant'],
                'insect': ['insec', 'insect'],
                'rock': ['rok', 'rock'],
                'pyro': ['fire', 'pyro'],
                'sea': ['see', 'sea'],
                'divine': ['divin', 'divine'],
                # Add compound word handling
                'metal flame': ['metalflame', 'metal flame'],
                'flame': ['flame', 'flam'],
                'metal': ['metal', 'mettle']
            }
            
            # Create phonetic alternatives
            lower_name = name.lower()
            for original, alternatives in substitutions.items():
                if original in lower_name:
                    for alt in alternatives:
                        variants.append(lower_name.replace(original, alt))
            
            # Add compound word variants (for cases like "metal flame" -> "metalflame")
            words = lower_name.split()
            if len(words) >= 2:
                # Add version with spaces removed
                variants.append(''.join(words))
                # Add version with different spacing
                for i in range(1, len(words)):
                    compound_variant = ''.join(words[:i]) + ' ' + ' '.join(words[i:])
                    variants.append(compound_variant)
            
            # Remove duplicates while preserving order
            seen = set()
            unique_variants = []
            for variant in variants:
                if variant.lower() not in seen:
                    seen.add(variant.lower())
                    unique_variants.append(variant)
            
            return unique_variants
        
        # Get card name variants for better matching
        search_variants = get_card_name_variants(card_name)
        print(f"[CARD SEARCH DEBUG] Generated search variants: {search_variants}")
        print(f"[CARD SEARCH DEBUG] Using rarity filter: '{rarity}'")
        
        for card in self.pack_session.set_cards:
            try:
                # Simplified and improved confidence calculation
                # Calculate name score using 3 key methods
                name_score = self.calculate_name_confidence(card_name, card['name'], search_variants)
                
                # Calculate rarity score (only if rarity was specified)
                rarity_score = 0
                if rarity:
                    rarity_score = self.calculate_rarity_confidence(rarity, card)
            except Exception as e:
                print(f"[NAME MATCH ERROR] Error processing card '{card.get('name', 'Unknown')}': {e}")
                continue
            
            # Weight card name more heavily than rarity (75% name + 25% rarity)
            # Getting the correct card is more important than identifying the correct rarity
            if rarity:
                # When rarity is specified, 75% name + 25% rarity
                total_score = (name_score * 0.75) + (rarity_score * 0.25)
                
                # Require minimum name score to prevent poor card matches
                # being boosted by perfect rarity matches
                if name_score < 40:
                    total_score = total_score * 0.5  # Heavily penalize poor name matches
            else:
                # When no rarity specified, 100% name matching
                total_score = name_score
            
            # Debug output for matches above threshold
            if total_score >= 50:
                print(f"[CARD SEARCH DEBUG] Match: '{card['name']}' (Name: {name_score:.1f}%, Rarity: {rarity_score:.1f}%, Total: {total_score:.1f}%)")
                card_sets = card.get('card_sets', [])
                for card_set in card_sets[:2]:  # Show first 2 rarities
                    print(f"  - Available: {card_set.get('set_rarity', 'N/A')}")
            
            if total_score >= 50:  # Lowered threshold since name is now weighted more heavily
                # Check if this is the same card name we already have
                existing_match = next((m for m in card_matches if m['name'] == card['name']), None)
                
                if existing_match:
                    # If we found a better scoring version of the same card, replace it
                    if total_score > existing_match['confidence']:
                        card_matches.remove(existing_match)
                        card_matches.append({
                            'card_data': card,
                            'confidence': total_score,
                            'name': card['name'],
                            'rarity': self.get_card_rarity_display(card, rarity)
                        })
                else:
                    # Add new card match
                    card_matches.append({
                        'card_data': card,
                        'confidence': total_score,
                        'name': card['name'],
                        'rarity': self.get_card_rarity_display(card, rarity)
                    })
        
        if not card_matches:
            print("[CARD SEARCH DEBUG] No matches found above 50% threshold")
            self.root.after(0, lambda: self.voice_status_var.set(f"No matches found for: {card_name}"))
            return
        
        # Now find ALL variants of the matched cards in the current set
        all_variants = []
        exact_name_matches = []  # Track exact name matches for prioritization
        
        for match in card_matches:
            card_name_to_find = match['name']
            print(f"[CARD SEARCH DEBUG] Finding all variants for: '{card_name_to_find}'")
            
            # Check if this is an exact or very close match to what the user said
            is_priority_match = (
                fuzz.token_set_ratio(card_name.lower(), card_name_to_find.lower()) >= 90 or
                card_name_to_find.lower() in card_name.lower() or
                card_name.lower() in card_name_to_find.lower()
            )
            
            # Find all cards with the same name but different rarities/variants
            for card in self.pack_session.set_cards:
                if card['name'] == card_name_to_find:
                    card_sets = card.get('card_sets', [])
                    for card_set in card_sets:
                        # Create a unique variant for each rarity
                        variant_key = f"{card['name']}_{card_set.get('set_rarity', 'Unknown')}_{card_set.get('set_code', 'N/A')}"
                        
                        # Check if we already added this exact variant
                        if not any(v.get('variant_key') == variant_key for v in all_variants):
                            # Recalculate confidence using the new simplified method
                            name_score = self.calculate_name_confidence(card_name, card['name'], search_variants)
                            
                            # Calculate rarity score for this specific variant
                            rarity_score = 0
                            if rarity:
                                # Direct rarity comparison for this variant
                                set_rarity = card_set.get('set_rarity', '').lower().strip()
                                input_rarity = rarity.lower().strip()
                                
                                if set_rarity == input_rarity:
                                    rarity_score = 100  # Perfect match
                                elif input_rarity in set_rarity or set_rarity in input_rarity:
                                    rarity_score = 80   # Good partial match
                                else:
                                    rarity_score = fuzz.ratio(input_rarity, set_rarity) * 0.7  # Fuzzy match scaled down
                            
                            # Weight card name more heavily than rarity (75% name + 25% rarity)
                            # Getting the correct card is more important than identifying the correct rarity
                            if rarity:
                                confidence = (name_score * 0.75) + (rarity_score * 0.25)
                                
                                # Require minimum name score to prevent poor card matches
                                # being boosted by perfect rarity matches
                                if name_score < 40:
                                    confidence = confidence * 0.5  # Heavily penalize poor name matches
                            else:
                                confidence = name_score
                            
                            # Apply caps to prevent over-confidence
                            if rarity and rarity_score == 100:  # Exact rarity match
                                confidence = min(95, confidence)
                            else:
                                confidence = min(85, confidence)
                            
                            variant = {
                                'card_data': card,
                                'confidence': confidence,
                                'name': card['name'],
                                'rarity': card_set.get('set_rarity', 'Unknown'),
                                'set_code': card_set.get('set_code', 'N/A'),
                                'variant_key': variant_key,
                                'is_priority': is_priority_match
                            }
                            
                            # Debug output
                            rarity_debug = f"(Name: {name_score:.1f}%, Rarity: {rarity_score:.1f}%)" if rarity else f"(Name: {name_score:.1f}%)"
                            
                            if is_priority_match:
                                exact_name_matches.append(variant)
                                print(f"[CARD SEARCH DEBUG] Priority variant: {card['name']} - {card_set.get('set_rarity')} ({confidence:.1f}%) {rarity_debug}")
                            else:
                                all_variants.append(variant)
                                print(f"[CARD SEARCH DEBUG] Added variant: {card['name']} - {card_set.get('set_rarity')} ({confidence:.1f}%) {rarity_debug}")
        
        # Combine priority matches first, then other variants
        final_variants = exact_name_matches + all_variants
        
        if not final_variants:
            print("[CARD SEARCH DEBUG] No variants found")
            self.root.after(0, lambda: self.voice_status_var.set(f"No variants found for: {card_name}"))
            return
        
        # Sort by confidence level (highest first) as requested by user
        # Apply bonuses for exact rarity matches and priority, but sort purely by final confidence
        def sort_key(variant):
            confidence = variant['confidence']
            # Sort by confidence only (highest first)
            return -confidence
        
        final_variants.sort(key=sort_key)
        
        # Ensure unique confidence scores to avoid ties
        self.ensure_unique_confidence_scores(final_variants)
        
        # Limit to top 8 variants (increased from 5) to capture more options
        final_variants = final_variants[:8]
        
        print(f"[CARD SEARCH DEBUG] Final variants to present: {len(final_variants)}")
        for i, variant in enumerate(final_variants, 1):
            priority_flag = "⭐ PRIORITY" if variant.get('is_priority') else ""
            # Check for exact rarity match
            exact_rarity_flag = ""
            if rarity:
                variant_rarity = variant.get('rarity', '').lower().strip()
                input_rarity = rarity.lower().strip()
                if variant_rarity == input_rarity:
                    exact_rarity_flag = "🎯 EXACT RARITY"
            
            print(f"  {i}. {variant['name']} - {variant['rarity']} ({variant['confidence']:.1f}%) {priority_flag} {exact_rarity_flag}")
        
        # Get auto-confirm threshold
        try:
            auto_threshold = int(self.auto_confirm_threshold_var.get())
        except ValueError:
            auto_threshold = 85
        
        # Debug auto-confirm state with explicit checkbox checking
        auto_confirm_enabled = self.auto_confirm_var.get()
        best_match = final_variants[0]
        
        # Additional debug: check the actual widget state
        print(f"[AUTO-CONFIRM DEBUG] BooleanVar.get(): {auto_confirm_enabled}")
        print(f"[AUTO-CONFIRM DEBUG] Threshold: {auto_threshold}%, Best match: {best_match['name']} ({best_match['confidence']:.1f}%)")
        print(f"[AUTO-CONFIRM DEBUG] Would auto-confirm: {auto_confirm_enabled and best_match['confidence'] >= auto_threshold}")
        
        # Check for auto-confirm - FIXED: Properly check the state
        if auto_confirm_enabled and best_match['confidence'] >= auto_threshold:
            # Auto-confirm the best match
            print(f"[AUTO-CONFIRM DEBUG] Auto-confirming {best_match['name']} - {best_match['rarity']}")
            # Lock status to prevent voice loop interference
            self.voice_status_locked = True
            
            # CRITICAL FIX: Schedule card addition on main thread to fix UI update issues
            def add_card_on_main_thread():
                print(f"[AUTO-CONFIRM DEBUG] Adding card on main thread")
                self.add_card_to_session(best_match['card_data'], art_variant, best_match['rarity'])
            
            self.root.after(0, add_card_on_main_thread)
            self.root.after(0, lambda: self.voice_status_var.set(f"Auto-confirmed: {best_match['name']} - {best_match['rarity']} ({best_match['confidence']}%)"))
            
            # Clear the auto-confirmed status after 2 seconds and return to listening mode
            def clear_auto_confirm_status():
                print("[VOICE DEBUG] Clearing auto-confirmed status and unlocking voice status")
                if self.voice_listening:
                    self.voice_status_var.set("🎤 Listening...")
                else:
                    self.voice_status_var.set("Voice recognition ready")
                # Unlock status to allow voice loop to continue
                self.voice_status_locked = False
                print("[VOICE DEBUG] Voice status unlocked after auto-confirm - ready for new input")
            
            # Use after_idle for higher priority to ensure voice status clearing is not blocked by other UI updates
            self.root.after(2000, lambda: self.root.after_idle(clear_auto_confirm_status))  # Clear after 2 seconds with high priority
            return
        
        # Present options to user
        print(f"[VOICE DEBUG] Showing dialog with {len(final_variants)} options to user")
        self.show_voice_card_options(final_variants, card_name, art_variant, rarity)
    
    def calculate_name_confidence(self, input_name, card_name, search_variants):
        """Calculate confidence score for name matching using simplified approach"""
        scores = []
        
        # Method 1: Best fuzzy match across all variants
        best_fuzzy = 0
        for variant in search_variants:
            # Use token_set_ratio as it handles word order differences well
            score = fuzz.token_set_ratio(variant.lower(), card_name.lower())
            best_fuzzy = max(best_fuzzy, score)
        scores.append(best_fuzzy)
        
        # Method 2: Enhanced substring/word matching for fantasy names
        clean_input = re.sub(r'[^\w\s]', '', input_name.lower())
        clean_card = re.sub(r'[^\w\s]', '', card_name.lower())
        
        input_words = clean_input.split()
        card_words = clean_card.split()
        
        # Calculate percentage of input words that have good matches in card name
        word_match_score = 0
        if input_words:
            matched_words = 0
            for input_word in input_words:
                if len(input_word) >= 2:  # Check meaningful words
                    best_word_match = 0
                    for card_word in card_words:
                        # Check for exact substring match first, but only for meaningful words
                        # Avoid false positives from very short words like "a", "i", "of", "the"
                        if len(card_word) >= 3 and len(input_word) >= 3:
                            if input_word in card_word or card_word in input_word:
                                best_word_match = 100
                                break
                        # Check fuzzy similarity for all words
                        word_similarity = fuzz.ratio(input_word, card_word)
                        best_word_match = max(best_word_match, word_similarity)
                    
                    # Consider it a match if similarity is high enough
                    if best_word_match >= 80:
                        matched_words += 1
            
            word_match_score = (matched_words / len(input_words)) * 100
        
        scores.append(word_match_score)
        
        # Method 3: Special handling for compound words (like "Metal flame" -> "Metalflame")
        # Remove spaces and check direct similarity
        no_space_input = clean_input.replace(' ', '')
        no_space_card = clean_card.replace(' ', '')
        compound_score = fuzz.ratio(no_space_input, no_space_card)
        scores.append(compound_score)
        
        # Take the best score from all methods
        raw_score = max(scores)
        
        # Apply length penalty for significantly different lengths
        input_len = len(clean_input.split())
        card_len = len(clean_card.split())
        
        # If input is much shorter than card name, apply penalty to avoid over-matching
        if input_len > 0 and card_len > 0:
            length_ratio = min(input_len, card_len) / max(input_len, card_len)
            if length_ratio < 0.5:  # Significant length difference
                length_penalty = 0.8  # 20% penalty
            elif length_ratio < 0.7:  # Moderate length difference
                length_penalty = 0.9  # 10% penalty
            else:
                length_penalty = 1.0  # No penalty
            
            final_score = raw_score * length_penalty
        else:
            final_score = raw_score
        
        # Debug output for significant matches
        if final_score >= 70:
            penalty_info = f" (penalty: {(1-length_penalty)*100:.0f}%)" if length_penalty < 1.0 else ""
            print(f"[NAME MATCH DEBUG] '{input_name}' -> '{card_name}': {final_score:.1f}% (fuzzy: {scores[0]}, word: {scores[1]}, compound: {scores[2]}){penalty_info}")
        
        return final_score
    
    def calculate_rarity_confidence(self, input_rarity, card):
        """Calculate confidence score for rarity matching"""
        card_sets = card.get('card_sets', [])
        best_rarity_score = 0
        
        input_rarity_clean = input_rarity.lower().strip()
        
        for card_set in card_sets:
            set_rarity = card_set.get('set_rarity', '').lower().strip()
            
            if set_rarity == input_rarity_clean:
                # Exact match gets highest score
                best_rarity_score = 100
                break
            elif input_rarity_clean in set_rarity or set_rarity in input_rarity_clean:
                # Partial match gets good score
                best_rarity_score = max(best_rarity_score, 80)
            else:
                # Use fuzzy matching for rarity names
                rarity_similarity = fuzz.ratio(input_rarity_clean, set_rarity)
                if rarity_similarity >= 70:
                    best_rarity_score = max(best_rarity_score, rarity_similarity * 0.7)  # Scale down fuzzy rarity matches
        
        return best_rarity_score
    
    def get_card_rarity_display(self, card, specified_rarity=None):
        """Get the best rarity display for a card"""
        if specified_rarity:
            return specified_rarity
        
        card_sets = card.get('card_sets', [])
        if card_sets:
            return card_sets[0].get('set_rarity', 'Unknown')
        
        return 'Unknown'
    
    def ensure_unique_confidence_scores(self, variants):
        """Ensure all variants have unique confidence scores to avoid ties"""
        if not variants:
            return
        
        # Track used confidence scores
        used_scores = set()
        
        for variant in variants:
            original_confidence = variant['confidence']
            confidence = original_confidence
            
            # If this confidence is already used, find a unique one
            while confidence in used_scores:
                confidence -= 0.1
                # Ensure we don't go below reasonable bounds
                if confidence < 10:
                    confidence = original_confidence + 0.1
                    while confidence in used_scores:
                        confidence += 0.1
                        if confidence > 99:
                            break
                    break
            
            # Round to one decimal place and update
            confidence = round(confidence, 1)
            variant['confidence'] = confidence
            used_scores.add(confidence)
            
            if confidence != original_confidence:
                print(f"[CARD SEARCH DEBUG] Adjusted confidence for uniqueness: {variant['name']} {original_confidence}% -> {confidence}%")
    
    def show_voice_card_options(self, card_matches, original_input, art_variant, rarity):
        """Show card options dialog for voice selection"""
        print(f"[DIALOG DEBUG] show_voice_card_options called with {len(card_matches)} matches")
        
        # Test: Try to create a simple dialog first to see if it's a dialog creation issue
        try:
            test_dialog = tk.Toplevel(self.root)
            test_dialog.title("Test Dialog")
            test_dialog.geometry("200x100")
            ttk.Label(test_dialog, text="Dialog creation test").pack()
            test_dialog.after(1000, test_dialog.destroy)  # Auto-close after 1 second
            print(f"[DIALOG DEBUG] Simple test dialog created successfully")
        except Exception as e:
            print(f"[DIALOG DEBUG] Failed to create test dialog: {e}")
        
        try:
            # Set pending state
            self.pending_voice_confirmation = True
            self.pending_card_options = card_matches
            self.pending_voice_data = {
                'original_input': original_input,
                'art_variant': art_variant,
                'rarity': rarity
            }
            
            # Validate we have options to show
            if not card_matches:
                print("[VOICE DEBUG] No card matches to display, cancelling selection")
                self.pending_voice_confirmation = False
                self.pending_card_options = []
                self.pending_voice_data = {}
                self.root.after(0, lambda: self.voice_status_var.set("No options to display"))
                return
            
            print(f"[VOICE DEBUG] Displaying {len(card_matches)} options in dialog")
            
            def select_option(index):
                try:
                    if 0 <= index < len(card_matches):
                        selected_option = card_matches[index]
                        
                        # Clear pending state BEFORE destroying dialog
                        self.pending_voice_confirmation = False
                        self.pending_card_options = []
                        
                        # Safely destroy dialog
                        try:
                            dialog.destroy()
                        except:
                            pass  # Dialog might already be destroyed
                        
                        # Add the selected card using the SPECIFIC rarity from the selected option
                        self.add_card_to_session(
                            selected_option['card_data'], 
                            art_variant, 
                            selected_option['rarity']  # Use the specific rarity from the selected option
                        )
                        self.pending_voice_data = {}
                        print(f"[VOICE DEBUG] Selected option {index + 1}: {selected_option['name']} - {selected_option['rarity']}")
                    else:
                        print(f"[VOICE DEBUG] Invalid option index: {index}")
                except Exception as e:
                    print(f"[VOICE DEBUG] Error selecting option: {e}")
                    # Ensure cleanup even on error
                    self.pending_voice_confirmation = False
                    self.pending_card_options = []
                    self.pending_voice_data = {}
                    self.root.after(0, lambda: self.voice_status_var.set("Error selecting option"))
            
            def cancel_selection():
                try:
                    # Clear pending state BEFORE destroying dialog
                    self.pending_voice_confirmation = False
                    self.pending_card_options = []
                    self.pending_voice_data = {}
                    
                    # Safely destroy dialog
                    try:
                        dialog.destroy()
                    except:
                        pass  # Dialog might already be destroyed
                        
                    self.root.after(0, lambda: self.voice_status_var.set("Selection cancelled, continue speaking..."))
                    print("[VOICE DEBUG] Selection cancelled by user")
                except Exception as e:
                    print(f"[VOICE DEBUG] Error cancelling selection: {e}")
                    # Ensure cleanup even on error
                    self.pending_voice_confirmation = False
                    self.pending_card_options = []
                    self.pending_voice_data = {}
            
            # Create options dialog
            dialog = tk.Toplevel(self.root)
            dialog.title("Select Card Match")
            dialog.geometry("600x500")
            dialog.transient(self.root)
            dialog.grab_set()
            
            # Make X button behave like Cancel button
            dialog.protocol("WM_DELETE_WINDOW", cancel_selection)
            
            # Ensure dialog is brought to front and properly focused
            dialog.lift()
            dialog.focus_force()
            dialog.update()  # Force update to ensure proper display
            
            # Center the dialog with better error handling
            try:
                dialog.update_idletasks()
                dialog_width = dialog.winfo_width()
                dialog_height = dialog.winfo_height()
                screen_width = dialog.winfo_screenwidth()
                screen_height = dialog.winfo_screenheight()
                
                x = max(0, (screen_width // 2) - (dialog_width // 2))
                y = max(0, (screen_height // 2) - (dialog_height // 2))
                dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
                
                # Ensure dialog stays on top and is visible
                dialog.attributes('-topmost', True)
                dialog.update()
                dialog.attributes('-topmost', False)  # Allow normal window behavior after showing
                
                print(f"[VOICE DEBUG] Dialog positioned at {x},{y} with size {dialog_width}x{dialog_height}")
            except Exception as e:
                print(f"[VOICE DEBUG] Error positioning dialog: {e}")
                # Fallback positioning
                dialog.geometry("600x500+100+100")
            
            # Dialog content
            ttk.Label(dialog, text="Card Match Options", font=("Arial", 16, "bold")).pack(pady=10)
            ttk.Label(dialog, text=f"You said: \"{original_input}\"", font=("Arial", 12)).pack(pady=5)
            
            if art_variant:
                ttk.Label(dialog, text=f"Art Variant: {art_variant}", font=("Arial", 10)).pack(pady=2)
            if rarity:
                ttk.Label(dialog, text=f"Specified Rarity: {rarity}", font=("Arial", 10)).pack(pady=2)
            
            ttk.Label(dialog, text="Select an option by clicking or saying the number:", 
                     font=("Arial", 11, "bold")).pack(pady=(15, 10))
            
            # Create scrollable frame for options
            canvas = tk.Canvas(dialog, height=300)
            scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Add options
            for i, match in enumerate(card_matches, 1):
                option_frame = ttk.LabelFrame(scrollable_frame, text=f"Option {i}", padding="10")
                option_frame.pack(fill=tk.X, padx=10, pady=5)
                
                # Card name and confidence
                name_label = ttk.Label(option_frame, text=match['name'], font=("Arial", 12, "bold"))
                name_label.pack(anchor=tk.W)
                
                confidence_label = ttk.Label(option_frame, text=f"Confidence: {match['confidence']:.1f}%", 
                                            font=("Arial", 10), foreground="#0066CC")
                confidence_label.pack(anchor=tk.W)
                
                # Card rarity
                rarity_label = ttk.Label(option_frame, text=f"Rarity: {match['rarity']}", 
                                       font=("Arial", 10))
                rarity_label.pack(anchor=tk.W)
                
                # Set code if available
                set_code = match.get('set_code', 'N/A')
                set_label = ttk.Label(option_frame, text=f"Set Code: {set_code}", 
                                    font=("Arial", 9), foreground="#666666")
                set_label.pack(anchor=tk.W)
                
                # Select button
                select_btn = ttk.Button(option_frame, text=f"Select Option {i}", 
                                      command=lambda idx=i-1: select_option(idx))
                select_btn.pack(anchor=tk.W, pady=(5, 0))
            
            canvas.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
            scrollbar.pack(side="right", fill="y", pady=10, padx=(0, 10))
            
            # Control buttons
            button_frame = ttk.Frame(dialog)
            button_frame.pack(pady=10)
            
            ttk.Button(button_frame, text="❌ Cancel", command=cancel_selection).pack(side=tk.LEFT, padx=5)
            
            # Voice instruction
            voice_instruction = ttk.Label(dialog, 
                                        text="💡 Say the option number (e.g., '1', '2') or 'cancel' to reject all options",
                                        font=("Arial", 9, "italic"), foreground="#666666")
            voice_instruction.pack(pady=(0, 10))
            
            # Start dedicated voice listener for dialog selections
            def dialog_voice_listener():
                """Dedicated voice listener for dialog selections to avoid context manager conflicts"""
                listen_count = 0
                max_listen_attempts = 20  # Limit attempts to prevent infinite loops
                
                while self.pending_voice_confirmation and listen_count < max_listen_attempts:
                    try:
                        # Use a shorter timeout for dialog listening
                        voice_text = self.voice_recognizer.listen_once(timeout=2)
                        if voice_text:
                            print(f"[DIALOG VOICE DEBUG] Dialog heard: '{voice_text}'")
                            # Process the voice selection in the main thread
                            self.root.after(0, lambda text=voice_text: self.handle_voice_selection(text))
                            break  # Exit after processing one command
                        listen_count += 1
                        time.sleep(0.1)  # Brief pause between attempts
                    except Exception as e:
                        print(f"[DIALOG VOICE DEBUG] Dialog voice error: {e}")
                        listen_count += 1
                        time.sleep(0.5)
                
                print(f"[DIALOG VOICE DEBUG] Dialog voice listener stopped after {listen_count} attempts")
            
            # Start dialog voice listener in separate thread
            dialog_voice_thread = threading.Thread(target=dialog_voice_listener, daemon=True)
            dialog_voice_thread.start()
            
            # Update voice status
            self.root.after(0, lambda: self.voice_status_var.set(f"Waiting for selection: Say 1-{len(card_matches)} or 'cancel'"))
            
            print(f"[VOICE DEBUG] Dialog displayed successfully with {len(card_matches)} options")
            
        except Exception as e:
            print(f"[VOICE DEBUG] Error creating voice options dialog: {e}")
            # Fallback: clear pending state and show error
            self.pending_voice_confirmation = False
            self.pending_card_options = []
            self.pending_voice_data = {}
            self.root.after(0, lambda: self.voice_status_var.set(f"Error displaying options: {str(e)}"))
    
    def add_card_to_session(self, card_data, art_variant=None, rarity=None):
        """Add a card to the current session with immediate placeholder, then fetch pricing asynchronously"""
        print(f"[ADD CARD DEBUG] Adding card: {card_data.get('name', 'Unknown')} - {rarity}")
        
        # Use the provided rarity parameter as the definitive rarity (this comes from the selected option)
        final_rarity = rarity or self.get_card_rarity_display(card_data, rarity)
        
        # Immediately add card with loading placeholder
        loading_card = {
            **card_data,
            'card_name': card_data.get('name', 'Unknown Card'),
            'card_rarity': final_rarity,  # Use the exact rarity from selection
            'art_variant': art_variant or 'None',
            'tcg_price': '⏳ Loading...',
            'tcg_market_price': '⏳ Loading...',
            'price_status': 'loading',
            'quantity': 1,  # Default quantity
            'timestamp': datetime.now().isoformat()
        }
        
        # Add to session immediately
        self.pack_session.add_card(loading_card)
        print(f"[ADD CARD DEBUG] Session now has {len(self.pack_session.cards)} cards")
        
        # Pre-load images for both modes to improve mode switching performance
        self.preload_card_images_background(card_data)
        
        # Update UI immediately with loading placeholder
        print(f"[ADD CARD DEBUG] About to call update_session_display() from main thread")
        print(f"[ADD CARD DEBUG] Current thread: {threading.current_thread().name}")
        print(f"[ADD CARD DEBUG] Is main thread: {threading.current_thread() == threading.main_thread()}")
        print(f"[ADD CARD DEBUG] Session tracker exists: {self.session_tracker is not None}")
        print(f"[ADD CARD DEBUG] Pack session active: {self.pack_session_active}")
        
        if not self.session_tracker:
            print(f"[ADD CARD DEBUG] ERROR: No session tracker available! Cannot update UI.")
            return
            
        if not self.pack_session_active:
            print(f"[ADD CARD DEBUG] WARNING: Pack session not active!")
            
        self.update_session_display()
        
        # CRITICAL FIX: Add fallback update mechanism with delay
        def delayed_update_fallback():
            print(f"[ADD CARD DEBUG] Executing delayed fallback update")
            if self.session_tracker:
                try:
                    self.session_tracker.safe_update_cards_display()
                    print(f"[ADD CARD DEBUG] Fallback update completed")
                except Exception as e:
                    print(f"[ADD CARD DEBUG] Fallback update failed: {e}")
            else:
                print(f"[ADD CARD DEBUG] No session tracker for fallback update")
        
        # Schedule a fallback update after 500ms in case the immediate one fails
        self.root.after(500, delayed_update_fallback)
        
        # Lock status if not already locked (auto-confirm already locks it)
        if not self.voice_status_locked:
            self.voice_status_locked = True
            
        self.root.after(0, lambda: self.voice_status_var.set(f"Added: {card_data.get('name', 'Card')} - {final_rarity} (fetching price...)"))
        
        # Clear the status message after 2 seconds and return to listening mode
        def clear_status_and_resume():
            print("[VOICE DEBUG] Clearing 'Added:' status and unlocking voice status")
            if self.voice_listening:
                self.voice_status_var.set("🎤 Listening...")
            else:
                self.voice_status_var.set("Voice recognition ready")
            # Unlock status to allow voice loop to continue
            self.voice_status_locked = False
            print("[VOICE DEBUG] Voice status unlocked - ready for new input")
        
        # Use after_idle for higher priority to ensure voice status clearing is not blocked by other UI updates
        self.root.after(2000, lambda: self.root.after_idle(clear_status_and_resume))  # Clear after 2 seconds with high priority
        
        # Fetch price in background without blocking
        def fetch_price_and_update():
            try:
                # Prepare payload for price API using the EXACT selected rarity
                payload = {
                    "card_number": card_data.get('card_sets', [{}])[0].get('set_code', ''),
                    "card_name": card_data.get('name', ''),
                    "card_rarity": final_rarity,  # Use the exact selected rarity
                    "art_variant": art_variant or '',
                    "force_refresh": False
                }
                
                print(f"[PRICE FETCH DEBUG] Fetching price for: {payload['card_name']} - {payload['card_rarity']}")
                
                # Make API call to get pricing
                url = f"{self.api_url}/cards/price"
                response = requests.post(url, json=payload, timeout=30)
                
                # Log the API call
                log_api_call("POST", url, body=payload, response=response)
                
                response.raise_for_status()
                
                price_data = response.json()
                
                # Find the card in our session and update it (match by name AND exact rarity)
                for i, session_card in enumerate(self.pack_session.cards):
                    if (session_card.get('name') == card_data.get('name') and 
                        session_card.get('price_status') == 'loading' and
                        session_card.get('card_rarity') == final_rarity):  # Match exact rarity
                        
                        if price_data.get('success'):
                            # Update with actual pricing data, preserving the selected rarity
                            updated_card = {
                                **session_card,
                                **price_data.get('data', {}),
                                'card_rarity': final_rarity,  # Ensure rarity doesn't get overwritten
                                'price_status': 'loaded'
                            }
                            print(f"[PRICE FETCH DEBUG] Price loaded for {card_data.get('name')} - {final_rarity}: Low=${price_data.get('data', {}).get('tcg_price', 'N/A')}, Market=${price_data.get('data', {}).get('tcg_market_price', 'N/A')}")
                        else:
                            # Update with failed status, preserving the selected rarity
                            updated_card = {
                                **session_card,
                                'tcg_price': 'Price unavailable',
                                'tcg_market_price': 'Price unavailable',
                                'card_rarity': final_rarity,  # Ensure rarity doesn't get overwritten
                                'price_status': 'failed'
                            }
                            print(f"[PRICE FETCH DEBUG] Price fetch failed for {card_data.get('name')} - {final_rarity}")
                        
                        # Replace the card in the session
                        self.pack_session.cards[i] = updated_card
                        
                        # Update UI on main thread - use immediate scheduling to ensure execution
                        print(f"[PRICE FETCH DEBUG] Scheduling UI update for {card_data.get('name')} - {final_rarity}")
                        self.root.after(0, self.update_session_display)  # Use after(0) to ensure execution
                        break
                
            except Exception as e:
                print(f"[PRICE FETCH DEBUG] Error fetching price for {card_data.get('name')} - {final_rarity}: {e}")
                # Find and update the card with error status, preserving the selected rarity
                for i, session_card in enumerate(self.pack_session.cards):
                    if (session_card.get('name') == card_data.get('name') and 
                        session_card.get('price_status') == 'loading' and
                        session_card.get('card_rarity') == final_rarity):  # Match exact rarity
                        
                        # Update with error status, preserving the selected rarity
                        updated_card = {
                            **session_card,
                            'tcg_price': '❌ Error',
                            'tcg_market_price': '❌ Error',
                            'card_rarity': final_rarity,  # Ensure rarity doesn't get overwritten
                            'price_status': 'error'
                        }
                        
                        self.pack_session.cards[i] = updated_card
                        self.root.after(0, self.update_session_display)  # Use after(0) to ensure execution
                        break
        
        # Start price fetching in background thread
        thread = threading.Thread(target=fetch_price_and_update)
        thread.daemon = True
        thread.start()
    
    def preload_card_images_background(self, card_data):
        """Pre-load card images in both modes in background thread for better performance"""
        def preload_images():
            try:
                if not PIL_AVAILABLE or not hasattr(self, 'image_manager'):
                    return
                    
                card_images = card_data.get('card_images', [])
                if not card_images or len(card_images) == 0:
                    return
                
                image_url = card_images[0].get('image_url')
                card_id = card_data.get('id')
                
                if image_url and card_id:
                    print(f"[IMAGE PRELOAD] Pre-loading images for card {card_data.get('name')} (ID: {card_id})")
                    self.image_manager.preload_image_for_both_modes(card_id, image_url)
                    print(f"[IMAGE PRELOAD] Completed pre-loading for card {card_data.get('name')}")
                    
            except Exception as e:
                print(f"[IMAGE PRELOAD] Error pre-loading images: {e}")
        
        # Run in background thread to avoid blocking UI
        thread = threading.Thread(target=preload_images)
        thread.daemon = True
        thread.start()
    
    def save_session_manual(self):
        """Manually save the current session"""
        if self.pack_session.save_session():
            messagebox.showinfo("Session Saved", f"Session saved successfully!\nFile: {self.pack_session.session_file}")
        else:
            messagebox.showerror("Save Error", "Failed to save session.")
    
    def load_session_manual(self):
        """Manually load a saved session"""
        if self.pack_session.load_session():
            self.update_session_display()
            messagebox.showinfo("Session Loaded", f"Session loaded successfully!\nCards: {len(self.pack_session.cards)}")
        else:
            messagebox.showerror("Load Error", "Failed to load session.")
    
    def export_to_excel(self):
        """Export session cards to Excel file"""
        if not self.pack_session.cards:
            messagebox.showwarning("No Data", "No cards in current session to export!")
            return
        
        # Show field selection dialog
        self.show_export_dialog()
    
    def show_export_dialog(self):
        """Show dialog for selecting fields to export"""
        def export_selected():
            selected_fields = []
            for field, var in field_vars.items():
                if var.get():
                    selected_fields.append(field)
            
            if not selected_fields:
                messagebox.showwarning("No Fields", "Please select at least one field to export.")
                return
            
            dialog.destroy()
            self.perform_excel_export(selected_fields)
        
        # Create export dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Export Fields")
        dialog.geometry("300x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Select fields to export:", font=("Arial", 12, "bold")).pack(pady=10)
        
        # Available fields with more comprehensive options from v2 API
        available_fields = {
            'card_name': 'Card Name',
            'card_rarity': 'Card Rarity',
            'art_variant': 'Art Variant',
            'tcg_price': 'TCGPlayer Low Price',
            'tcg_market_price': 'TCGPlayer Market Price',
            'set_code': 'Set Code',
            'booster_set_name': 'Set Name',
            'card_number': 'Card Number',
            'card_art_variant': 'Card Art Variant',
            'scrape_success': 'Scrape Success',
            'source_url': 'Source URL',
            'last_price_updt': 'Last Updated',
            'timestamp': 'Added Timestamp',
            'error_message': 'Error Message (if any)'
        }
        
        field_vars = {}
        for field, label in available_fields.items():
            var = tk.BooleanVar(value=True)  # Default to selected
            field_vars[field] = var
            ttk.Checkbutton(dialog, text=label, variable=var).pack(anchor=tk.W, padx=20, pady=2)
        
        # Buttons
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="📊 Export", command=export_selected).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)
    
    def perform_excel_export(self, selected_fields):
        """Perform the actual Excel export"""
        # Ask for file save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Save Excel Export"
        )
        
        if not file_path:
            return
        
        try:
            # Create workbook
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Pack Session Cards"
            
            # Field labels with expanded v2 API fields
            field_labels = {
                'card_name': 'Card Name',
                'card_rarity': 'Card Rarity',
                'art_variant': 'Art Variant',
                'tcg_price': 'TCGPlayer Low Price',
                'tcg_market_price': 'TCGPlayer Market Price',
                'set_code': 'Set Code',
                'booster_set_name': 'Set Name',
                'card_number': 'Card Number',
                'card_art_variant': 'Card Art Variant',
                'scrape_success': 'Scrape Success',
                'source_url': 'Source URL',
                'last_price_updt': 'Last Updated',
                'timestamp': 'Added Timestamp',
                'error_message': 'Error Message (if any)'
            }
            
            # Write headers
            headers = [field_labels.get(field, field) for field in selected_fields]
            worksheet.append(headers)
            
            # Style headers
            for cell in worksheet["1:1"]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data
            for card in self.pack_session.cards:
                row_data = []
                for field in selected_fields:
                    value = card.get(field, 'N/A')
                    if value is None:
                        value = 'N/A'
                    row_data.append(value)
                worksheet.append(row_data)
            
            # Auto-size columns
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(file_path)
            messagebox.showinfo("Export Success", f"Excel file created successfully!\n\nFile: {file_path}\nCards exported: {len(self.pack_session.cards)}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to create Excel file:\n{str(e)}")
    
    def show_cache_info(self):
        """Show image cache information and management"""
        cache_size = self.image_manager.get_cache_size()
        dialog = tk.Toplevel(self.root)
        dialog.title("Image Cache Information")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        info_text = f"""Image Cache Status:
        
Cache Size: {cache_size:.2f} MB
Cache Location: {self.image_manager.image_cache_dir}

The image cache stores downloaded card images locally to comply with YGO API guidelines and improve performance.
        """
        
        ttk.Label(dialog, text=info_text, justify=tk.LEFT, padding="15").pack(expand=True, fill=tk.BOTH)
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        
        def clear_cache():
            if messagebox.askyesno("Clear Cache", "Are you sure you want to clear the image cache?"):
                self.image_manager.clear_cache()
                messagebox.showinfo("Cache Cleared", "Image cache has been cleared successfully!")
                dialog.destroy()
        
        ttk.Button(button_frame, text="🗑️ Clear Cache", command=clear_cache).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="✅ Close", command=dialog.destroy).pack(side=tk.LEFT, padx=10)
    
    def show_display_settings(self):
        """Show display customization settings"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Display Settings")
        dialog.geometry("500x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Customize Session Tracker Display", font=("Arial", 14, "bold")).pack(pady=10)
        
        # Available display fields
        display_fields = {
            'show_images': 'Show Card Images',
            'show_card_name': 'Show Card Name',
            'show_rarity': 'Show Rarity',
            'show_art_variant': 'Show Art Variant',
            'show_tcg_price': 'Show TCGPlayer Low Price',
            'show_tcg_market_price': 'Show TCGPlayer Market Price',
            'show_set_info': 'Show Set Information',
            'show_timestamps': 'Show Add Timestamp'
        }
        
        # Initialize settings if not exists
        if not hasattr(self, 'display_settings'):
            self.display_settings = {key: True for key in display_fields.keys()}
        
        field_vars = {}
        for field, label in display_fields.items():
            var = tk.BooleanVar(value=self.display_settings.get(field, True))
            field_vars[field] = var
            ttk.Checkbutton(dialog, text=label, variable=var).pack(anchor=tk.W, padx=20, pady=2)
        
        def apply_settings():
            # Update settings
            for field, var in field_vars.items():
                self.display_settings[field] = var.get()
            
            # Refresh display
            self.update_session_display()
            dialog.destroy()
            messagebox.showinfo("Settings Applied", "Display settings have been updated!")
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="✅ Apply", command=apply_settings).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="❌ Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def auto_load_card_sets(self):
        """Automatically load card sets from cache on startup"""
        def load_sets_thread():
            try:
                # Use cache endpoint to get all sets
                url = f"{self.api_url}/card-sets/from-cache"
                response = requests.get(url, timeout=30)
                
                # Log the API call
                log_api_call("GET", url, response=response)
                
                response.raise_for_status()
                
                data = response.json()
                if data.get('success'):
                    sets_data = data.get('data', [])
                    self.available_sets = sets_data
                    self.filtered_sets = sets_data
                    
                    # Update UI on main thread
                    self.root.after(0, self.update_set_combobox)
                    print(f"Auto-loaded {len(sets_data)} card sets from cache")
                else:
                    print(f"Failed to auto-load sets: {data.get('message', 'Unknown error')}")
            except requests.exceptions.RequestException as e:
                # Log the error
                log_api_call("GET", url, error=e)
                print(f"Failed to auto-load sets: {str(e)}")
        
        # Run in separate thread to avoid blocking UI startup
        thread = threading.Thread(target=load_sets_thread)
        thread.daemon = True
        thread.start()

def main():
    root = tk.Tk()
    app = YGORipperUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()